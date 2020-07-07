#!/usr/bin/env python
# coding: utf-8

import os
import sys
from PyQt5.QtWidgets import *
import cv2
import numpy as np
import glob
import pytesseract
import openpyxl
from threading import Thread
from queue import Queue
from concurrent.futures import ThreadPoolExecutor
from time import sleep

pytesseract.pytesseract.tesseract_cmd = 'C:/Tesseract-OCR/tesseract.exe'
# pytesseract.pytesseract.tesseract_cmd = 'C:/Tesseract-OCR-4.0/tesseract.exe'
pool = ThreadPoolExecutor(8)
pool2 = ThreadPoolExecutor(3)
options = []

print('header loading start')
header_type01_type02 = cv2.imread('headers/type01_type02.jpg', cv2.IMREAD_COLOR)
header_type03 = cv2.imread('headers/type03.jpg', cv2.IMREAD_COLOR)
header_type04 = cv2.imread('headers/type04.jpg', cv2.IMREAD_COLOR)
header_type05 = cv2.imread('headers/type05.jpg', cv2.IMREAD_COLOR)
header_type06 = cv2.imread('headers/type06.jpg', cv2.IMREAD_COLOR)
header_type07_1 = cv2.imread('headers/type07_1.jpg', cv2.IMREAD_COLOR)
header_type07_2 = cv2.imread('headers/type07_2.jpg', cv2.IMREAD_COLOR)
header_type08 = cv2.imread('headers/type08.jpg', cv2.IMREAD_COLOR)
print('header loading end')
print('\nplease wait for initializing...')

before_image_files_counter = 0
type01_img_cnt = 0
type02_img_cnt = 0
type03_img_cnt = 0
type04_img_cnt = 0
type05_img_cnt = 0
type06_img_cnt = 0
type07_img_cnt = 0
type08_img_cnt = 0
typeUK_img_cnt = 0


def detect(image, template, ratio):
    resized = cv2.resize(image, dsize=(0, 0), fx=ratio, fy=ratio, interpolation=cv2.INTER_CUBIC)
    res = cv2.matchTemplate(resized, template, cv2.TM_CCOEFF)
    (_, max_val, _, max_loc) = cv2.minMaxLoc(res)
    return [max_val, max_loc, resized]


def get_max_matched_res(image, template):
    fs = []
    ratio = 0.980
    for i in range(1, 6 + 1):
        fs.append(pool2.submit(detect, image, template, ratio))
        ratio += 0.0005

    max_template_match_val  = -1
    max_template_match_loc = -1
    max_template_match_img = -1
    for f in fs:
        res = f.result()
        if res[0] > max_template_match_val:
            max_template_match_val = res[0]
            max_template_match_loc = res[1]
            max_template_match_img = res[2]
    
    return [max_template_match_img, max_template_match_loc]


def process_and_get_arr(image, file_name, header, table_w, table_h):
    res = get_max_matched_res(image, header)
    img = res[0]
    loc = res[1]
    header_height, header_width, header_channel = header.shape
    table_x = loc[0]
    table_y = loc[1] + header_height
    table = img[table_y:table_y+table_h, table_x:table_x+table_w]
    zoomed_table = cv2.resize(table, dsize=(0, 0), fx=2, fy=2, interpolation=cv2.INTER_CUBIC)

    # ocr
    ocr_res = pytesseract.image_to_string(zoomed_table, config='-psm 6 digits') # ver 3
    # ocr_res = pytesseract.image_to_string(zoomed_table, config='--psm 6 --oem 0') # ver 4
    # ocr_res = pytesseract.image_to_string(zoomed_table, config='--psm 6 tessedit_char_whitelist=0123456789.-')
    # ocr_res = pytesseract.image_to_string(zoomed_table, config='--psm 6 tessedit_char_whitelist=0123456789.-' + white_list)

    # split ocr result by line & space
    sp = ocr_res.split('\n')
    arr = []
    arr_1d = []
    for s in sp:
        if len(s.strip()) != 0:
            sp = s.split(' ')
            line_arr = []
            for ns in sp:
                if ns.startswith('.'):
                    ns = ns.replace('.', '-', 1)
                line_arr.append(ns)
                arr_1d.append(ns)
            arr.append(line_arr)


    file_res = []
    # detect x, y position using contour
    proc = cv2.resize(table, dsize=(0, 0), fx=0.5, fy=0.5, interpolation=cv2.INTER_CUBIC)
    proc = cv2.cvtColor(proc, cv2.COLOR_BGR2GRAY)
    th, proc =  cv2.threshold(proc, 190, 255, cv2.THRESH_BINARY)
    proc = cv2.bitwise_not(proc)
    kernel = np.ones((1, 5), np.uint8)
    proc = cv2.dilate(proc, kernel, iterations=1)
    contours, hierachy = cv2.findContours(proc, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    table_copy = table.copy()
    table_h, table_w, table_ch = table_copy.shape
    i = 0
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        x *= 2
        y *= 2
        w *= 2
        h *= 2

        if w < table_w - 10:
            cv2.rectangle(table_copy, (x, y), (x + w, y + h), (0, 0, 255), 2)
            x += table_x
            y += table_y
            cv2.rectangle(image, (x, y), (x + w, y + h), (0, 0, 255), 2)
            json = {
                'val': len(arr_1d) > i and arr_1d[i] or '',
                'x': str(x),
                'y': str(y),
                'w': str(w),
                'h': str(h),
            }
            file_res.append(json)
            i += 1
            # print('' + str(x) + ', ' + str(y) + ', ' + str(w) + ', ' + str(h))
    
    cv2.imshow('proc', proc)
    cv2.imshow('table_copy', table_copy)
    cv2.imshow('image', image)
    cv2.waitKey(0)
            
    # save result to file
    if os.path.exists ('result') == 0:
        os.mkdir('result')
    sp = file_name.split('.')
    file = open('result/' + sp[0] + '.txt', mode='wt', encoding='utf-8')
    # for line in arr:
    #     for ns in line:
    #         file.write(ns + ' ')
    #     file.write('\n')
    file.write(str(file_res).replace('\'', '\"'))
    file.close()

    # pre-process image if options exist
    if os.path.exists ('progress') == 0:
        os.mkdir('progress')
    sp = file_name.split('.')
    sp[0] = 'progress/' + sp[0]
    i = 0
    for option in options:
        if option == 'grayscale':
            tmp_gray = cv2.cvtColor(table, cv2.COLOR_BGR2GRAY)
            cv2.imwrite('%s_%s_grayscale.%s' % (sp[0], i, sp[1]), tmp_gray)
        elif option == 'threshold':
            tmp, tmp_binary = cv2.threshold(table, 190, 255, cv2.THRESH_BINARY)
            cv2.imwrite('%s_%s_threshold.%s' % (sp[0], i, sp[1]), tmp_binary)
        elif option == 'blur':
            tmp_blur = cv2.blur(table, (2, 2))
            cv2.imwrite('%s_%s_blur.%s' % (sp[0], i, sp[1]), tmp_blur)
        elif option == 'invert':
            tmp_invert = cv2.bitwise_not(table)
            cv2.imwrite('%s_%s_invert.%s' % (sp[0], i, sp[1]), tmp_invert)
            print('invert')
        i += 1
        
    return arr


def ocr_cell(cell):
    cell = cv2.resize(cell, dsize=(0, 0), fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
    return pytesseract.image_to_string(cell, config='-psm 6 digits')


def ocr(q, var_name, submat):
    submat = cv2.resize(submat, dsize=(0, 0), fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
    res = pytesseract.image_to_string(submat, config='-psm 6 digits')
    q.put([var_name, res])
    return res


def ocr_for_title_searching(submat):
    submat = cv2.resize(submat, dsize=(0, 0), fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
    submat = cv2.copyMakeBorder(submat, 20, 20, 20, 20, cv2.BORDER_CONSTANT, value=[255, 255, 255])
    res = pytesseract.image_to_string(submat, config='-psm 6')
    return res


def process_chart_and_get_v_gar(cur_before_image_file_path):
    sleep(0)
    global before_image_files_counter
    global type01_img_cnt
    global type02_img_cnt
    global type03_img_cnt
    global type04_img_cnt
    global type05_img_cnt
    global type06_img_cnt
    global type07_img_cnt
    global type08_img_cnt
    global typeUK_img_cnt

    g_var = new_g_var()
    cur_before_image_file_name = os.path.basename(cur_before_image_file_path)
    chart_image = cv2.imread(cur_before_image_file_path)

    q = Queue()
    fs = []
    arr = []
    
    #[s_y:e_y, s_x:e_x]
    if ocr_for_title_searching(chart_image[167:223, 239:445]) == 'Methacholine':
        g_var['img_type'] = 'type01'
        arr = process_and_get_arr(chart_image, cur_before_image_file_name, header_type01_type02, 790, 360)
        
        # fvc dose
        for i in range(0, 5 + 1):
            g_var['img_fvc_dose_lv' + str(i + 1)] = arr[0][i]

        # fvc liters
        g_var['img_fvc_ref'] = arr[1][0]
        g_var['img_fvc_pre'] = arr[1][1]
        for i in range(2, 7 + 1):
            g_var['img_fvc_lv' + str(i - 1)] = arr[1][i]

        # fvc % ref
        g_var['img_fvc_pref_pre'] = arr[2][0]
        for i in range(1, 6 + 1):
            g_var['img_fvc_pref_lv' + str(i)] = arr[2][i]

        # fvc % chg
        for i in range(0, 5 + 1):
            g_var['img_fvc_pchg_lv' + str(i + 1)] = arr[3][i]

        # fev1 dose
        for i in range(0, 5 + 1):
            g_var['img_fev1_dose_lv' + str(i + 1)] = arr[4][i]

        # fev1 liters
        g_var['img_fev1_ref'] = arr[5][0]
        g_var['img_fev1_pre'] = arr[5][1]
        for i in range(2, 7 + 1):
            g_var['img_fev1_lv' + str(i - 1)] = arr[5][i]

        # fev1 % ref
        g_var['img_fev1_pref_pre'] = arr[6][0]
        for i in range(1, 6 + 1):
            g_var['img_fev1_pref_lv' + str(i)] = arr[6][i]

        # fev1 % chg
        for i in range(0, 5 + 1):
            g_var['img_fev1_pchg_lv' + str(i + 1)] = arr[7][i]

        # fef25-75% dose
        for i in range(0, 5 + 1):
            g_var['img_fef25_75_lv' + str(i + 1)] = arr[8][i]

        # fef25-75%
        for i in range(0, 5 + 1):
            g_var['img_fef25_75_dose_lv' + str(i + 1)] = arr[9][i]

        # fef25-75% % ref
        g_var['img_fef25_75_pre'] = arr[10][0]
        for i in range(1, 6 + 1):
            g_var['img_fef25_75_lv' + str(i)] = arr[10][i]

        # fef25-75% % chg
        for i in range(0, 5 + 1):
            g_var['img_fef25_75_pchg_lv' + str(i + 1)] = arr[11][i]
        
        # pef lsec dose
        for i in range(0, 5 + 1):
            g_var['img_pef_dose_lv' + str(i + 1)] = arr[12][i]

        # pef lsec
        g_var['img_pef_ref'] = arr[13][0]
        g_var['img_pef_pre'] = arr[13][1]
        for i in range(2, 7 + 1):
            g_var['img_pef_lv' + str(i - 1)] = arr[13][i]

        # pef lsec % ref
        g_var['img_pef_pref_pre'] = arr[14][0]
        for i in range(1, 6 + 1):
            g_var['img_pef_pref_lv' +  str(i)] = arr[14][i]

        # pef lsec % chg
        for i in range(0, 5 + 1):
            g_var['img_pef_pchg_lv' + str(i + 1)] = arr[15][i]

        type01_img_cnt += 1
        pass

    elif ocr_for_title_searching(chart_image[177:217, 156:247]) == 'aridol':
        g_var['img_type'] = 'type02'
        arr = process_and_get_arr(chart_image, cur_before_image_file_name, header_type01_type02, 790, 360)

        # fvc dose
        for i in range(0, 8 + 1):
            g_var['img_fvc_dose_lv' + str(i + 1)] = arr[0][i]

        # fvc liters
        g_var['img_fvc_ref'] = arr[1][0]
        g_var['img_fvc_pre'] = arr[1][1]
        for i in range(2, 10 + 1):
            g_var['img_fvc_lv' + str(i - 1)] = arr[1][i]

        # fvc % ref
        g_var['img_fvc_pref_pre'] = arr[2][0]
        for i in range(1, 9 + 1):
            g_var['img_fvc_pref_lv' + str(i)] = arr[2][i]

        # fvc % chg
        for i in range(0, 8 + 1):
            g_var['img_fvc_pchg_lv' + str(i + 1)] = arr[3][i]

        # fev1 dose
        for i in range(0, 8 + 1):
            g_var['img_fev1_dose_lv' + str(i + 1)] = arr[4][i]

        # fev1 liters
        g_var['img_fev1_ref'] = arr[5][0]
        g_var['img_fev1_pre'] = arr[5][1]
        for i in range(2, 10 + 1):
            g_var['img_fev1_lv' + str(i - 1)] = arr[5][i]

        # fev1 % ref
        g_var['img_fev1_pref_pre'] = arr[6][0]
        for i in range(1, 9 + 1):
            g_var['img_fev1_pref_lv' + str(i)] = arr[6][i]

        # fev1 % chg
        for i in range(0, 8 + 1):
            g_var['img_fev1_pchg_lv' + str(i + 1)] = arr[7][i]

        # fef25-75% dose
        for i in range(0, 8 + 1):
            g_var['img_fef25_75_lv' + str(i + 1)] = arr[8][i]

        # fef25-75%
        for i in range(0, 10 + 1):
            g_var['img_fef25_75_dose_lv' + str(i + 1)] = arr[9][i]

        # fef25-75% % ref
        g_var['img_fef25_75_pre'] = arr[10][0]
        for i in range(1, 9 + 1):
            g_var['img_fef25_75_lv' + str(i)] = arr[10][i]

        # fef25-75% % chg
        for i in range(0, 8 + 1):
            g_var['img_fef25_75_pchg_lv' + str(i + 1)] = arr[11][i]
        
        # pef lsec dose
        for i in range(0, 8 + 1):
            g_var['img_pef_dose_lv' + str(i + 1)] = arr[12][i]

        # pef lsec
        g_var['img_pef_ref'] = arr[13][0]
        g_var['img_pef_pre'] = arr[13][1]
        for i in range(2, 10 + 1):
            g_var['img_pef_lv' + str(i - 1)] = arr[13][i]

        # pef lsec % ref
        g_var['img_pef_pref_pre'] = arr[14][0]
        for i in range(1, 9 + 1):
            g_var['img_pef_pref_lv' +  str(i)] = arr[14][i]

        # pef lsec % chg
        for i in range(0, 8 + 1):
            g_var['img_pef_pchg_lv' + str(i + 1)] = arr[15][i]

        type02_img_cnt += 1
        pass

    elif ocr_for_title_searching(chart_image[360:390, 38:111]) == 'Diffusing':
        g_var['img_type'] = 'type03'
        arr = process_and_get_arr(chart_image, cur_before_image_file_name, header_type03, 220, 140)

        g_var['img_dlco_ref'] = arr[0][0]
        g_var['img_dlco_pre'] = arr[0][1]
        g_var['img_dlco_pref_pre'] = arr[0][2]
        g_var['img_dladj_ref'] = arr[1][0]
        g_var['img_dladj_pre'] = arr[1][1]
        g_var['img_dladj_pref_pre'] = arr[1][2]
        g_var['img_dlcodva_ref'] = arr[2][0]
        g_var['img_dlcodva_pre'] = arr[2][1]
        g_var['img_dlcodva_pref_pre'] = arr[2][2]
        g_var['img_dldvaadj_ref'] = arr[3][0]
        g_var['img_dldvaadj_pre'] = arr[3][1]
        g_var['img_dldvaadj_pref_pre'] = arr[3][2]
        g_var['img_va_pre'] = arr[4][0]
        g_var['img_ivc_pre'] = arr[5][0]
        g_var['img_dlcoecode_pre'] = arr[6][0]

        type03_img_cnt += 1
        pass

    elif ocr_for_title_searching(chart_image[266:295, 1:89]) == 'Spirometry':
        g_var['img_type'] = 'type04'

        arr = process_and_get_arr(chart_image, cur_before_image_file_name, header_type04, 430, 900)

        # spirometry section start
        g_var['img_fvc_ref'] = arr[0][0]
        g_var['img_fvc_pre'] = arr[0][1]
        g_var['img_fvc_pref_pre'] = arr[0][2]
        if len(arr[0]) == 6:
            g_var['img_fvc_post'] = arr[0][3]
            g_var['img_fvc_pref_post'] = arr[0][4]
            g_var['img_fvc_pchg'] = arr[0][5]
        
        g_var['img_fev05_ref'] = arr[1][0]
        g_var['img_fev05_pre'] = arr[1][1]
        g_var['img_fev05_pref_pre'] = arr[1][2]
        if len(arr[1]) == 6:
            g_var['img_fev05_post'] = arr[1][3]
            g_var['img_fev05_pref_post'] = arr[1][4]
            g_var['img_fev05_pchg'] = arr[1][5]

        g_var['img_fev1_ref'] = arr[2][0]
        g_var['img_fev1_pre'] = arr[2][1]
        g_var['img_fev1_pref_pre'] = arr[2][2]
        if len(arr[2]) == 6:
            g_var['img_fev1_post'] = arr[2][3]
            g_var['img_fev1_pref_post'] = arr[2][4]
            g_var['img_fev1_pchg'] = arr[2][5]

        g_var['img_fev1dfvc_ref'] = arr[3][0]
        g_var['img_fev1dfvc_pre'] = arr[3][1]
        if len(arr[3]) == 3:
            g_var['img_fev1dfvc_post'] = arr[3][2]

        g_var['img_fev3dfvc_ref'] = arr[4][0]
        g_var['img_fev3dfvc_pre'] = arr[4][1]
        if len(arr[4]) == 3:
            g_var['img_fev3dfvc_post'] = arr[4][2]

        g_var['img_fef25_75_ref'] = arr[5][0]
        g_var['img_fef25_75_pre'] = arr[5][1]
        g_var['img_fef25_75_pref_pre'] = arr[5][2]
        if len(arr[5]) == 6:
            g_var['img_fef25_75_post'] = arr[5][3]
            g_var['img_fef25_75_pref_post'] = arr[5][4]
            g_var['img_fef25_75_pchg'] = arr[5][5]

        g_var['img_fef75_85_ref'] = arr[6][0]
        g_var['img_fef75_85_pre'] = arr[6][1]
        g_var['img_fef75_85_pref_pre'] = arr[6][2]
        if len(arr[6]) == 6:
            g_var['img_fef75_85_post'] = arr[6][3]
            g_var['img_fef75_85_pref_post'] = arr[6][4]
            g_var['img_fef75_85_pchg'] = arr[6][5]

        g_var['img_fef25_pre'] = arr[7][0]
        if len(arr[7]) == 3:
            g_var['img_fef25_post'] = arr[7][1]
            g_var['img_fef25_pchg'] = arr[7][2]

        g_var['img_fef50_ref'] = arr[8][0]
        g_var['img_fef50_pre'] = arr[8][1]
        g_var['img_fef50_pref_pre'] = arr[8][2]
        if len(arr[8]) == 6:
            g_var['img_fef50_post'] = arr[8][3]
            g_var['img_fef50_pref_post'] = arr[8][4]
            g_var['img_fef50_pchg'] = arr[8][5]

        g_var['img_fef75_ref'] = arr[9][0]
        g_var['img_fef75_pre'] = arr[9][1]
        g_var['img_fef75_pref_pre'] = arr[9][2]
        if len(arr[9]) == 6:
            g_var['img_fef75_post'] = arr[9][3]
            g_var['img_fef75_pref_post'] = arr[9][4]
            g_var['img_fef75_pchg'] = arr[9][5]

        g_var['img_fef200_1200_ref'] = arr[10][0]
        g_var['img_fef200_1200_pre'] = arr[10][1]
        g_var['img_fef200_1200_pref_pre'] = arr[10][2]
        if len(arr[10]) == 6:
            g_var['img_fef200_1200_post'] = arr[10][3]
            g_var['img_fef200_1200_pref_post'] = arr[10][4]
            g_var['img_fef200_1200_pchg'] = arr[10][5]

        g_var['img_pef_ref'] = arr[11][0]
        g_var['img_pef_pre'] = arr[11][1]
        g_var['img_pef_pref_pre'] = arr[11][2]
        if len(arr[11]) == 6:
            g_var['img_pef_post'] = arr[11][3]
            g_var['img_pef_pref_post'] = arr[11][4]
            g_var['img_pef_pchg'] = arr[11][5]

        g_var['img_fivc_ref'] = arr[12][0]
        g_var['img_fivc_pre'] = arr[12][1]
        g_var['img_fivc_pref_pre'] = arr[12][2]
        if len(arr[12]) == 6:
            g_var['img_fivc_post'] = arr[12][3]
            g_var['img_fivc_pref_post'] = arr[12][4]
            g_var['img_fivc_pchg'] = arr[12][5]

        g_var['img_fvlecode_pre'] = arr[13][0]
        if len(arr[13]) == 2:
            g_var['img_fvlecode_post'] = arr[13][1]
        # spirometry section end

        # lung volumes section start
        g_var['img_tlc_ref'] = arr[14][0]
        g_var['img_tlc_pre'] = arr[14][1]
        g_var['img_tlc_pref_pre'] = arr[14][2]

        g_var['img_vc_ref'] = arr[15][0]
        g_var['img_vc_pre'] = arr[15][1]
        g_var['img_vc_pref_pre'] = arr[15][2]

        g_var['img_ic_ref'] = arr[16][0]
        g_var['img_ic_pre'] = arr[16][1]
        g_var['img_ic_pref_pre'] = arr[16][2]

        g_var['img_frcpl_ref'] = arr[17][0]
        g_var['img_frcpl_pre'] = arr[17][1]
        g_var['img_frcpl_pref_pre'] = arr[17][2]

        g_var['img_erv_ref'] = arr[18][0]
        g_var['img_erv_pre'] = arr[18][1]
        g_var['img_erv_pref_pre'] = arr[18][2]

        g_var['img_rv_ref'] = arr[19][0]
        g_var['img_rv_pre'] = arr[19][1]
        g_var['img_rv_pref_pre'] = arr[19][2]

        g_var['img_rvdtlc_ref'] = arr[20][0]
        g_var['img_rvdtlc_pre'] = arr[20][1]

        g_var['img_vtg_pre'] = arr[21][0]

        g_var['img_vt_pre'] = arr[22][0]
        # lung volumes section end

        # diffusing capacity section start
        g_var['img_dlco_ref'] = arr[23][0]
        g_var['img_dlco_pre'] = arr[23][1]
        g_var['img_dlco_pref_pre'] = arr[23][2]

        g_var['img_dladj_ref'] = arr[24][0]
        g_var['img_dladj_pre'] = arr[24][1]
        g_var['img_dladj_pref_pre'] = arr[24][2]

        g_var['img_dlcodva_ref'] = arr[25][0]
        g_var['img_dlcodva_pre'] = arr[25][1]
        g_var['img_dlcodva_pref_pre'] = arr[25][2]

        g_var['img_dldvaadj_ref'] = arr[26][0]
        g_var['img_dldvaadj_pre'] = arr[26][1]
        g_var['img_dldvaadj_pref_pre'] = arr[26][2]

        g_var['img_va_pre'] = arr[27][0]

        g_var['img_ivc_pre'] = arr[28][0]
        # diffusing capacity section end

        # resistence section start
        g_var['img_rawtotal_pre'] = arr[29][0]

        g_var['img_rawinsp_pre'] = arr[30][0]

        g_var['img_rawexp_pre'] = arr[31][0]

        g_var['img_raw_ref'] = arr[32][0]
        g_var['img_raw_pre'] = arr[32][1]
        g_var['img_raw_pref_pre'] = arr[32][2]

        g_var['img_gaw_ref'] = arr[33][0]
        g_var['img_gaw_pre'] = arr[33][1]
        g_var['img_gaw_pref_pre'] = arr[33][2]

        g_var['img_sraw_ref'] = arr[34][0]
        g_var['img_sraw_pre'] = arr[34][1]
        g_var['img_sraw_pref_pre'] = arr[34][2]

        g_var['img_sgaw_ref'] = arr[35][0]
        g_var['img_sgaw_pre'] = arr[35][1]
        g_var['img_sgaw_pref_pre'] = arr[35][2]

        g_var['img_rawvtg_pre'] = arr[36][0]

        g_var['img_rawf_pre'] = arr[37][0]
        # resistence section start

        type04_img_cnt += 1
        pass

    elif ocr_for_title_searching(chart_image[6:40, 245:367]) == 'CATHOLIC' and ocr_for_title_searching(chart_image[323:596, 657:914]) == '':
        g_var['img_type'] = 'type05'
        arr = process_and_get_arr(chart_image, cur_before_image_file_name, header_type05, 330, 300)
        
        i = 0
        g_var['img_fvc_pred'] = arr[i][0]
        g_var['img_fvc_pre'] = arr[i][1]
        g_var['img_fvc_pre_ppred'] = arr[i][2]

        i += 1
        g_var['img_fev1_pred'] = arr[i][0]
        g_var['img_fev1_pre'] = arr[i][1]
        g_var['img_fev1_pre_ppred'] = arr[i][2]

        i += 1
        g_var['img_fev1dfvc_pred'] = arr[i][0]
        g_var['img_fev1dfvc_pre'] = arr[i][1]

        i += 1
        g_var['img_fef25_75_pred'] = arr[i][0]
        g_var['img_fef25_75_pre'] = arr[i][1]
        g_var['img_fef25_75_pre_ppred'] = arr[i][2]

        i += 1
        g_var['img_isofef25_75_pred'] = arr[i][0]
        g_var['img_isofef25_75_pre'] = arr[i][1]
        g_var['img_isofef25_75_pre_ppred'] = arr[i][2]

        i += 1
        g_var['img_fef75_85_pred'] = arr[i][0]
        g_var['img_fef75_85_pre'] = arr[i][1]
        g_var['img_fef75_85_pre_ppred'] = arr[i][2]

        i += 1
        g_var['img_pef_pred'] = arr[i][0]
        g_var['img_pef_pre'] = arr[i][1]
        g_var['img_pef_pre_ppred'] = arr[i][2]

        i += 1
        g_var['img_peft_pre'] = arr[i][0]

        i += 1
        g_var['img_fet100_pre'] = arr[i][0]

        i += 1
        g_var['img_fivc_pred'] = arr[i][0]
        g_var['img_fivc_pre'] = arr[i][1]
        g_var['img_fivc_pre_ppred'] = arr[i][2]

        if len(arr) == 15:
            i += 1
            g_var['img_fiv1_pre'] = arr[i][0]

        i += 1
        g_var['img_fefdfif50_pred'] = arr[i][0]
        g_var['img_fefdfif50_pre'] = arr[i][1]

        i += 1
        g_var['img_volextrap_pre'] = arr[i][0]

        i += 1
        g_var['img_fvlecode_pre'] = arr[i][0]

        i += 1
        g_var['img_mvv_pred'] = arr[i][0]

        type05_img_cnt += 1
        pass

    elif ocr_for_title_searching(chart_image[6:40, 245:367]) == 'CATHOLIC':
        g_var['img_type'] = 'type06'

        arr = process_and_get_arr(chart_image, cur_before_image_file_name, header_type06, 610, 310)

        i = 0
        g_var['img_fvc_pred'] = arr[i][0]
        g_var['img_fvc_pre'] = arr[i][1]
        g_var['img_fvc_pre_ppred'] = arr[i][2]
        g_var['img_fvc_post'] = arr[i][3]
        g_var['img_fvc_post_ppred'] = arr[i][4]
        g_var['img_fvc_pchg'] = arr[i][5]

        i += 1
        g_var['img_fev1_pred'] = arr[i][0]
        g_var['img_fev1_pre'] = arr[i][1]
        g_var['img_fev1_pre_ppred'] = arr[i][2]
        g_var['img_fev1_post'] = arr[i][3]
        g_var['img_fev1_post_ppred'] = arr[i][4]
        g_var['img_fev1_pchg'] = arr[i][5]

        i += 1
        g_var['img_fev1dfvc_pred'] = arr[i][0]
        g_var['img_fev1dfvc_pre'] = arr[i][1]
        g_var['img_fev1dfvc_post'] = arr[i][2]

        i += 1
        g_var['img_fef25_75_pred'] = arr[i][0]
        g_var['img_fef25_75_pre'] = arr[i][1]
        g_var['img_fef25_75_pre_ppred'] = arr[i][2]
        g_var['img_fef25_75_post'] = arr[i][3]
        g_var['img_fef25_75_post_ppred'] = arr[i][4]
        g_var['img_fef25_75_pchg'] = arr[i][5]

        i += 1
        g_var['img_isofef25_75_pred'] = arr[i][0]
        g_var['img_isofef25_75_pre'] = arr[i][1]
        g_var['img_isofef25_75_pre_ppred'] = arr[i][2]
        g_var['img_isofef25_75_post'] = arr[i][3]
        g_var['img_isofef25_75_post_ppred'] = arr[i][4]
        g_var['img_isofef25_75_pchg'] = arr[i][5]

        i += 1
        g_var['img_fef75_85_pred'] = arr[i][0]
        g_var['img_fef75_85_pre'] = arr[i][1]
        g_var['img_fef75_85_pre_ppred'] = arr[i][2]
        g_var['img_fef75_85_post'] = arr[i][3]
        g_var['img_fef75_85_post_ppred'] = arr[i][4]
        g_var['img_fef75_85_pchg'] = arr[i][5]

        i += 1
        g_var['img_pef_pred'] = arr[i][0]
        g_var['img_pef_pre'] = arr[i][1]
        g_var['img_pef_pre_ppred'] = arr[i][2]
        g_var['img_pef_post'] = arr[i][3]
        g_var['img_pef_post_ppred'] = arr[i][4]
        g_var['img_pef_pchg'] = arr[i][5]

        i += 1
        g_var['img_fet100_pre'] = arr[i][0]
        g_var['img_fet100_post'] = arr[i][1]
        g_var['img_fet100_pchg'] = arr[i][2]

        i += 1
        g_var['img_fivc_pred'] = arr[i][0]
        g_var['img_fivc_pre'] = arr[i][1]
        g_var['img_fivc_pre_ppred'] = arr[i][2]
        g_var['img_fivc_post'] = arr[i][3]
        g_var['img_fivc_post_ppred'] = arr[i][4]
        g_var['img_fivc_pchg'] = arr[i][5]

        i += 1
        g_var['img_fev1_pred'] = arr[i][0]
        g_var['img_fev1_pre'] = arr[i][1]
        g_var['img_fev1_pre_ppred'] = arr[i][2]
        g_var['img_fev1_post'] = arr[i][3]
        g_var['img_fev1_post_ppred'] = arr[i][4]
        g_var['img_fev1_pchg'] = arr[i][5]

        if len(arr) == 15:
            i += 1
            g_var['img_fiv1_pre'] = arr[i][0]
            g_var['img_fiv1_post'] = arr[i][1]
            g_var['img_fiv1_pchg'] = arr[i][2]

        i += 1
        g_var['img_fefdfif50_pred'] = arr[i][0]
        g_var['img_fefdfif50_pre'] = arr[i][1]
        g_var['img_fefdfif50_post'] = arr[i][2]
        g_var['img_fefdfif50_pchg'] = arr[i][3]

        i += 1
        g_var['img_volextrap_pre'] = arr[i][0]
        g_var['img_volextrap_post'] = arr[i][1]
        g_var['img_volextrap_pchg'] = arr[i][2]

        i += 1
        g_var['img_fvlecode_pre'] = arr[i][0]
        g_var['img_fvlecode_post'] = arr[i][1]

        i += 1
        g_var['img_mvv_pred'] = arr[i][0]

        type06_img_cnt += 1
        pass

    elif ocr_for_title_searching(chart_image[66:99, 538:647]) == 'REPORT':
        g_var['img_type'] = 'type07'

        arr = process_and_get_arr(chart_image, cur_before_image_file_name, header_type07_1, 400, 185)

        # best data chart start
        i = 0
        g_var['img_fvc_ref'] = arr[i][0]
        g_var['img_fvc_pre'] = arr[i][1]
        g_var['img_fvc_pref_pre'] = arr[i][2]
        g_var['img_fvc_post'] = arr[i][3]
        g_var['img_fvc_pref_post'] = arr[i][4]
        g_var['img_fvc_pchg'] = arr[i][5]

        i += 1
        g_var['img_fev1_ref'] = arr[i][0]
        g_var['img_fev1_pre'] = arr[i][1]
        g_var['img_fev1_pref_pre'] = arr[i][2]
        g_var['img_fev1_post'] = arr[i][3]
        g_var['img_fev1_pref_post'] = arr[i][4]
        g_var['img_fev1_pchg'] = arr[i][5]

        i += 1
        g_var['img_fev1dfvc_ref'] = arr[i][0]
        g_var['img_fev1dfvc_pre'] = arr[i][1]
        g_var['img_fev1dfvc_post'] = arr[i][2]

        i += 1
        g_var['img_fef25_75_ref'] = arr[i][0]
        g_var['img_fef25_75_pre'] = arr[i][1]
        g_var['img_fef25_75_pref_pre'] = arr[i][2]
        g_var['img_fef25_75_post'] = arr[i][3]
        g_var['img_fef25_75_pref_post'] = arr[i][4]
        g_var['img_fef25_75_pchg'] = arr[i][5]

        i += 1
        g_var['img_pef_ref'] = arr[i][0]
        g_var['img_pef_pre'] = arr[i][1]
        g_var['img_pef_pref_pre'] = arr[i][2]
        g_var['img_pef_post'] = arr[i][3]
        g_var['img_pef_pref_post'] = arr[i][4]
        g_var['img_pef_pchg'] = arr[i][5]

        i += 1
        g_var['img_peft_pre'] = arr[i][0]
        g_var['img_peft_post'] = arr[i][1]
        g_var['img_peft_pchg'] = arr[i][2]

        i += 1
        g_var['img_fet100_pre'] = arr[i][0]
        g_var['img_fet100_post'] = arr[i][1]
        g_var['img_fet100_pchg'] = arr[i][2]

        i += 1
        g_var['img_fivc_ref'] = arr[i][0]
        g_var['img_fivc_pre'] = arr[i][1]
        g_var['img_fivc_pref_pre'] = arr[i][2]
        g_var['img_fivc_post'] = arr[i][3]
        g_var['img_fivc_pref_post'] = arr[i][4]
        g_var['img_fivc_pchg'] = arr[i][5]

        i += 1
        if len(arr[i]) == 3:
            g_var['img_volextrap_pre'] = arr[i][0]
            g_var['img_volextrap_post'] = arr[i][1]
            g_var['img_volextrap_pchg'] = arr[i][2]
        else:
            g_var['img_volextrap_post'] = arr[i][0]

        i += 1
        g_var['img_fvlecode_pre'] = arr[i][0]
        g_var['img_fvlecode_post'] = arr[i][1]
        # best data chart end

        # all trials chart start
        arr = process_and_get_arr(chart_image, cur_before_image_file_name, header_type07_2, 400, 190)

        i = 0
        g_var['img_fvc_tri1'] = len(arr[i]) > 0 and arr[i][0] or ''
        g_var['img_fvc_tri2'] = len(arr[i]) > 1 and arr[i][1] or ''
        g_var['img_fvc_tri3'] = len(arr[i]) > 2 and arr[i][2] or ''
        g_var['img_fvc_tri4'] = len(arr[i]) > 3 and arr[i][3] or ''
        g_var['img_fvc_tri5'] = len(arr[i]) > 4 and arr[i][4] or ''
        g_var['img_fvc_tri6'] = len(arr[i]) > 5 and arr[i][5] or ''
        g_var['img_fvc_tri7'] = len(arr[i]) > 6 and arr[i][6] or ''
        g_var['img_fvc_tri8'] = len(arr[i]) > 7 and arr[i][7] or ''

        i += 1
        g_var['img_fev1_tri1'] = len(arr[i]) > 0 and arr[i][0] or ''
        g_var['img_fev1_tri2'] = len(arr[i]) > 1 and arr[i][1] or ''
        g_var['img_fev1_tri3'] = len(arr[i]) > 2 and arr[i][2] or ''
        g_var['img_fev1_tri4'] = len(arr[i]) > 3 and arr[i][3] or ''
        g_var['img_fev1_tri5'] = len(arr[i]) > 4 and arr[i][4] or ''
        g_var['img_fev1_tri6'] = len(arr[i]) > 5 and arr[i][5] or ''
        g_var['img_fev1_tri7'] = len(arr[i]) > 6 and arr[i][6] or ''
        g_var['img_fev1_tri8'] = len(arr[i]) > 7 and arr[i][7] or ''

        i += 1
        g_var['img_fev1dfvc_tri1'] = len(arr[i]) > 0 and arr[i][0] or ''
        g_var['img_fev1dfvc_tri2'] = len(arr[i]) > 1 and arr[i][1] or ''
        g_var['img_fev1dfvc_tri3'] = len(arr[i]) > 2 and arr[i][2] or ''
        g_var['img_fev1dfvc_tri4'] = len(arr[i]) > 3 and arr[i][3] or ''
        g_var['img_fev1dfvc_tri5'] = len(arr[i]) > 4 and arr[i][4] or ''
        g_var['img_fev1dfvc_tri6'] = len(arr[i]) > 5 and arr[i][5] or ''
        g_var['img_fev1dfvc_tri7'] = len(arr[i]) > 6 and arr[i][6] or ''
        g_var['img_fev1dfvc_tri8'] = len(arr[i]) > 7 and arr[i][7] or ''

        i += 1
        g_var['img_fef25_75_tri1'] = len(arr[i]) > 0 and arr[i][0] or ''
        g_var['img_fef25_75_tri2'] = len(arr[i]) > 1 and arr[i][1] or ''
        g_var['img_fef25_75_tri3'] = len(arr[i]) > 2 and arr[i][2] or ''
        g_var['img_fef25_75_tri4'] = len(arr[i]) > 3 and arr[i][3] or ''
        g_var['img_fef25_75_tri5'] = len(arr[i]) > 4 and arr[i][4] or ''
        g_var['img_fef25_75_tri6'] = len(arr[i]) > 5 and arr[i][5] or ''
        g_var['img_fef25_75_tri7'] = len(arr[i]) > 6 and arr[i][6] or ''
        g_var['img_fef25_75_tri8'] = len(arr[i]) > 7 and arr[i][7] or ''

        i += 1
        g_var['img_pef_tri1'] = len(arr[i]) > 0 and arr[i][0] or ''
        g_var['img_pef_tri2'] = len(arr[i]) > 1 and arr[i][1] or ''
        g_var['img_pef_tri3'] = len(arr[i]) > 2 and arr[i][2] or ''
        g_var['img_pef_tri4'] = len(arr[i]) > 3 and arr[i][3] or ''
        g_var['img_pef_tri5'] = len(arr[i]) > 4 and arr[i][4] or ''
        g_var['img_pef_tri6'] = len(arr[i]) > 5 and arr[i][5] or ''
        g_var['img_pef_tri7'] = len(arr[i]) > 6 and arr[i][6] or ''
        g_var['img_pef_tri8'] = len(arr[i]) > 7 and arr[i][7] or ''

        i += 1
        g_var['img_peft_tri1'] = len(arr[i]) > 0 and arr[i][0] or ''
        g_var['img_peft_tri2'] = len(arr[i]) > 1 and arr[i][1] or ''
        g_var['img_peft_tri3'] = len(arr[i]) > 2 and arr[i][2] or ''
        g_var['img_peft_tri4'] = len(arr[i]) > 3 and arr[i][3] or ''
        g_var['img_peft_tri5'] = len(arr[i]) > 4 and arr[i][4] or ''
        g_var['img_peft_tri6'] = len(arr[i]) > 5 and arr[i][5] or ''
        g_var['img_peft_tri7'] = len(arr[i]) > 6 and arr[i][6] or ''
        g_var['img_peft_tri8'] = len(arr[i]) > 7 and arr[i][7] or ''

        i += 1
        g_var['img_fet100_tri1'] = len(arr[i]) > 0 and arr[i][0] or ''
        g_var['img_fet100_tri2'] = len(arr[i]) > 1 and arr[i][1] or ''
        g_var['img_fet100_tri3'] = len(arr[i]) > 2 and arr[i][2] or ''
        g_var['img_fet100_tri4'] = len(arr[i]) > 3 and arr[i][3] or ''
        g_var['img_fet100_tri5'] = len(arr[i]) > 4 and arr[i][4] or ''
        g_var['img_fet100_tri6'] = len(arr[i]) > 5 and arr[i][5] or ''
        g_var['img_fet100_tri7'] = len(arr[i]) > 6 and arr[i][6] or ''
        g_var['img_fet100_tri8'] = len(arr[i]) > 7 and arr[i][7] or ''

        i += 1
        g_var['img_fivc_tri1'] = len(arr[i]) > 0 and arr[i][0] or ''
        g_var['img_fivc_tri2'] = len(arr[i]) > 1 and arr[i][1] or ''
        g_var['img_fivc_tri3'] = len(arr[i]) > 2 and arr[i][2] or ''
        g_var['img_fivc_tri4'] = len(arr[i]) > 3 and arr[i][3] or ''
        g_var['img_fivc_tri5'] = len(arr[i]) > 4 and arr[i][4] or ''
        g_var['img_fivc_tri6'] = len(arr[i]) > 5 and arr[i][5] or ''
        g_var['img_fivc_tri7'] = len(arr[i]) > 6 and arr[i][6] or ''
        g_var['img_fivc_tri8'] = len(arr[i]) > 7 and arr[i][7] or ''

        i += 1
        g_var['img_volextrap_tri1'] = len(arr[i]) > 0 and arr[i][0] or ''
        g_var['img_volextrap_tri2'] = len(arr[i]) > 1 and arr[i][1] or ''
        g_var['img_volextrap_tri3'] = len(arr[i]) > 2 and arr[i][2] or ''
        g_var['img_volextrap_tri4'] = len(arr[i]) > 3 and arr[i][3] or ''
        g_var['img_volextrap_tri5'] = len(arr[i]) > 4 and arr[i][4] or ''
        g_var['img_volextrap_tri6'] = len(arr[i]) > 5 and arr[i][5] or ''
        g_var['img_volextrap_tri7'] = len(arr[i]) > 6 and arr[i][6] or ''
        g_var['img_volextrap_tri8'] = len(arr[i]) > 7 and arr[i][7] or ''

        i += 1
        g_var['img_fvlecode_tri1'] = len(arr[i]) > 0 and arr[i][0] or ''
        g_var['img_fvlecode_tri2'] = len(arr[i]) > 1 and arr[i][1] or ''
        g_var['img_fvlecode_tri3'] = len(arr[i]) > 2 and arr[i][2] or ''
        g_var['img_fvlecode_tri4'] = len(arr[i]) > 3 and arr[i][3] or ''
        g_var['img_fvlecode_tri5'] = len(arr[i]) > 4 and arr[i][4] or ''
        g_var['img_fvlecode_tri6'] = len(arr[i]) > 5 and arr[i][5] or ''
        g_var['img_fvlecode_tri7'] = len(arr[i]) > 6 and arr[i][6] or ''
        g_var['img_fvlecode_tri8'] = len(arr[i]) > 7 and arr[i][7] or ''
        # all trials chart end

        type07_img_cnt += 1
        pass

    elif ocr_for_title_searching(chart_image[266:303, 3:44]) == 'Lung':
        g_var['img_type'] = 'type08'
        arr = process_and_get_arr(chart_image, cur_before_image_file_name, header_type08, 215, 390)

        i = 0
        g_var['img_tlc_ref'] = arr[i][0]
        g_var['img_tlc_pre'] = arr[i][1]
        g_var['img_tlc_pref_pre'] = arr[i][2]

        i += 1
        g_var['img_vc_ref'] = arr[i][0]
        g_var['img_vc_pre'] = arr[i][1]
        g_var['img_vc_pref_pre'] = arr[i][2]

        i += 1
        g_var['img_ic_ref'] = arr[i][0]
        g_var['img_ic_pre'] = arr[i][1]
        g_var['img_ic_pref_pre'] = arr[i][2]

        i += 1
        g_var['img_frcpl_ref'] = arr[i][0]
        g_var['img_frcpl_pre'] = arr[i][1]
        g_var['img_frcpl_pref_pre'] = arr[i][2]

        i += 1
        g_var['img_erv_ref'] = arr[i][0]
        g_var['img_erv_pre'] = arr[i][1]
        g_var['img_erv_pref_pre'] = arr[i][2]

        i += 1
        g_var['img_rv_ref'] = arr[i][0]
        g_var['img_rv_pre'] = arr[i][1]
        g_var['img_rv_pref_pre'] = arr[i][2]

        i += 1
        g_var['img_rvdtlc_ref'] = arr[i][0]
        g_var['img_rvdtlc_pre'] = arr[i][1]

        i += 1
        g_var['img_vtg_pre'] = arr[i][0]

        i += 1
        g_var['img_vt_pre'] = arr[i][0]

        i += 1
        g_var['img_rawinsp_pre'] = arr[i][0]

        i += 1
        g_var['img_rawexp_pre'] = arr[i][0]

        i += 1
        g_var['img_raw_ref'] = arr[i][0]
        g_var['img_raw_pre'] = arr[i][1]
        g_var['img_raw_pref_pre'] = arr[i][2]

        i += 1
        g_var['img_gaw_ref'] = arr[i][0]
        g_var['img_gaw_pre'] = arr[i][1]
        g_var['img_gaw_pref_pre'] = arr[i][2]

        i += 1
        g_var['img_sraw_ref'] = arr[i][0]
        g_var['img_sraw_pre'] = arr[i][1]
        g_var['img_sraw_pref_pre'] = arr[i][2]

        i += 1
        g_var['img_sgaw_ref'] = arr[i][0]
        g_var['img_sgaw_pre'] = arr[i][1]
        g_var['img_sgaw_pref_pre'] = arr[i][2]

        i += 1
        g_var['img_rawvtg_pre'] = arr[i][0]

        i += 1
        g_var['img_rawf_pre'] = arr[i][0]
        
        # for ar in arr:
        #     for ns in ar:
        #         print(ns)
        #     print()
        type08_img_cnt += 1
        pass

    else:
        g_var['img_type'] = 'typeUK'
        typeUK_img_cnt += 1
        pass
            
    for f in fs:
        f.result()
    
    q.put(['END', True])
    while True:
        q_val = q.get()
        if q_val[0] == 'END':
            break
        g_var[q_val[0]] = q_val[1]
    
    # copy image to sorted dir
    before_image_files_counter += 1
    # cv2.imwrite(g_var['img_type'] + "/" + g_var['img_type'] + "_" + cur_before_image_file_name, chart_image)

    return [cur_before_image_file_path, g_var, arr, cur_before_image_file_name, chart_image]

    
def process(before_path):
    before_image_file_paths = []
    if len(before_path) == 1 and os.path.isdir(before_path[0]) == 1:
        before_image_file_paths = glob.glob(before_path[0] + "/*.jpg")
    else:
        before_image_file_paths = before_path

    ws, wb = init_worksheet()
    pfs = []
    for cur_before_image_file_path in before_image_file_paths:
        pfs.append(pool.submit(process_chart_and_get_v_gar, cur_before_image_file_path))

    for f in pfs:
        f_result = f.result()
        img_path = f_result[0]
        g_var = f_result[1]
        arr = f_result[2]
        img_name = f_result[3]
        img = f_result[4]

        for ar in arr:
            for ns in ar:
                print(ns)
            print()
        print(img_path + " -> " + g_var['img_type'])
        print()

        # create new directory
        if os.path.exists (g_var['img_type']) == 0:
            os.mkdir(g_var['img_type'])

        cv2.imwrite(g_var['img_type'] + "/" + g_var['img_type'] + "_" + img_name, img)
        # add data into OOMII_DB
        append_g_var_to_excel(ws, wb, g_var)

    print()
    print('Total: ' + str(before_image_files_counter))
    print('type01: ' + str(type01_img_cnt))
    print('type02: ' + str(type02_img_cnt))
    print('type03: ' + str(type03_img_cnt))
    print('type04: ' + str(type04_img_cnt))
    print('type05: ' + str(type05_img_cnt))
    print('type06: ' + str(type06_img_cnt))
    print('type07: ' + str(type07_img_cnt))
    print('type08: ' + str(type08_img_cnt))
    print('typeUK: ' + str(typeUK_img_cnt))


def init_worksheet():
    # create new excel file 
    if os.path.isfile('OOMII.xlsx') == 0:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws = wb.create_sheet("OOMII_DB", 0)
        ws ["A1"] = "type"
        ws ["B1"] = "pid"
        ws ["C1"] = "date"
        ws ["D1"] = "age"
        ws ["E1"] = "height"
        ws ["F1"] = "weight"
        ws ["G1"] = "gender"
        ws ["H1"] = "fvc_dose_lv1"
        ws ["I1"] = "fvc_dose_lv2"
        ws ["J1"] = "fvc_dose_lv3"
        ws ["K1"] = "fvc_dose_lv4"
        ws ["L1"] = "fvc_dose_lv5"
        ws ["M1"] = "fvc_dose_lv6"
        ws ["N1"] = "fvc_dose_lv7"
        ws ["O1"] = "fvc_dose_lv8"
        ws ["P1"] = "fvc_dose_lv9"
        ws ["Q1"] = "fvc_ref"
        ws ["R1"] = "fvc_pre"
        ws ["S1"] = "fvc_post"
        ws ["T1"] = "fvc_lv1"
        ws ["U1"] = "fvc_lv2"
        ws ["V1"] = "fvc_lv3"
        ws ["W1"] = "fvc_lv4"
        ws ["X1"] = "fvc_lv5"
        ws ["Y1"] = "fvc_lv6"
        ws ["Z1"] = "fvc_lv7"
        ws ["AA1"] = "fvc_lv8"
        ws ["AB1"] = "fvc_lv9"
        ws ["AC1"] = "fvc_tri1"
        ws ["AD1"] = "fvc_tri2"
        ws ["AE1"] = "fvc_tri3"
        ws ["AF1"] = "fvc_tri4"
        ws ["AG1"] = "fvc_tri5"
        ws ["AH1"] = "fvc_tri6"
        ws ["AI1"] = "fvc_tri7"
        ws ["AJ1"] = "fvc_tri8"
        ws ["AK1"] = "fvc_pref_pre"
        ws ["AL1"] = "fvc_pref_post"
        ws ["AM1"] = "fvc_pref_lv1"
        ws ["AN1"] = "fvc_pref_lv2"
        ws ["AO1"] = "fvc_pref_lv3"
        ws ["AP1"] = "fvc_pref_lv4"
        ws ["AQ1"] = "fvc_pref_lv5"
        ws ["AR1"] = "fvc_pref_lv6"
        ws ["AS1"] = "fvc_pref_lv7"
        ws ["AT1"] = "fvc_pref_lv8"
        ws ["AU1"] = "fvc_pref_lv9"
        ws ["AV1"] = "fvc_pchg"
        ws ["AW1"] = "fvc_pchg_lv1"
        ws ["AX1"] = "fvc_pchg_lv2"
        ws ["AY1"] = "fvc_pchg_lv3"
        ws ["AZ1"] = "fvc_pchg_lv4"
        ws ["BA1"] = "fvc_pchg_lv5"
        ws ["BB1"] = "fvc_pchg_lv6"
        ws ["BC1"] = "fvc_pchg_lv7"
        ws ["BD1"] = "fvc_pchg_lv8"
        ws ["BE1"] = "fvc_pchg_lv9"
        ws ["BF1"] = "fvc_pred"
        ws ["BG1"] = "fvc_pre_ppred"
        ws ["BH1"] = "fvc_post_ppred"
        ws ["BI1"] = "fev05_ref"
        ws ["BJ1"] = "fev05_pre"
        ws ["BK1"] = "fev05_post"
        ws ["BL1"] = "fev05_pref_pre"
        ws ["BM1"] = "fev05_pref_post"
        ws ["BN1"] = "fev05_pchg"
        ws ["BO1"] = "fev1_dose_lv1"
        ws ["BP1"] = "fev1_dose_lv2"
        ws ["BQ1"] = "fev1_dose_lv3"
        ws ["BR1"] = "fev1_dose_lv4"
        ws ["BS1"] = "fev1_dose_lv5"
        ws ["BT1"] = "fev1_dose_lv6"
        ws ["BU1"] = "fev1_dose_lv7"
        ws ["BV1"] = "fev1_dose_lv8"
        ws ["BW1"] = "fev1_dose_lv9"
        ws ["BX1"] = "fev1_ref"
        ws ["BY1"] = "fev1_pre"
        ws ["BZ1"] = "fev1_post"
        ws ["CA1"] = "fev1_lv1"
        ws ["CB1"] = "fev1_lv2"
        ws ["CC1"] = "fev1_lv3"
        ws ["CD1"] = "fev1_lv4"
        ws ["CE1"] = "fev1_lv5"
        ws ["CF1"] = "fev1_lv6"
        ws ["CG1"] = "fev1_lv7"
        ws ["CH1"] = "fev1_lv8"
        ws ["CI1"] = "fev1_lv9"
        ws ["CJ1"] = "fev1_tri1"
        ws ["CK1"] = "fev1_tri2"
        ws ["CL1"] = "fev1_tri3"
        ws ["CM1"] = "fev1_tri4"
        ws ["CN1"] = "fev1_tri5"
        ws ["CO1"] = "fev1_tri6"
        ws ["CP1"] = "fev1_tri7"
        ws ["CQ1"] = "fev1_tri8"
        ws ["CR1"] = "fev1_pref_pre"
        ws ["CS1"] = "fev1_pref_post"
        ws ["CT1"] = "fev1_pref_lv1"
        ws ["CU1"] = "fev1_pref_lv2"
        ws ["CV1"] = "fev1_pref_lv3"
        ws ["CW1"] = "fev1_pref_lv4"
        ws ["CX1"] = "fev1_pref_lv5"
        ws ["CY1"] = "fev1_pref_lv6"
        ws ["CZ1"] = "fev1_pref_lv7"
        ws ["DA1"] = "fev1_pref_lv8"
        ws ["DB1"] = "fev1_pref_lv9"
        ws ["DC1"] = "fev1_pchg"
        ws ["DD1"] = "fev1_pchg_lv1"
        ws ["DE1"] = "fev1_pchg_lv2"
        ws ["DF1"] = "fev1_pchg_lv3"
        ws ["DG1"] = "fev1_pchg_lv4"
        ws ["DH1"] = "fev1_pchg_lv5"
        ws ["DI1"] = "fev1_pchg_lv6"
        ws ["DJ1"] = "fev1_pchg_lv7"
        ws ["DK1"] = "fev1_pchg_lv8"
        ws ["DL1"] = "fev1_pchg_lv9"
        ws ["DM1"] = "fev1_pred"
        ws ["DN1"] = "fev1_pre_ppred"
        ws ["DO1"] = "fev1_post_ppred"
        ws ["DP1"] = "pc_fev1"
        ws ["DQ1"] = "fev1_pc"
        ws ["DR1"] = "fev1dfvc_ref"
        ws ["DS1"] = "fev1dfvc_pre"
        ws ["DT1"] = "fev1dfvc_post"
        ws ["DU1"] = "fev1dfvc_pred"
        ws ["DV1"] = "fev1dfvc_tri1"
        ws ["DW1"] = "fev1dfvc_tri2"
        ws ["DX1"] = "fev1dfvc_tri3"
        ws ["DY1"] = "fev1dfvc_tri4"
        ws ["DZ1"] = "fev1dfvc_tri5"
        ws ["EA1"] = "fev1dfvc_tri6"
        ws ["EB1"] = "fev1dfvc_tri7"
        ws ["EC1"] = "fev1dfvc_tri8"
        ws ["ED1"] = "fev3dfvc_ref"
        ws ["EE1"] = "fev3dfvc_pre"
        ws ["EF1"] = "fev3dfvc_post"
        ws ["EG1"] = "fev3dfvc_pref_pre"
        ws ["EH1"] = "fev3dfvc_pref_post"
        ws ["EI1"] = "fev3dfvc_pchg"
        ws ["EJ1"] = "fef25_75_dose_lv1"
        ws ["EK1"] = "fef25_75_dose_lv2"
        ws ["EL1"] = "fef25_75_dose_lv3"
        ws ["EM1"] = "fef25_75_dose_lv4"
        ws ["EN1"] = "fef25_75_dose_lv5"
        ws ["EO1"] = "fef25_75_dose_lv6"
        ws ["EP1"] = "fef25_75_dose_lv7"
        ws ["EQ1"] = "fef25_75_dose_lv8"
        ws ["ER1"] = "fef25_75_dose_lv9"
        ws ["ES1"] = "fef25_75_ref"
        ws ["ET1"] = "fef25_75_pre"
        ws ["EU1"] = "fef25_75_post"
        ws ["EV1"] = "fef25_75_lv1"
        ws ["EW1"] = "fef25_75_lv2"
        ws ["EX1"] = "fef25_75_lv3"
        ws ["EY1"] = "fef25_75_lv4"
        ws ["EZ1"] = "fef25_75_lv5"
        ws ["FA1"] = "fef25_75_lv6"
        ws ["FB1"] = "fef25_75_lv7"
        ws ["FC1"] = "fef25_75_lv8"
        ws ["FD1"] = "fef25_75_lv9"
        ws ["FE1"] = "fef25_75_tri1"
        ws ["FF1"] = "fef25_75_tri2"
        ws ["FG1"] = "fef25_75_tri3"
        ws ["FH1"] = "fef25_75_tri4"
        ws ["FI1"] = "fef25_75_tri5"
        ws ["FJ1"] = "fef25_75_tri6"
        ws ["FK1"] = "fef25_75_tri7"
        ws ["FL1"] = "fef25_75_tri8"
        ws ["FM1"] = "fef25_75_pref_pre"
        ws ["FN1"] = "fef25_75_pref_post"
        ws ["FO1"] = "fef25_75_pref_lv1"
        ws ["FP1"] = "fef25_75_pref_lv2"
        ws ["FQ1"] = "fef25_75_pref_lv3"
        ws ["FR1"] = "fef25_75_pref_lv4"
        ws ["FS1"] = "fef25_75_pref_lv5"
        ws ["FT1"] = "fef25_75_pref_lv6"
        ws ["FU1"] = "fef25_75_pref_lv7"
        ws ["FV1"] = "fef25_75_pref_lv8"
        ws ["FW1"] = "fef25_75_pref_lv9"
        ws ["FX1"] = "fef25_75_pchg"
        ws ["FY1"] = "fef25_75_pchg_lv1"
        ws ["FZ1"] = "fef25_75_pchg_lv2"
        ws ["GA1"] = "fef25_75_pchg_lv3"
        ws ["GB1"] = "fef25_75_pchg_lv4"
        ws ["GC1"] = "fef25_75_pchg_lv5"
        ws ["GD1"] = "fef25_75_pchg_lv6"
        ws ["GE1"] = "fef25_75_pchg_lv7"
        ws ["GF1"] = "fef25_75_pchg_lv8"
        ws ["GG1"] = "fef25_75_pchg_lv9"
        ws ["GH1"] = "fef25_75_pred"
        ws ["GI1"] = "fef25_75_pre_ppred"
        ws ["GJ1"] = "fef25_75_post_ppred"
        ws ["GK1"] = "isofef25_75_pchg"
        ws ["GL1"] = "isofef25_75_pred"
        ws ["GM1"] = "isofef25_75_pre"
        ws ["GN1"] = "isofef25_75_pre_ppred"
        ws ["GO1"] = "isofef25_75_post"
        ws ["GP1"] = "isofef25_75_post_ppred"
        ws ["GQ1"] = "fef75_85_ref"
        ws ["GR1"] = "fef75_85_pre"
        ws ["GS1"] = "fef75_85_post"
        ws ["GT1"] = "fef75_85_pref_pre"
        ws ["GU1"] = "fef75_85_pref_post"
        ws ["GV1"] = "fef75_85_pchg"
        ws ["GW1"] = "fef75_85_pred"
        ws ["GX1"] = "fef75_85_pre_ppred"
        ws ["GY1"] = "fef75_85_post_ppred"
        ws ["GZ1"] = "fef25_ref"
        ws ["HA1"] = "fef25_pre"
        ws ["HB1"] = "fef25_post"
        ws ["HC1"] = "fef25_pref_pre"
        ws ["HD1"] = "fef25_pref_post"
        ws ["HE1"] = "fef25_pchg"
        ws ["HF1"] = "fef50_ref"
        ws ["HG1"] = "fef50_pre"
        ws ["HH1"] = "fef50_post"
        ws ["HI1"] = "fef50_pref_pre"
        ws ["HJ1"] = "fef50_pref_post"
        ws ["HK1"] = "fef50_pchg"
        ws ["HL1"] = "fef75_ref"
        ws ["HM1"] = "fef75_pre"
        ws ["HN1"] = "fef75_post"
        ws ["HO1"] = "fef75_pref_pre"
        ws ["HP1"] = "fef75_pref_post"
        ws ["HQ1"] = "fef75_pchg"
        ws ["HR1"] = "fef200_1200_ref"
        ws ["HS1"] = "fef200_1200_pre"
        ws ["HT1"] = "fef200_1200_post"
        ws ["HU1"] = "fef200_1200_pref_pre"
        ws ["HV1"] = "fef200_1200_pref_post"
        ws ["HW1"] = "fef200_1200_pchg"
        ws ["HX1"] = "pef_dose_lv1"
        ws ["HY1"] = "pef_dose_lv2"
        ws ["HZ1"] = "pef_dose_lv3"
        ws ["IA1"] = "pef_dose_lv4"
        ws ["IB1"] = "pef_dose_lv5"
        ws ["IC1"] = "pef_dose_lv6"
        ws ["ID1"] = "pef_dose_lv7"
        ws ["IE1"] = "pef_dose_lv8"
        ws ["IF1"] = "pef_dose_lv9"
        ws ["IG1"] = "pef_ref"
        ws ["IH1"] = "pef_pre"
        ws ["II1"] = "pef_post"
        ws ["IJ1"] = "pef_lv1"
        ws ["IK1"] = "pef_lv2"
        ws ["IL1"] = "pef_lv3"
        ws ["IM1"] = "pef_lv4"
        ws ["IN1"] = "pef_lv5"
        ws ["IO1"] = "pef_lv6"
        ws ["IP1"] = "pef_lv7"
        ws ["IQ1"] = "pef_lv8"
        ws ["IR1"] = "pef_lv9"
        ws ["IS1"] = "pef_tri1"
        ws ["IT1"] = "pef_tri2"
        ws ["IU1"] = "pef_tri3"
        ws ["IV1"] = "pef_tri4"
        ws ["IW1"] = "pef_tri5"
        ws ["IX1"] = "pef_tri6"
        ws ["IY1"] = "pef_tri7"
        ws ["IZ1"] = "pef_tri8"
        ws ["JA1"] = "pef_pref_pre"
        ws ["JB1"] = "pef_pref_post"
        ws ["JC1"] = "pef_pref_lv1"
        ws ["JD1"] = "pef_pref_lv2"
        ws ["JE1"] = "pef_pref_lv3"
        ws ["JF1"] = "pef_pref_lv4"
        ws ["JG1"] = "pef_pref_lv5"
        ws ["JH1"] = "pef_pref_lv6"
        ws ["JI1"] = "pef_pref_lv7"
        ws ["JJ1"] = "pef_pref_lv8"
        ws ["JK1"] = "pef_pref_lv9"
        ws ["JL1"] = "pef_pchg"
        ws ["JM1"] = "pef_pchg_lv1"
        ws ["JN1"] = "pef_pchg_lv2"
        ws ["JO1"] = "pef_pchg_lv3"
        ws ["JP1"] = "pef_pchg_lv4"
        ws ["JQ1"] = "pef_pchg_lv5"
        ws ["JR1"] = "pef_pchg_lv6"
        ws ["JS1"] = "pef_pchg_lv7"
        ws ["JT1"] = "pef_pchg_lv8"
        ws ["JU1"] = "pef_pchg_lv9"
        ws ["JV1"] = "pef_pred"
        ws ["JW1"] = "pef_pre_ppred"
        ws ["JX1"] = "pef_post_ppred"
        ws ["JY1"] = "peft_pchg"
        ws ["JZ1"] = "peft_pre"
        ws ["KA1"] = "peft_post"
        ws ["KB1"] = "peft_tri1"
        ws ["KC1"] = "peft_tri2"
        ws ["KD1"] = "peft_tri3"
        ws ["KE1"] = "peft_tri4"
        ws ["KF1"] = "peft_tri5"
        ws ["KG1"] = "peft_tri6"
        ws ["KH1"] = "peft_tri7"
        ws ["KI1"] = "peft_tri8"
        ws ["KJ1"] = "fet100_pchg"
        ws ["KK1"] = "fet100_pre"
        ws ["KL1"] = "fet100_post"
        ws ["KM1"] = "fet100_tri1"
        ws ["KN1"] = "fet100_tri2"
        ws ["KO1"] = "fet100_tri3"
        ws ["KP1"] = "fet100_tri4"
        ws ["KQ1"] = "fet100_tri5"
        ws ["KR1"] = "fet100_tri6"
        ws ["KS1"] = "fet100_tri7"
        ws ["KT1"] = "fet100_tri8"
        ws ["KU1"] = "fivc_ref"
        ws ["KV1"] = "fivc_pre"
        ws ["KW1"] = "fivc_post"
        ws ["KX1"] = "fivc_pref_pre"
        ws ["KY1"] = "fivc_pref_post"
        ws ["KZ1"] = "fivc_pchg"
        ws ["LA1"] = "fivc_pred"
        ws ["LB1"] = "fivc_pre_ppred"
        ws ["LC1"] = "fivc_post_ppred"
        ws ["LD1"] = "fivc_tri1"
        ws ["LE1"] = "fivc_tri2"
        ws ["LF1"] = "fivc_tri3"
        ws ["LG1"] = "fivc_tri4"
        ws ["LH1"] = "fivc_tri5"
        ws ["LI1"] = "fivc_tri6"
        ws ["LJ1"] = "fivc_tri7"
        ws ["LK1"] = "fivc_tri8"
        ws ["LL1"] = "fiv1_pchg"
        ws ["LM1"] = "fiv1_pre"
        ws ["LN1"] = "fiv1_post"
        ws ["LO1"] = "fefdfif50_pchg"
        ws ["LP1"] = "fefdfif50_pred"
        ws ["LQ1"] = "fefdfif50_pre"
        ws ["LR1"] = "fefdfif50_pre_ppred"
        ws ["LS1"] = "fefdfif50_post"
        ws ["LT1"] = "fefdfif50_post_ppred"
        ws ["LU1"] = "volextrap_pchg"
        ws ["LV1"] = "volextrap_pre"
        ws ["LW1"] = "volextrap_post"
        ws ["LX1"] = "volextrap_tri1"
        ws ["LY1"] = "volextrap_tri2"
        ws ["LZ1"] = "volextrap_tri3"
        ws ["MA1"] = "volextrap_tri4"
        ws ["MB1"] = "volextrap_tri5"
        ws ["MC1"] = "volextrap_tri6"
        ws ["MD1"] = "volextrap_tri7"
        ws ["ME1"] = "volextrap_tri8"
        ws ["MF1"] = "fvlecode_pre"
        ws ["MG1"] = "fvlecode_post"
        ws ["MH1"] = "fvlecode_tri1"
        ws ["MI1"] = "fvlecode_tri2"
        ws ["MJ1"] = "fvlecode_tri3"
        ws ["MK1"] = "fvlecode_tri4"
        ws ["ML1"] = "fvlecode_tri5"
        ws ["MM1"] = "fvlecode_tri6"
        ws ["MN1"] = "fvlecode_tri7"
        ws ["MO1"] = "fvlecode_tri8"
        ws ["MP1"] = "mvv_pred"
        ws ["MQ1"] = "tlc_ref"
        ws ["MR1"] = "tlc_pre"
        ws ["MS1"] = "tlc_pref_pre"
        ws ["MT1"] = "vc_ref"
        ws ["MU1"] = "vc_pre"
        ws ["MV1"] = "vc_pref_pre"
        ws ["MW1"] = "ic_ref"
        ws ["MX1"] = "ic_pre"
        ws ["MY1"] = "ic_pref_pre"
        ws ["MZ1"] = "frcpl_ref"
        ws ["NA1"] = "frcpl_pre"
        ws ["NB1"] = "frcpl_pref_pre"
        ws ["NC1"] = "erv_ref"
        ws ["ND1"] = "erv_pre"
        ws ["NE1"] = "erv_pref_pre"
        ws ["NF1"] = "rv_ref"
        ws ["NG1"] = "rv_pre"
        ws ["NH1"] = "rv_pref_pre"
        ws ["NI1"] = "rvdtlc_ref"
        ws ["NJ1"] = "rvdtlc_pre"
        ws ["NK1"] = "vtg_pre"
        ws ["NL1"] = "vt_pre"
        ws ["NM1"] = "dlco_ref"
        ws ["NN1"] = "dlco_pre"
        ws ["NO1"] = "dlco_pref_pre"
        ws ["NP1"] = "dladj_ref"
        ws ["NQ1"] = "dladj_pre"
        ws ["NR1"] = "dladj_pref_pre"
        ws ["NS1"] = "dlcodva_ref"
        ws ["NT1"] = "dlcodva_pre"
        ws ["NU1"] = "dlcodva_pref_pre"
        ws ["NV1"] = "dldvaadj_ref"
        ws ["NW1"] = "dldvaadj_pre"
        ws ["NX1"] = "dldvaadj_pref_pre"
        ws ["NY1"] = "va_pre"
        ws ["NZ1"] = "ivc_pre"
        ws ["OA1"] = "dlcoecode_pre"
        ws ["OB1"] = "rawtotal_pre"
        ws ["OC1"] = "rawinsp_pre"
        ws ["OD1"] = "rawexp_pre"
        ws ["OE1"] = "raw_ref"
        ws ["OF1"] = "raw_pre"
        ws ["OG1"] = "raw_pref_pre"
        ws ["OH1"] = "gaw_ref"
        ws ["OI1"] = "gaw_pre"
        ws ["OJ1"] = "gaw_pref_pre"
        ws ["OK1"] = "sraw_ref"
        ws ["OL1"] = "sraw_pre"
        ws ["OM1"] = "sraw_pref_pre"
        ws ["ON1"] = "sgaw_ref"
        ws ["OO1"] = "sgaw_pre"
        ws ["OP1"] = "sgaw_pref_pre"
        ws ["OQ1"] = "rawvtg_pre"
        ws ["OR1"] = "rawf_pre"
        wb.save("OOMII.xlsx")
    else:
        wb = openpyxl.load_workbook('OOMII.xlsx')
        ws = wb["OOMII_DB"]
    return ws, wb


def new_g_var() :
    g_var = {}
    g_var['img_type'] = ""
    g_var['img_pid'] = ""
    g_var['img_date'] = ""
    g_var['img_age'] = ""
    g_var['img_height'] = ""
    g_var['img_weight'] = ""
    g_var['img_gender'] = ""
    g_var['img_fvc_dose_lv1'] = ""
    g_var['img_fvc_dose_lv2'] = ""
    g_var['img_fvc_dose_lv3'] = ""
    g_var['img_fvc_dose_lv4'] = ""
    g_var['img_fvc_dose_lv5'] = ""
    g_var['img_fvc_dose_lv6'] = ""
    g_var['img_fvc_dose_lv7'] = ""
    g_var['img_fvc_dose_lv8'] = ""
    g_var['img_fvc_dose_lv9'] = ""
    g_var['img_fvc_ref'] = ""
    g_var['img_fvc_pre'] = ""
    g_var['img_fvc_post'] = ""
    g_var['img_fvc_lv1'] = ""
    g_var['img_fvc_lv2'] = ""
    g_var['img_fvc_lv3'] = ""
    g_var['img_fvc_lv4'] = ""
    g_var['img_fvc_lv5'] = ""
    g_var['img_fvc_lv6'] = ""
    g_var['img_fvc_lv7'] = ""
    g_var['img_fvc_lv8'] = ""
    g_var['img_fvc_lv9'] = ""
    g_var['img_fvc_tri1'] = ""
    g_var['img_fvc_tri2'] = ""
    g_var['img_fvc_tri3'] = ""
    g_var['img_fvc_tri4'] = ""
    g_var['img_fvc_tri5'] = ""
    g_var['img_fvc_tri6'] = ""
    g_var['img_fvc_tri7'] = ""
    g_var['img_fvc_tri8'] = ""
    g_var['img_fvc_pref_pre'] = ""
    g_var['img_fvc_pref_post'] = ""
    g_var['img_fvc_pref_lv1'] = ""
    g_var['img_fvc_pref_lv2'] = ""
    g_var['img_fvc_pref_lv3'] = ""
    g_var['img_fvc_pref_lv4'] = ""
    g_var['img_fvc_pref_lv5'] = ""
    g_var['img_fvc_pref_lv6'] = ""
    g_var['img_fvc_pref_lv7'] = ""
    g_var['img_fvc_pref_lv8'] = ""
    g_var['img_fvc_pref_lv9'] = ""
    g_var['img_fvc_pchg'] = ""
    g_var['img_fvc_pchg_lv1'] = ""
    g_var['img_fvc_pchg_lv2'] = ""
    g_var['img_fvc_pchg_lv3'] = ""
    g_var['img_fvc_pchg_lv4'] = ""
    g_var['img_fvc_pchg_lv5'] = ""
    g_var['img_fvc_pchg_lv6'] = ""
    g_var['img_fvc_pchg_lv7'] = ""
    g_var['img_fvc_pchg_lv8'] = ""
    g_var['img_fvc_pchg_lv9'] = ""
    g_var['img_fvc_pred'] = ""
    g_var['img_fvc_pre_ppred'] = ""
    g_var['img_fvc_post_ppred'] = ""
    g_var['img_fev05_ref'] = ""
    g_var['img_fev05_pre'] = ""
    g_var['img_fev05_post'] = ""
    g_var['img_fev05_pref_pre'] = ""
    g_var['img_fev05_pref_post'] = ""
    g_var['img_fev05_pchg'] = ""
    g_var['img_fev1_dose_lv1'] = ""
    g_var['img_fev1_dose_lv2'] = ""
    g_var['img_fev1_dose_lv3'] = ""
    g_var['img_fev1_dose_lv4'] = ""
    g_var['img_fev1_dose_lv5'] = ""
    g_var['img_fev1_dose_lv6'] = ""
    g_var['img_fev1_dose_lv7'] = ""
    g_var['img_fev1_dose_lv8'] = ""
    g_var['img_fev1_dose_lv9'] = ""
    g_var['img_fev1_ref'] = ""
    g_var['img_fev1_pre'] = ""
    g_var['img_fev1_post'] = ""
    g_var['img_fev1_lv1'] = ""
    g_var['img_fev1_lv2'] = ""
    g_var['img_fev1_lv3'] = ""
    g_var['img_fev1_lv4'] = ""
    g_var['img_fev1_lv5'] = ""
    g_var['img_fev1_lv6'] = ""
    g_var['img_fev1_lv7'] = ""
    g_var['img_fev1_lv8'] = ""
    g_var['img_fev1_lv9'] = ""
    g_var['img_fev1_tri1'] = ""
    g_var['img_fev1_tri2'] = ""
    g_var['img_fev1_tri3'] = ""
    g_var['img_fev1_tri4'] = ""
    g_var['img_fev1_tri5'] = ""
    g_var['img_fev1_tri6'] = ""
    g_var['img_fev1_tri7'] = ""
    g_var['img_fev1_tri8'] = ""
    g_var['img_fev1_pref_pre'] = ""
    g_var['img_fev1_pref_post'] = ""
    g_var['img_fev1_pref_lv1'] = ""
    g_var['img_fev1_pref_lv2'] = ""
    g_var['img_fev1_pref_lv3'] = ""
    g_var['img_fev1_pref_lv4'] = ""
    g_var['img_fev1_pref_lv5'] = ""
    g_var['img_fev1_pref_lv6'] = ""
    g_var['img_fev1_pref_lv7'] = ""
    g_var['img_fev1_pref_lv8'] = ""
    g_var['img_fev1_pref_lv9'] = ""
    g_var['img_fev1_pchg'] = ""
    g_var['img_fev1_pchg_lv1'] = ""
    g_var['img_fev1_pchg_lv2'] = ""
    g_var['img_fev1_pchg_lv3'] = ""
    g_var['img_fev1_pchg_lv4'] = ""
    g_var['img_fev1_pchg_lv5'] = ""
    g_var['img_fev1_pchg_lv6'] = ""
    g_var['img_fev1_pchg_lv7'] = ""
    g_var['img_fev1_pchg_lv8'] = ""
    g_var['img_fev1_pchg_lv9'] = ""
    g_var['img_fev1_pred'] = ""
    g_var['img_fev1_pre_ppred'] = ""
    g_var['img_fev1_post_ppred'] = ""
    g_var['img_pc_fev1'] = ""
    g_var['img_fev1_pc'] = ""
    g_var['img_fev1dfvc_ref'] = ""
    g_var['img_fev1dfvc_pre'] = ""
    g_var['img_fev1dfvc_post'] = ""
    g_var['img_fev1dfvc_pred'] = ""
    g_var['img_fev1dfvc_tri1'] = ""
    g_var['img_fev1dfvc_tri2'] = ""
    g_var['img_fev1dfvc_tri3'] = ""
    g_var['img_fev1dfvc_tri4'] = ""
    g_var['img_fev1dfvc_tri5'] = ""
    g_var['img_fev1dfvc_tri6'] = ""
    g_var['img_fev1dfvc_tri7'] = ""
    g_var['img_fev1dfvc_tri8'] = ""
    g_var['img_fev3dfvc_ref'] = ""
    g_var['img_fev3dfvc_pre'] = ""
    g_var['img_fev3dfvc_post'] = ""
    g_var['img_fev3dfvc_pref_pre'] = ""
    g_var['img_fev3dfvc_pref_post'] = ""
    g_var['img_fev3dfvc_pchg'] = ""
    g_var['img_fef25_75_dose_lv1'] = ""
    g_var['img_fef25_75_dose_lv2'] = ""
    g_var['img_fef25_75_dose_lv3'] = ""
    g_var['img_fef25_75_dose_lv4'] = ""
    g_var['img_fef25_75_dose_lv5'] = ""
    g_var['img_fef25_75_dose_lv6'] = ""
    g_var['img_fef25_75_dose_lv7'] = ""
    g_var['img_fef25_75_dose_lv8'] = ""
    g_var['img_fef25_75_dose_lv9'] = ""
    g_var['img_fef25_75_ref'] = ""
    g_var['img_fef25_75_pre'] = ""
    g_var['img_fef25_75_post'] = ""
    g_var['img_fef25_75_lv1'] = ""
    g_var['img_fef25_75_lv2'] = ""
    g_var['img_fef25_75_lv3'] = ""
    g_var['img_fef25_75_lv4'] = ""
    g_var['img_fef25_75_lv5'] = ""
    g_var['img_fef25_75_lv6'] = ""
    g_var['img_fef25_75_lv7'] = ""
    g_var['img_fef25_75_lv8'] = ""
    g_var['img_fef25_75_lv9'] = ""
    g_var['img_fef25_75_tri1'] = ""
    g_var['img_fef25_75_tri2'] = ""
    g_var['img_fef25_75_tri3'] = ""
    g_var['img_fef25_75_tri4'] = ""
    g_var['img_fef25_75_tri5'] = ""
    g_var['img_fef25_75_tri6'] = ""
    g_var['img_fef25_75_tri7'] = ""
    g_var['img_fef25_75_tri8'] = ""
    g_var['img_fef25_75_pref_pre'] = ""
    g_var['img_fef25_75_pref_post'] = ""
    g_var['img_fef25_75_pref_lv1'] = ""
    g_var['img_fef25_75_pref_lv2'] = ""
    g_var['img_fef25_75_pref_lv3'] = ""
    g_var['img_fef25_75_pref_lv4'] = ""
    g_var['img_fef25_75_pref_lv5'] = ""
    g_var['img_fef25_75_pref_lv6'] = ""
    g_var['img_fef25_75_pref_lv7'] = ""
    g_var['img_fef25_75_pref_lv8'] = ""
    g_var['img_fef25_75_pref_lv9'] = ""
    g_var['img_fef25_75_pchg'] = ""
    g_var['img_fef25_75_pchg_lv1'] = ""
    g_var['img_fef25_75_pchg_lv2'] = ""
    g_var['img_fef25_75_pchg_lv3'] = ""
    g_var['img_fef25_75_pchg_lv4'] = ""
    g_var['img_fef25_75_pchg_lv5'] = ""
    g_var['img_fef25_75_pchg_lv6'] = ""
    g_var['img_fef25_75_pchg_lv7'] = ""
    g_var['img_fef25_75_pchg_lv8'] = ""
    g_var['img_fef25_75_pchg_lv9'] = ""
    g_var['img_fef25_75_pred'] = ""
    g_var['img_fef25_75_pre_ppred'] = ""
    g_var['img_fef25_75_post_ppred'] = ""
    g_var['img_isofef25_75_pchg'] = ""
    g_var['img_isofef25_75_pred'] = ""
    g_var['img_isofef25_75_pre'] = ""
    g_var['img_isofef25_75_pre_ppred'] = ""
    g_var['img_isofef25_75_post'] = ""
    g_var['img_isofef25_75_post_ppred'] = ""
    g_var['img_fef75_85_ref'] = ""
    g_var['img_fef75_85_pre'] = ""
    g_var['img_fef75_85_post'] = ""
    g_var['img_fef75_85_pref_pre'] = ""
    g_var['img_fef75_85_pref_post'] = ""
    g_var['img_fef75_85_pchg'] = ""
    g_var['img_fef75_85_pred'] = ""
    g_var['img_fef75_85_pre_ppred'] = ""
    g_var['img_fef75_85_post_ppred'] = ""
    g_var['img_fef25_ref'] = ""
    g_var['img_fef25_pre'] = ""
    g_var['img_fef25_post'] = ""
    g_var['img_fef25_pref_pre'] = ""
    g_var['img_fef25_pref_post'] = ""
    g_var['img_fef25_pchg'] = ""
    g_var['img_fef50_ref'] = ""
    g_var['img_fef50_pre'] = ""
    g_var['img_fef50_post'] = ""
    g_var['img_fef50_pref_pre'] = ""
    g_var['img_fef50_pref_post'] = ""
    g_var['img_fef50_pchg'] = ""
    g_var['img_fef75_ref'] = ""
    g_var['img_fef75_pre'] = ""
    g_var['img_fef75_post'] = ""
    g_var['img_fef75_pref_pre'] = ""
    g_var['img_fef75_pref_post'] = ""
    g_var['img_fef75_pchg'] = ""
    g_var['img_fef200_1200_ref'] = ""
    g_var['img_fef200_1200_pre'] = ""
    g_var['img_fef200_1200_post'] = ""
    g_var['img_fef200_1200_pref_pre'] = ""
    g_var['img_fef200_1200_pref_post'] = ""
    g_var['img_fef200_1200_pchg'] = ""
    g_var['img_pef_dose_lv1'] = ""
    g_var['img_pef_dose_lv2'] = ""
    g_var['img_pef_dose_lv3'] = ""
    g_var['img_pef_dose_lv4'] = ""
    g_var['img_pef_dose_lv5'] = ""
    g_var['img_pef_dose_lv6'] = ""
    g_var['img_pef_dose_lv7'] = ""
    g_var['img_pef_dose_lv8'] = ""
    g_var['img_pef_dose_lv9'] = ""
    g_var['img_pef_ref'] = ""
    g_var['img_pef_pre'] = ""
    g_var['img_pef_post'] = ""
    g_var['img_pef_lv1'] = ""
    g_var['img_pef_lv2'] = ""
    g_var['img_pef_lv3'] = ""
    g_var['img_pef_lv4'] = ""
    g_var['img_pef_lv5'] = ""
    g_var['img_pef_lv6'] = ""
    g_var['img_pef_lv7'] = ""
    g_var['img_pef_lv8'] = ""
    g_var['img_pef_lv9'] = ""
    g_var['img_pef_tri1'] = ""
    g_var['img_pef_tri2'] = ""
    g_var['img_pef_tri3'] = ""
    g_var['img_pef_tri4'] = ""
    g_var['img_pef_tri5'] = ""
    g_var['img_pef_tri6'] = ""
    g_var['img_pef_tri7'] = ""
    g_var['img_pef_tri8'] = ""
    g_var['img_pef_pref_pre'] = ""
    g_var['img_pef_pref_post'] = ""
    g_var['img_pef_pref_lv1'] = ""
    g_var['img_pef_pref_lv2'] = ""
    g_var['img_pef_pref_lv3'] = ""
    g_var['img_pef_pref_lv4'] = ""
    g_var['img_pef_pref_lv5'] = ""
    g_var['img_pef_pref_lv6'] = ""
    g_var['img_pef_pref_lv7'] = ""
    g_var['img_pef_pref_lv8'] = ""
    g_var['img_pef_pref_lv9'] = ""
    g_var['img_pef_pchg'] = ""
    g_var['img_pef_pchg_lv1'] = ""
    g_var['img_pef_pchg_lv2'] = ""
    g_var['img_pef_pchg_lv3'] = ""
    g_var['img_pef_pchg_lv4'] = ""
    g_var['img_pef_pchg_lv5'] = ""
    g_var['img_pef_pchg_lv6'] = ""
    g_var['img_pef_pchg_lv7'] = ""
    g_var['img_pef_pchg_lv8'] = ""
    g_var['img_pef_pchg_lv9'] = ""
    g_var['img_pef_pred'] = ""
    g_var['img_pef_pre_ppred'] = ""
    g_var['img_pef_post_ppred'] = ""
    g_var['img_peft_pchg'] = ""
    g_var['img_peft_pre'] = ""
    g_var['img_peft_post'] = ""
    g_var['img_peft_tri1'] = ""
    g_var['img_peft_tri2'] = ""
    g_var['img_peft_tri3'] = ""
    g_var['img_peft_tri4'] = ""
    g_var['img_peft_tri5'] = ""
    g_var['img_peft_tri6'] = ""
    g_var['img_peft_tri7'] = ""
    g_var['img_peft_tri8'] = ""
    g_var['img_fet100_pchg'] = ""
    g_var['img_fet100_pre'] = ""
    g_var['img_fet100_post'] = ""
    g_var['img_fet100_tri1'] = ""
    g_var['img_fet100_tri2'] = ""
    g_var['img_fet100_tri3'] = ""
    g_var['img_fet100_tri4'] = ""
    g_var['img_fet100_tri5'] = ""
    g_var['img_fet100_tri6'] = ""
    g_var['img_fet100_tri7'] = ""
    g_var['img_fet100_tri8'] = ""
    g_var['img_fivc_ref'] = ""
    g_var['img_fivc_pre'] = ""
    g_var['img_fivc_post'] = ""
    g_var['img_fivc_pref_pre'] = ""
    g_var['img_fivc_pref_post'] = ""
    g_var['img_fivc_pchg'] = ""
    g_var['img_fivc_pred'] = ""
    g_var['img_fivc_pre_ppred'] = ""
    g_var['img_fivc_post_ppred'] = ""
    g_var['img_fivc_tri1'] = ""
    g_var['img_fivc_tri2'] = ""
    g_var['img_fivc_tri3'] = ""
    g_var['img_fivc_tri4'] = ""
    g_var['img_fivc_tri5'] = ""
    g_var['img_fivc_tri6'] = ""
    g_var['img_fivc_tri7'] = ""
    g_var['img_fivc_tri8'] = ""
    g_var['img_fiv1_pchg'] = ""
    g_var['img_fiv1_pre'] = ""
    g_var['img_fiv1_post'] = ""
    g_var['img_fefdfif50_pchg'] = ""
    g_var['img_fefdfif50_pred'] = ""
    g_var['img_fefdfif50_pre'] = ""
    g_var['img_fefdfif50_pre_ppred'] = ""
    g_var['img_fefdfif50_post'] = ""
    g_var['img_fefdfif50_post_ppred'] = ""
    g_var['img_volextrap_pchg'] = ""
    g_var['img_volextrap_pre'] = ""
    g_var['img_volextrap_post'] = ""
    g_var['img_volextrap_tri1'] = ""
    g_var['img_volextrap_tri2'] = ""
    g_var['img_volextrap_tri3'] = ""
    g_var['img_volextrap_tri4'] = ""
    g_var['img_volextrap_tri5'] = ""
    g_var['img_volextrap_tri6'] = ""
    g_var['img_volextrap_tri7'] = ""
    g_var['img_volextrap_tri8'] = ""
    g_var['img_fvlecode_pre'] = ""
    g_var['img_fvlecode_post'] = ""
    g_var['img_fvlecode_tri1'] = ""
    g_var['img_fvlecode_tri2'] = ""
    g_var['img_fvlecode_tri3'] = ""
    g_var['img_fvlecode_tri4'] = ""
    g_var['img_fvlecode_tri5'] = ""
    g_var['img_fvlecode_tri6'] = ""
    g_var['img_fvlecode_tri7'] = ""
    g_var['img_fvlecode_tri8'] = ""
    g_var['img_mvv_pred'] = ""
    g_var['img_tlc_ref'] = ""
    g_var['img_tlc_pre'] = ""
    g_var['img_tlc_pref_pre'] = ""
    g_var['img_vc_ref'] = ""
    g_var['img_vc_pre'] = ""
    g_var['img_vc_pref_pre'] = ""
    g_var['img_ic_ref'] = ""
    g_var['img_ic_pre'] = ""
    g_var['img_ic_pref_pre'] = ""
    g_var['img_frcpl_ref'] = ""
    g_var['img_frcpl_pre'] = ""
    g_var['img_frcpl_pref_pre'] = ""
    g_var['img_erv_ref'] = ""
    g_var['img_erv_pre'] = ""
    g_var['img_erv_pref_pre'] = ""
    g_var['img_rv_ref'] = ""
    g_var['img_rv_pre'] = ""
    g_var['img_rv_pref_pre'] = ""
    g_var['img_rvdtlc_ref'] = ""
    g_var['img_rvdtlc_pre'] = ""
    g_var['img_vtg_pre'] = ""
    g_var['img_vt_pre'] = ""
    g_var['img_dlco_ref'] = ""
    g_var['img_dlco_pre'] = ""
    g_var['img_dlco_pref_pre'] = ""
    g_var['img_dladj_ref'] = ""
    g_var['img_dladj_pre'] = ""
    g_var['img_dladj_pref_pre'] = ""
    g_var['img_dlcodva_ref'] = ""
    g_var['img_dlcodva_pre'] = ""
    g_var['img_dlcodva_pref_pre'] = ""
    g_var['img_dldvaadj_ref'] = ""
    g_var['img_dldvaadj_pre'] = ""
    g_var['img_dldvaadj_pref_pre'] = ""
    g_var['img_va_pre'] = ""
    g_var['img_ivc_pre'] = ""
    g_var['img_dlcoecode_pre'] = ""
    g_var['img_rawtotal_pre'] = ""
    g_var['img_rawinsp_pre'] = ""
    g_var['img_rawexp_pre'] = ""
    g_var['img_raw_ref'] = ""
    g_var['img_raw_pre'] = ""
    g_var['img_raw_pref_pre'] = ""
    g_var['img_gaw_ref'] = ""
    g_var['img_gaw_pre'] = ""
    g_var['img_gaw_pref_pre'] = ""
    g_var['img_sraw_ref'] = ""
    g_var['img_sraw_pre'] = ""
    g_var['img_sraw_pref_pre'] = ""
    g_var['img_sgaw_ref'] = ""
    g_var['img_sgaw_pre'] = ""
    g_var['img_sgaw_pref_pre'] = ""
    g_var['img_rawvtg_pre'] = ""
    g_var['img_rawf_pre'] = ""
    return g_var


def append_g_var_to_excel(ws, wb, g_var):
    ws_cnt = str(ws.max_row + 1)
    print ("ws_cnt: " + ws_cnt)
    ws ["A" + ws_cnt] = g_var['img_type']
    ws ["B" + ws_cnt] = g_var['img_pid']
    ws ["C" + ws_cnt] = g_var['img_date']
    ws ["D" + ws_cnt] = g_var['img_age']
    ws ["E" + ws_cnt] = g_var['img_height']
    ws ["F" + ws_cnt] = g_var['img_weight']
    ws ["G" + ws_cnt] = g_var['img_gender']
    ws ["H" + ws_cnt] = g_var['img_fvc_dose_lv1']
    ws ["I" + ws_cnt] = g_var['img_fvc_dose_lv2']
    ws ["J" + ws_cnt] = g_var['img_fvc_dose_lv3']
    ws ["K" + ws_cnt] = g_var['img_fvc_dose_lv4']
    ws ["L" + ws_cnt] = g_var['img_fvc_dose_lv5']
    ws ["M" + ws_cnt] = g_var['img_fvc_dose_lv6']
    ws ["N" + ws_cnt] = g_var['img_fvc_dose_lv7']
    ws ["O" + ws_cnt] = g_var['img_fvc_dose_lv8']
    ws ["P" + ws_cnt] = g_var['img_fvc_dose_lv9']
    ws ["Q" + ws_cnt] = g_var['img_fvc_ref']
    ws ["R" + ws_cnt] = g_var['img_fvc_pre']
    ws ["S" + ws_cnt] = g_var['img_fvc_post']
    ws ["T" + ws_cnt] = g_var['img_fvc_lv1']
    ws ["U" + ws_cnt] = g_var['img_fvc_lv2']
    ws ["V" + ws_cnt] = g_var['img_fvc_lv3']
    ws ["W" + ws_cnt] = g_var['img_fvc_lv4']
    ws ["X" + ws_cnt] = g_var['img_fvc_lv5']
    ws ["Y" + ws_cnt] = g_var['img_fvc_lv6']
    ws ["Z" + ws_cnt] = g_var['img_fvc_lv7']
    ws ["AA" + ws_cnt] = g_var['img_fvc_lv8']
    ws ["AB" + ws_cnt] = g_var['img_fvc_lv9']
    ws ["AC" + ws_cnt] = g_var['img_fvc_tri1']
    ws ["AD" + ws_cnt] = g_var['img_fvc_tri2']
    ws ["AE" + ws_cnt] = g_var['img_fvc_tri3']
    ws ["AF" + ws_cnt] = g_var['img_fvc_tri4']
    ws ["AG" + ws_cnt] = g_var['img_fvc_tri5']
    ws ["AH" + ws_cnt] = g_var['img_fvc_tri6']
    ws ["AI" + ws_cnt] = g_var['img_fvc_tri7']
    ws ["AJ" + ws_cnt] = g_var['img_fvc_tri8']
    ws ["AK" + ws_cnt] = g_var['img_fvc_pref_pre']
    ws ["AL" + ws_cnt] = g_var['img_fvc_pref_post']
    ws ["AM" + ws_cnt] = g_var['img_fvc_pref_lv1']
    ws ["AN" + ws_cnt] = g_var['img_fvc_pref_lv2']
    ws ["AO" + ws_cnt] = g_var['img_fvc_pref_lv3']
    ws ["AP" + ws_cnt] = g_var['img_fvc_pref_lv4']
    ws ["AQ" + ws_cnt] = g_var['img_fvc_pref_lv5']
    ws ["AR" + ws_cnt] = g_var['img_fvc_pref_lv6']
    ws ["AS" + ws_cnt] = g_var['img_fvc_pref_lv7']
    ws ["AT" + ws_cnt] = g_var['img_fvc_pref_lv8']
    ws ["AU" + ws_cnt] = g_var['img_fvc_pref_lv9']
    ws ["AV" + ws_cnt] = g_var['img_fvc_pchg']
    ws ["AW" + ws_cnt] = g_var['img_fvc_pchg_lv1']
    ws ["AX" + ws_cnt] = g_var['img_fvc_pchg_lv2']
    ws ["AY" + ws_cnt] = g_var['img_fvc_pchg_lv3']
    ws ["AZ" + ws_cnt] = g_var['img_fvc_pchg_lv4']
    ws ["BA" + ws_cnt] = g_var['img_fvc_pchg_lv5']
    ws ["BB" + ws_cnt] = g_var['img_fvc_pchg_lv6']
    ws ["BC" + ws_cnt] = g_var['img_fvc_pchg_lv7']
    ws ["BD" + ws_cnt] = g_var['img_fvc_pchg_lv8']
    ws ["BE" + ws_cnt] = g_var['img_fvc_pchg_lv9']
    ws ["BF" + ws_cnt] = g_var['img_fvc_pred']
    ws ["BG" + ws_cnt] = g_var['img_fvc_pre_ppred']
    ws ["BH" + ws_cnt] = g_var['img_fvc_post_ppred']
    ws ["BI" + ws_cnt] = g_var['img_fev05_ref']
    ws ["BJ" + ws_cnt] = g_var['img_fev05_pre']
    ws ["BK" + ws_cnt] = g_var['img_fev05_post']
    ws ["BL" + ws_cnt] = g_var['img_fev05_pref_pre']
    ws ["BM" + ws_cnt] = g_var['img_fev05_pref_post']
    ws ["BN" + ws_cnt] = g_var['img_fev05_pchg']
    ws ["BO" + ws_cnt] = g_var['img_fev1_dose_lv1']
    ws ["BP" + ws_cnt] = g_var['img_fev1_dose_lv2']
    ws ["BQ" + ws_cnt] = g_var['img_fev1_dose_lv3']
    ws ["BR" + ws_cnt] = g_var['img_fev1_dose_lv4']
    ws ["BS" + ws_cnt] = g_var['img_fev1_dose_lv5']
    ws ["BT" + ws_cnt] = g_var['img_fev1_dose_lv6']
    ws ["BU" + ws_cnt] = g_var['img_fev1_dose_lv7']
    ws ["BV" + ws_cnt] = g_var['img_fev1_dose_lv8']
    ws ["BW" + ws_cnt] = g_var['img_fev1_dose_lv9']
    ws ["BX" + ws_cnt] = g_var['img_fev1_ref']
    ws ["BY" + ws_cnt] = g_var['img_fev1_pre']
    ws ["BZ" + ws_cnt] = g_var['img_fev1_post']
    ws ["CA" + ws_cnt] = g_var['img_fev1_lv1']
    ws ["CB" + ws_cnt] = g_var['img_fev1_lv2']
    ws ["CC" + ws_cnt] = g_var['img_fev1_lv3']
    ws ["CD" + ws_cnt] = g_var['img_fev1_lv4']
    ws ["CE" + ws_cnt] = g_var['img_fev1_lv5']
    ws ["CF" + ws_cnt] = g_var['img_fev1_lv6']
    ws ["CG" + ws_cnt] = g_var['img_fev1_lv7']
    ws ["CH" + ws_cnt] = g_var['img_fev1_lv8']
    ws ["CI" + ws_cnt] = g_var['img_fev1_lv9']
    ws ["CJ" + ws_cnt] = g_var['img_fev1_tri1']
    ws ["CK" + ws_cnt] = g_var['img_fev1_tri2']
    ws ["CL" + ws_cnt] = g_var['img_fev1_tri3']
    ws ["CM" + ws_cnt] = g_var['img_fev1_tri4']
    ws ["CN" + ws_cnt] = g_var['img_fev1_tri5']
    ws ["CO" + ws_cnt] = g_var['img_fev1_tri6']
    ws ["CP" + ws_cnt] = g_var['img_fev1_tri7']
    ws ["CQ" + ws_cnt] = g_var['img_fev1_tri8']
    ws ["CR" + ws_cnt] = g_var['img_fev1_pref_pre']
    ws ["CS" + ws_cnt] = g_var['img_fev1_pref_post']
    ws ["CT" + ws_cnt] = g_var['img_fev1_pref_lv1']
    ws ["CU" + ws_cnt] = g_var['img_fev1_pref_lv2']
    ws ["CV" + ws_cnt] = g_var['img_fev1_pref_lv3']
    ws ["CW" + ws_cnt] = g_var['img_fev1_pref_lv4']
    ws ["CX" + ws_cnt] = g_var['img_fev1_pref_lv5']
    ws ["CY" + ws_cnt] = g_var['img_fev1_pref_lv6']
    ws ["CZ" + ws_cnt] = g_var['img_fev1_pref_lv7']
    ws ["DA" + ws_cnt] = g_var['img_fev1_pref_lv8']
    ws ["DB" + ws_cnt] = g_var['img_fev1_pref_lv9']
    ws ["DC" + ws_cnt] = g_var['img_fev1_pchg']
    ws ["DD" + ws_cnt] = g_var['img_fev1_pchg_lv1']
    ws ["DE" + ws_cnt] = g_var['img_fev1_pchg_lv2']
    ws ["DF" + ws_cnt] = g_var['img_fev1_pchg_lv3']
    ws ["DG" + ws_cnt] = g_var['img_fev1_pchg_lv4']
    ws ["DH" + ws_cnt] = g_var['img_fev1_pchg_lv5']
    ws ["DI" + ws_cnt] = g_var['img_fev1_pchg_lv6']
    ws ["DJ" + ws_cnt] = g_var['img_fev1_pchg_lv7']
    ws ["DK" + ws_cnt] = g_var['img_fev1_pchg_lv8']
    ws ["DL" + ws_cnt] = g_var['img_fev1_pchg_lv9']
    ws ["DM" + ws_cnt] = g_var['img_fev1_pred']
    ws ["DN" + ws_cnt] = g_var['img_fev1_pre_ppred']
    ws ["DO" + ws_cnt] = g_var['img_fev1_post_ppred']
    ws ["DP" + ws_cnt] = g_var['img_pc_fev1']
    ws ["DQ" + ws_cnt] = g_var['img_fev1_pc']
    ws ["DR" + ws_cnt] = g_var['img_fev1dfvc_ref']
    ws ["DS" + ws_cnt] = g_var['img_fev1dfvc_pre']
    ws ["DT" + ws_cnt] = g_var['img_fev1dfvc_post']
    ws ["DU" + ws_cnt] = g_var['img_fev1dfvc_pred']
    ws ["DV" + ws_cnt] = g_var['img_fev1dfvc_tri1']
    ws ["DW" + ws_cnt] = g_var['img_fev1dfvc_tri2']
    ws ["DX" + ws_cnt] = g_var['img_fev1dfvc_tri3']
    ws ["DY" + ws_cnt] = g_var['img_fev1dfvc_tri4']
    ws ["DZ" + ws_cnt] = g_var['img_fev1dfvc_tri5']
    ws ["EA" + ws_cnt] = g_var['img_fev1dfvc_tri6']
    ws ["EB" + ws_cnt] = g_var['img_fev1dfvc_tri7']
    ws ["EC" + ws_cnt] = g_var['img_fev1dfvc_tri8']
    ws ["ED" + ws_cnt] = g_var['img_fev3dfvc_ref']
    ws ["EE" + ws_cnt] = g_var['img_fev3dfvc_pre']
    ws ["EF" + ws_cnt] = g_var['img_fev3dfvc_post']
    ws ["EG" + ws_cnt] = g_var['img_fev3dfvc_pref_pre']
    ws ["EH" + ws_cnt] = g_var['img_fev3dfvc_pref_post']
    ws ["EI" + ws_cnt] = g_var['img_fev3dfvc_pchg']
    ws ["EJ" + ws_cnt] = g_var['img_fef25_75_dose_lv1']
    ws ["EK" + ws_cnt] = g_var['img_fef25_75_dose_lv2']
    ws ["EL" + ws_cnt] = g_var['img_fef25_75_dose_lv3']
    ws ["EM" + ws_cnt] = g_var['img_fef25_75_dose_lv4']
    ws ["EN" + ws_cnt] = g_var['img_fef25_75_dose_lv5']
    ws ["EO" + ws_cnt] = g_var['img_fef25_75_dose_lv6']
    ws ["EP" + ws_cnt] = g_var['img_fef25_75_dose_lv7']
    ws ["EQ" + ws_cnt] = g_var['img_fef25_75_dose_lv8']
    ws ["ER" + ws_cnt] = g_var['img_fef25_75_dose_lv9']
    ws ["ES" + ws_cnt] = g_var['img_fef25_75_ref']
    ws ["ET" + ws_cnt] = g_var['img_fef25_75_pre']
    ws ["EU" + ws_cnt] = g_var['img_fef25_75_post']
    ws ["EV" + ws_cnt] = g_var['img_fef25_75_lv1']
    ws ["EW" + ws_cnt] = g_var['img_fef25_75_lv2']
    ws ["EX" + ws_cnt] = g_var['img_fef25_75_lv3']
    ws ["EY" + ws_cnt] = g_var['img_fef25_75_lv4']
    ws ["EZ" + ws_cnt] = g_var['img_fef25_75_lv5']
    ws ["FA" + ws_cnt] = g_var['img_fef25_75_lv6']
    ws ["FB" + ws_cnt] = g_var['img_fef25_75_lv7']
    ws ["FC" + ws_cnt] = g_var['img_fef25_75_lv8']
    ws ["FD" + ws_cnt] = g_var['img_fef25_75_lv9']
    ws ["FE" + ws_cnt] = g_var['img_fef25_75_tri1']
    ws ["FF" + ws_cnt] = g_var['img_fef25_75_tri2']
    ws ["FG" + ws_cnt] = g_var['img_fef25_75_tri3']
    ws ["FH" + ws_cnt] = g_var['img_fef25_75_tri4']
    ws ["FI" + ws_cnt] = g_var['img_fef25_75_tri5']
    ws ["FJ" + ws_cnt] = g_var['img_fef25_75_tri6']
    ws ["FK" + ws_cnt] = g_var['img_fef25_75_tri7']
    ws ["FL" + ws_cnt] = g_var['img_fef25_75_tri8']
    ws ["FM" + ws_cnt] = g_var['img_fef25_75_pref_pre']
    ws ["FN" + ws_cnt] = g_var['img_fef25_75_pref_post']
    ws ["FO" + ws_cnt] = g_var['img_fef25_75_pref_lv1']
    ws ["FP" + ws_cnt] = g_var['img_fef25_75_pref_lv2']
    ws ["FQ" + ws_cnt] = g_var['img_fef25_75_pref_lv3']
    ws ["FR" + ws_cnt] = g_var['img_fef25_75_pref_lv4']
    ws ["FS" + ws_cnt] = g_var['img_fef25_75_pref_lv5']
    ws ["FT" + ws_cnt] = g_var['img_fef25_75_pref_lv6']
    ws ["FU" + ws_cnt] = g_var['img_fef25_75_pref_lv7']
    ws ["FV" + ws_cnt] = g_var['img_fef25_75_pref_lv8']
    ws ["FW" + ws_cnt] = g_var['img_fef25_75_pref_lv9']
    ws ["FX" + ws_cnt] = g_var['img_fef25_75_pchg']
    ws ["FY" + ws_cnt] = g_var['img_fef25_75_pchg_lv1']
    ws ["FZ" + ws_cnt] = g_var['img_fef25_75_pchg_lv2']
    ws ["GA" + ws_cnt] = g_var['img_fef25_75_pchg_lv3']
    ws ["GB" + ws_cnt] = g_var['img_fef25_75_pchg_lv4']
    ws ["GC" + ws_cnt] = g_var['img_fef25_75_pchg_lv5']
    ws ["GD" + ws_cnt] = g_var['img_fef25_75_pchg_lv6']
    ws ["GE" + ws_cnt] = g_var['img_fef25_75_pchg_lv7']
    ws ["GF" + ws_cnt] = g_var['img_fef25_75_pchg_lv8']
    ws ["GG" + ws_cnt] = g_var['img_fef25_75_pchg_lv9']
    ws ["GH" + ws_cnt] = g_var['img_fef25_75_pred']
    ws ["GI" + ws_cnt] = g_var['img_fef25_75_pre_ppred']
    ws ["GJ" + ws_cnt] = g_var['img_fef25_75_post_ppred']
    ws ["GK" + ws_cnt] = g_var['img_isofef25_75_pchg']
    ws ["GL" + ws_cnt] = g_var['img_isofef25_75_pred']
    ws ["GM" + ws_cnt] = g_var['img_isofef25_75_pre']
    ws ["GN" + ws_cnt] = g_var['img_isofef25_75_pre_ppred']
    ws ["GO" + ws_cnt] = g_var['img_isofef25_75_post']
    ws ["GP" + ws_cnt] = g_var['img_isofef25_75_post_ppred']
    ws ["GQ" + ws_cnt] = g_var['img_fef75_85_ref']
    ws ["GR" + ws_cnt] = g_var['img_fef75_85_pre']
    ws ["GS" + ws_cnt] = g_var['img_fef75_85_post']
    ws ["GT" + ws_cnt] = g_var['img_fef75_85_pref_pre']
    ws ["GU" + ws_cnt] = g_var['img_fef75_85_pref_post']
    ws ["GV" + ws_cnt] = g_var['img_fef75_85_pchg']
    ws ["GW" + ws_cnt] = g_var['img_fef75_85_pred']
    ws ["GX" + ws_cnt] = g_var['img_fef75_85_pre_ppred']
    ws ["GY" + ws_cnt] = g_var['img_fef75_85_post_ppred']
    ws ["GZ" + ws_cnt] = g_var['img_fef25_ref']
    ws ["HA" + ws_cnt] = g_var['img_fef25_pre']
    ws ["HB" + ws_cnt] = g_var['img_fef25_post']
    ws ["HC" + ws_cnt] = g_var['img_fef25_pref_pre']
    ws ["HD" + ws_cnt] = g_var['img_fef25_pref_post']
    ws ["HE" + ws_cnt] = g_var['img_fef25_pchg']
    ws ["HF" + ws_cnt] = g_var['img_fef50_ref']
    ws ["HG" + ws_cnt] = g_var['img_fef50_pre']
    ws ["HH" + ws_cnt] = g_var['img_fef50_post']
    ws ["HI" + ws_cnt] = g_var['img_fef50_pref_pre']
    ws ["HJ" + ws_cnt] = g_var['img_fef50_pref_post']
    ws ["HK" + ws_cnt] = g_var['img_fef50_pchg']
    ws ["HL" + ws_cnt] = g_var['img_fef75_ref']
    ws ["HM" + ws_cnt] = g_var['img_fef75_pre']
    ws ["HN" + ws_cnt] = g_var['img_fef75_post']
    ws ["HO" + ws_cnt] = g_var['img_fef75_pref_pre']
    ws ["HP" + ws_cnt] = g_var['img_fef75_pref_post']
    ws ["HQ" + ws_cnt] = g_var['img_fef75_pchg']
    ws ["HR" + ws_cnt] = g_var['img_fef200_1200_ref']
    ws ["HS" + ws_cnt] = g_var['img_fef200_1200_pre']
    ws ["HT" + ws_cnt] = g_var['img_fef200_1200_post']
    ws ["HU" + ws_cnt] = g_var['img_fef200_1200_pref_pre']
    ws ["HV" + ws_cnt] = g_var['img_fef200_1200_pref_post']
    ws ["HW" + ws_cnt] = g_var['img_fef200_1200_pchg']
    ws ["HX" + ws_cnt] = g_var['img_pef_dose_lv1']
    ws ["HY" + ws_cnt] = g_var['img_pef_dose_lv2']
    ws ["HZ" + ws_cnt] = g_var['img_pef_dose_lv3']
    ws ["IA" + ws_cnt] = g_var['img_pef_dose_lv4']
    ws ["IB" + ws_cnt] = g_var['img_pef_dose_lv5']
    ws ["IC" + ws_cnt] = g_var['img_pef_dose_lv6']
    ws ["ID" + ws_cnt] = g_var['img_pef_dose_lv7']
    ws ["IE" + ws_cnt] = g_var['img_pef_dose_lv8']
    ws ["IF" + ws_cnt] = g_var['img_pef_dose_lv9']
    ws ["IG" + ws_cnt] = g_var['img_pef_ref']
    ws ["IH" + ws_cnt] = g_var['img_pef_pre']
    ws ["II" + ws_cnt] = g_var['img_pef_post']
    ws ["IJ" + ws_cnt] = g_var['img_pef_lv1']
    ws ["IK" + ws_cnt] = g_var['img_pef_lv2']
    ws ["IL" + ws_cnt] = g_var['img_pef_lv3']
    ws ["IM" + ws_cnt] = g_var['img_pef_lv4']
    ws ["IN" + ws_cnt] = g_var['img_pef_lv5']
    ws ["IO" + ws_cnt] = g_var['img_pef_lv6']
    ws ["IP" + ws_cnt] = g_var['img_pef_lv7']
    ws ["IQ" + ws_cnt] = g_var['img_pef_lv8']
    ws ["IR" + ws_cnt] = g_var['img_pef_lv9']
    ws ["IS" + ws_cnt] = g_var['img_pef_tri1']
    ws ["IT" + ws_cnt] = g_var['img_pef_tri2']
    ws ["IU" + ws_cnt] = g_var['img_pef_tri3']
    ws ["IV" + ws_cnt] = g_var['img_pef_tri4']
    ws ["IW" + ws_cnt] = g_var['img_pef_tri5']
    ws ["IX" + ws_cnt] = g_var['img_pef_tri6']
    ws ["IY" + ws_cnt] = g_var['img_pef_tri7']
    ws ["IZ" + ws_cnt] = g_var['img_pef_tri8']
    ws ["JA" + ws_cnt] = g_var['img_pef_pref_pre']
    ws ["JB" + ws_cnt] = g_var['img_pef_pref_post']
    ws ["JC" + ws_cnt] = g_var['img_pef_pref_lv1']
    ws ["JD" + ws_cnt] = g_var['img_pef_pref_lv2']
    ws ["JE" + ws_cnt] = g_var['img_pef_pref_lv3']
    ws ["JF" + ws_cnt] = g_var['img_pef_pref_lv4']
    ws ["JG" + ws_cnt] = g_var['img_pef_pref_lv5']
    ws ["JH" + ws_cnt] = g_var['img_pef_pref_lv6']
    ws ["JI" + ws_cnt] = g_var['img_pef_pref_lv7']
    ws ["JJ" + ws_cnt] = g_var['img_pef_pref_lv8']
    ws ["JK" + ws_cnt] = g_var['img_pef_pref_lv9']
    ws ["JL" + ws_cnt] = g_var['img_pef_pchg']
    ws ["JM" + ws_cnt] = g_var['img_pef_pchg_lv1']
    ws ["JN" + ws_cnt] = g_var['img_pef_pchg_lv2']
    ws ["JO" + ws_cnt] = g_var['img_pef_pchg_lv3']
    ws ["JP" + ws_cnt] = g_var['img_pef_pchg_lv4']
    ws ["JQ" + ws_cnt] = g_var['img_pef_pchg_lv5']
    ws ["JR" + ws_cnt] = g_var['img_pef_pchg_lv6']
    ws ["JS" + ws_cnt] = g_var['img_pef_pchg_lv7']
    ws ["JT" + ws_cnt] = g_var['img_pef_pchg_lv8']
    ws ["JU" + ws_cnt] = g_var['img_pef_pchg_lv9']
    ws ["JV" + ws_cnt] = g_var['img_pef_pred']
    ws ["JW" + ws_cnt] = g_var['img_pef_pre_ppred']
    ws ["JX" + ws_cnt] = g_var['img_pef_post_ppred']
    ws ["JY" + ws_cnt] = g_var['img_peft_pchg']
    ws ["JZ" + ws_cnt] = g_var['img_peft_pre']
    ws ["KA" + ws_cnt] = g_var['img_peft_post']
    ws ["KB" + ws_cnt] = g_var['img_peft_tri1']
    ws ["KC" + ws_cnt] = g_var['img_peft_tri2']
    ws ["KD" + ws_cnt] = g_var['img_peft_tri3']
    ws ["KE" + ws_cnt] = g_var['img_peft_tri4']
    ws ["KF" + ws_cnt] = g_var['img_peft_tri5']
    ws ["KG" + ws_cnt] = g_var['img_peft_tri6']
    ws ["KH" + ws_cnt] = g_var['img_peft_tri7']
    ws ["KI" + ws_cnt] = g_var['img_peft_tri8']
    ws ["KJ" + ws_cnt] = g_var['img_fet100_pchg']
    ws ["KK" + ws_cnt] = g_var['img_fet100_pre']
    ws ["KL" + ws_cnt] = g_var['img_fet100_post']
    ws ["KM" + ws_cnt] = g_var['img_fet100_tri1']
    ws ["KN" + ws_cnt] = g_var['img_fet100_tri2']
    ws ["KO" + ws_cnt] = g_var['img_fet100_tri3']
    ws ["KP" + ws_cnt] = g_var['img_fet100_tri4']
    ws ["KQ" + ws_cnt] = g_var['img_fet100_tri5']
    ws ["KR" + ws_cnt] = g_var['img_fet100_tri6']
    ws ["KS" + ws_cnt] = g_var['img_fet100_tri7']
    ws ["KT" + ws_cnt] = g_var['img_fet100_tri8']
    ws ["KU" + ws_cnt] = g_var['img_fivc_ref']
    ws ["KV" + ws_cnt] = g_var['img_fivc_pre']
    ws ["KW" + ws_cnt] = g_var['img_fivc_post']
    ws ["KX" + ws_cnt] = g_var['img_fivc_pref_pre']
    ws ["KY" + ws_cnt] = g_var['img_fivc_pref_post']
    ws ["KZ" + ws_cnt] = g_var['img_fivc_pchg']
    ws ["LA" + ws_cnt] = g_var['img_fivc_pred']
    ws ["LB" + ws_cnt] = g_var['img_fivc_pre_ppred']
    ws ["LC" + ws_cnt] = g_var['img_fivc_post_ppred']
    ws ["LD" + ws_cnt] = g_var['img_fivc_tri1']
    ws ["LE" + ws_cnt] = g_var['img_fivc_tri2']
    ws ["LF" + ws_cnt] = g_var['img_fivc_tri3']
    ws ["LG" + ws_cnt] = g_var['img_fivc_tri4']
    ws ["LH" + ws_cnt] = g_var['img_fivc_tri5']
    ws ["LI" + ws_cnt] = g_var['img_fivc_tri6']
    ws ["LJ" + ws_cnt] = g_var['img_fivc_tri7']
    ws ["LK" + ws_cnt] = g_var['img_fivc_tri8']
    ws ["LL" + ws_cnt] = g_var['img_fiv1_pchg']
    ws ["LM" + ws_cnt] = g_var['img_fiv1_pre']
    ws ["LN" + ws_cnt] = g_var['img_fiv1_post']
    ws ["LO" + ws_cnt] = g_var['img_fefdfif50_pchg']
    ws ["LP" + ws_cnt] = g_var['img_fefdfif50_pred']
    ws ["LQ" + ws_cnt] = g_var['img_fefdfif50_pre']
    ws ["LR" + ws_cnt] = g_var['img_fefdfif50_pre_ppred']
    ws ["LS" + ws_cnt] = g_var['img_fefdfif50_post']
    ws ["LT" + ws_cnt] = g_var['img_fefdfif50_post_ppred']
    ws ["LU" + ws_cnt] = g_var['img_volextrap_pchg']
    ws ["LV" + ws_cnt] = g_var['img_volextrap_pre']
    ws ["LW" + ws_cnt] = g_var['img_volextrap_post']
    ws ["LX" + ws_cnt] = g_var['img_volextrap_tri1']
    ws ["LY" + ws_cnt] = g_var['img_volextrap_tri2']
    ws ["LZ" + ws_cnt] = g_var['img_volextrap_tri3']
    ws ["MA" + ws_cnt] = g_var['img_volextrap_tri4']
    ws ["MB" + ws_cnt] = g_var['img_volextrap_tri5']
    ws ["MC" + ws_cnt] = g_var['img_volextrap_tri6']
    ws ["MD" + ws_cnt] = g_var['img_volextrap_tri7']
    ws ["ME" + ws_cnt] = g_var['img_volextrap_tri8']
    ws ["MF" + ws_cnt] = g_var['img_fvlecode_pre']
    ws ["MG" + ws_cnt] = g_var['img_fvlecode_post']
    ws ["MH" + ws_cnt] = g_var['img_fvlecode_tri1']
    ws ["MI" + ws_cnt] = g_var['img_fvlecode_tri2']
    ws ["MJ" + ws_cnt] = g_var['img_fvlecode_tri3']
    ws ["MK" + ws_cnt] = g_var['img_fvlecode_tri4']
    ws ["ML" + ws_cnt] = g_var['img_fvlecode_tri5']
    ws ["MM" + ws_cnt] = g_var['img_fvlecode_tri6']
    ws ["MN" + ws_cnt] = g_var['img_fvlecode_tri7']
    ws ["MO" + ws_cnt] = g_var['img_fvlecode_tri8']
    ws ["MP" + ws_cnt] = g_var['img_mvv_pred']
    ws ["MQ" + ws_cnt] = g_var['img_tlc_ref']
    ws ["MR" + ws_cnt] = g_var['img_tlc_pre']
    ws ["MS" + ws_cnt] = g_var['img_tlc_pref_pre']
    ws ["MT" + ws_cnt] = g_var['img_vc_ref']
    ws ["MU" + ws_cnt] = g_var['img_vc_pre']
    ws ["MV" + ws_cnt] = g_var['img_vc_pref_pre']
    ws ["MW" + ws_cnt] = g_var['img_ic_ref']
    ws ["MX" + ws_cnt] = g_var['img_ic_pre']
    ws ["MY" + ws_cnt] = g_var['img_ic_pref_pre']
    ws ["MZ" + ws_cnt] = g_var['img_frcpl_ref']
    ws ["NA" + ws_cnt] = g_var['img_frcpl_pre']
    ws ["NB" + ws_cnt] = g_var['img_frcpl_pref_pre']
    ws ["NC" + ws_cnt] = g_var['img_erv_ref']
    ws ["ND" + ws_cnt] = g_var['img_erv_pre']
    ws ["NE" + ws_cnt] = g_var['img_erv_pref_pre']
    ws ["NF" + ws_cnt] = g_var['img_rv_ref']
    ws ["NG" + ws_cnt] = g_var['img_rv_pre']
    ws ["NH" + ws_cnt] = g_var['img_rv_pref_pre']
    ws ["NI" + ws_cnt] = g_var['img_rvdtlc_ref']
    ws ["NJ" + ws_cnt] = g_var['img_rvdtlc_pre']
    ws ["NK" + ws_cnt] = g_var['img_vtg_pre']
    ws ["NL" + ws_cnt] = g_var['img_vt_pre']
    ws ["NM" + ws_cnt] = g_var['img_dlco_ref']
    ws ["NN" + ws_cnt] = g_var['img_dlco_pre']
    ws ["NO" + ws_cnt] = g_var['img_dlco_pref_pre']
    ws ["NP" + ws_cnt] = g_var['img_dladj_ref']
    ws ["NQ" + ws_cnt] = g_var['img_dladj_pre']
    ws ["NR" + ws_cnt] = g_var['img_dladj_pref_pre']
    ws ["NS" + ws_cnt] = g_var['img_dlcodva_ref']
    ws ["NT" + ws_cnt] = g_var['img_dlcodva_pre']
    ws ["NU" + ws_cnt] = g_var['img_dlcodva_pref_pre']
    ws ["NV" + ws_cnt] = g_var['img_dldvaadj_ref']
    ws ["NW" + ws_cnt] = g_var['img_dldvaadj_pre']
    ws ["NX" + ws_cnt] = g_var['img_dldvaadj_pref_pre']
    ws ["NY" + ws_cnt] = g_var['img_va_pre']
    ws ["NZ" + ws_cnt] = g_var['img_ivc_pre']
    ws ["OA" + ws_cnt] = g_var['img_dlcoecode_pre']
    ws ["OB" + ws_cnt] = g_var['img_rawtotal_pre']
    ws ["OC" + ws_cnt] = g_var['img_rawinsp_pre']
    ws ["OD" + ws_cnt] = g_var['img_rawexp_pre']
    ws ["OE" + ws_cnt] = g_var['img_raw_ref']
    ws ["OF" + ws_cnt] = g_var['img_raw_pre']
    ws ["OG" + ws_cnt] = g_var['img_raw_pref_pre']
    ws ["OH" + ws_cnt] = g_var['img_gaw_ref']
    ws ["OI" + ws_cnt] = g_var['img_gaw_pre']
    ws ["OJ" + ws_cnt] = g_var['img_gaw_pref_pre']
    ws ["OK" + ws_cnt] = g_var['img_sraw_ref']
    ws ["OL" + ws_cnt] = g_var['img_sraw_pre']
    ws ["OM" + ws_cnt] = g_var['img_sraw_pref_pre']
    ws ["ON" + ws_cnt] = g_var['img_sgaw_ref']
    ws ["OO" + ws_cnt] = g_var['img_sgaw_pre']
    ws ["OP" + ws_cnt] = g_var['img_sgaw_pref_pre']
    ws ["OQ" + ws_cnt] = g_var['img_rawvtg_pre']
    ws ["OR" + ws_cnt] = g_var['img_rawf_pre']
    wb.save("OOMII.xlsx")
    return


if __name__ == '__main__':
    file = open('result.txt', mode='wt', encoding='utf-8')
    file.write('0')
    file.close()
    if len(sys.argv) > 1: 
        paths = []
        for i in range(1, len(sys.argv)):
            if sys.argv[i].find('path=') > -1:
                sp = sys.argv[i].split('=')
                if len(sp) == 2:
                    sp = sp[1].split('*')
                    for path in sp:
                        paths.append(path)

            elif sys.argv[i].find('option=') > -1:
                sp = sys.argv[i].split('=')
                if len(sp) == 2:
                    option_params = sp[1].split('*')
                    for option in option_params:
                        options.append(option)
            else:
                print('invalid argument: ' + sys.argv[i])
        process(paths)
        file = open('result.txt', mode='wt', encoding='utf-8')
        file.write('1')
        file.close()
    else:
        print('NEED ARGUMENT FOR IMAGE FILE PATHS OR DIR NAME')