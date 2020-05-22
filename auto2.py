#!/usr/bin/env python
# coding: utf-8

import os
import sys
from PyQt5.QtWidgets import *
import cv2
import numpy as np
import glob
import pytesseract
import openpyxl

pytesseract.pytesseract.tesseract_cmd = 'E:/Tesseract-OCR/tesseract.exe'
class MyApp(QWidget):
    
    global ori_path, new_path
    ori_path = new_path = ""

    def __init__(self):
        global ori_path, new_path
        super().__init__()
        self.initUI()
        ori_path = "C:/inz/before"
        self.label1.setText(ori_path)
        new_path = "C:/inz/after"
        self.label2.setText(new_path)
        self.pushButtonForOpenImageFileClicked();

    def initUI(self):
        self.setGeometry(800, 200, 300, 300)
        self.setWindowTitle("OOMII :)")
        
        self.pushButton1 = QPushButton("Set Original Image Directory")
        self.pushButton1.clicked.connect(self.pushButtonForOriImagePathClicked)
        self.label1 = QLabel()
        
        self.pushButton2 = QPushButton("Set New Image Directory")
        self.pushButton2.clicked.connect(self.pushButtonForNewImagePathClicked)
        self.label2 = QLabel()
        
        self.pushButton3 = QPushButton("Classify PFT Sheets")
        self.pushButton3.clicked.connect(self.pushButtonForOpenImageFileClicked)
        self.label3 = QLabel()

        layout = QVBoxLayout()
        
        layout.addWidget(self.pushButton1)
        layout.addWidget(self.label1)
        
        layout.addWidget(self.pushButton2)
        layout.addWidget(self.label2)
        
        layout.addWidget(self.pushButton3)
        layout.addWidget(self.label3)

        self.setLayout(layout)
        
                
    def pushButtonForOriImagePathClicked(self):
        global ori_path
        
        ori_path = QFileDialog.getExistingDirectory(self)
        self.label1.setText(ori_path)
        
        
    def pushButtonForNewImagePathClicked(self):
        global new_path
        
        new_path = QFileDialog.getExistingDirectory(self)
        self.label2.setText(new_path)

    def pushButtonForOpenImageFileClicked(self):
        global new_path, wb, ws
 
        ori_image_files = glob.glob(ori_path + "/*.jpg")
        ori_image_files_counter = 0
        type01_img_cnt = 0
        type02_img_cnt = 0
        type03_img_cnt = 0
        type04_img_cnt = 0
        type05_img_cnt = 0
        type06_img_cnt = 0
        type07_img_cnt = 0
        type08_img_cnt = 0
        typeUK_img_cnt = 0

        # resize image
        def resizing(mod_img):
            return cv2.resize(mod_img, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)

        # get grayscale image
        def grayscaling(mod_img):
            return cv2.cvtColor(mod_img, cv2.COLOR_BGR2GRAY)

        # create new excel file 
        if os.path.isfile(new_path + '/OOMII.xlsx') == 0:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws = wb.create_sheet("OOMII_DB", 0)
            ws ["A1"] = "type"
            ws ["B1"] = "pid"
            ws ["C1"] = "date"
            ws ["D1"] = "age"
            ws ["E1"] = "height"
            ws ["F1"] = "weight"
            ws ["G1"] = "gender"
            ws ["H1"] = "fvc_dose_lv1"
            ws ["I1"] = "fvc_dose_lv2"
            ws ["J1"] = "fvc_dose_lv3"
            ws ["K1"] = "fvc_dose_lv4"
            ws ["L1"] = "fvc_dose_lv5"
            ws ["M1"] = "fvc_dose_lv6"
            ws ["N1"] = "fvc_dose_lv7"
            ws ["O1"] = "fvc_dose_lv8"
            ws ["P1"] = "fvc_dose_lv9"
            ws ["Q1"] = "fvc_ref"
            ws ["R1"] = "fvc_pre"
            ws ["S1"] = "fvc_post"
            ws ["T1"] = "fvc_lv1"
            ws ["U1"] = "fvc_lv2"
            ws ["V1"] = "fvc_lv3"
            ws ["W1"] = "fvc_lv4"
            ws ["X1"] = "fvc_lv5"
            ws ["Y1"] = "fvc_lv6"
            ws ["Z1"] = "fvc_lv7"
            ws ["AA1"] = "fvc_lv8"
            ws ["AB1"] = "fvc_lv9"
            ws ["AC1"] = "fvc_tri1"
            ws ["AD1"] = "fvc_tri2"
            ws ["AE1"] = "fvc_tri3"
            ws ["AF1"] = "fvc_tri4"
            ws ["AG1"] = "fvc_tri5"
            ws ["AH1"] = "fvc_tri6"
            ws ["AI1"] = "fvc_tri7"
            ws ["AJ1"] = "fvc_tri8"
            ws ["AK1"] = "fvc_pref_pre"
            ws ["AL1"] = "fvc_pref_post"
            ws ["AM1"] = "fvc_pref_lv1"
            ws ["AN1"] = "fvc_pref_lv2"
            ws ["AO1"] = "fvc_pref_lv3"
            ws ["AP1"] = "fvc_pref_lv4"
            ws ["AQ1"] = "fvc_pref_lv5"
            ws ["AR1"] = "fvc_pref_lv6"
            ws ["AS1"] = "fvc_pref_lv7"
            ws ["AT1"] = "fvc_pref_lv8"
            ws ["AU1"] = "fvc_pref_lv9"
            ws ["AV1"] = "fvc_pchg"
            ws ["AW1"] = "fvc_pchg_lv1"
            ws ["AX1"] = "fvc_pchg_lv2"
            ws ["AY1"] = "fvc_pchg_lv3"
            ws ["AZ1"] = "fvc_pchg_lv4"
            ws ["BA1"] = "fvc_pchg_lv5"
            ws ["BB1"] = "fvc_pchg_lv6"
            ws ["BC1"] = "fvc_pchg_lv7"
            ws ["BD1"] = "fvc_pchg_lv8"
            ws ["BE1"] = "fvc_pchg_lv9"
            ws ["BF1"] = "fvc_pred"
            ws ["BG1"] = "fvc_pre_ppred"
            ws ["BH1"] = "fvc_post_ppred"
            ws ["BI1"] = "fev05_ref"
            ws ["BJ1"] = "fev05_pre"
            ws ["BK1"] = "fev05_post"
            ws ["BL1"] = "fev05_pref_pre"
            ws ["BM1"] = "fev05_pref_post"
            ws ["BN1"] = "fev05_pchg"
            ws ["BO1"] = "fev1_dose_lv1"
            ws ["BP1"] = "fev1_dose_lv2"
            ws ["BQ1"] = "fev1_dose_lv3"
            ws ["BR1"] = "fev1_dose_lv4"
            ws ["BS1"] = "fev1_dose_lv5"
            ws ["BT1"] = "fev1_dose_lv6"
            ws ["BU1"] = "fev1_dose_lv7"
            ws ["BV1"] = "fev1_dose_lv8"
            ws ["BW1"] = "fev1_dose_lv9"
            ws ["BX1"] = "fev1_ref"
            ws ["BY1"] = "fev1_pre"
            ws ["BZ1"] = "fev1_post"
            ws ["CA1"] = "fev1_lv1"
            ws ["CB1"] = "fev1_lv2"
            ws ["CC1"] = "fev1_lv3"
            ws ["CD1"] = "fev1_lv4"
            ws ["CE1"] = "fev1_lv5"
            ws ["CF1"] = "fev1_lv6"
            ws ["CG1"] = "fev1_lv7"
            ws ["CH1"] = "fev1_lv8"
            ws ["CI1"] = "fev1_lv9"
            ws ["CJ1"] = "fev1_tri1"
            ws ["CK1"] = "fev1_tri2"
            ws ["CL1"] = "fev1_tri3"
            ws ["CM1"] = "fev1_tri4"
            ws ["CN1"] = "fev1_tri5"
            ws ["CO1"] = "fev1_tri6"
            ws ["CP1"] = "fev1_tri7"
            ws ["CQ1"] = "fev1_tri8"
            ws ["CR1"] = "fev1_pref_pre"
            ws ["CS1"] = "fev1_pref_post"
            ws ["CT1"] = "fev1_pref_lv1"
            ws ["CU1"] = "fev1_pref_lv2"
            ws ["CV1"] = "fev1_pref_lv3"
            ws ["CW1"] = "fev1_pref_lv4"
            ws ["CX1"] = "fev1_pref_lv5"
            ws ["CY1"] = "fev1_pref_lv6"
            ws ["CZ1"] = "fev1_pref_lv7"
            ws ["DA1"] = "fev1_pref_lv8"
            ws ["DB1"] = "fev1_pref_lv9"
            ws ["DC1"] = "fev1_pchg"
            ws ["DD1"] = "fev1_pchg_lv1"
            ws ["DE1"] = "fev1_pchg_lv2"
            ws ["DF1"] = "fev1_pchg_lv3"
            ws ["DG1"] = "fev1_pchg_lv4"
            ws ["DH1"] = "fev1_pchg_lv5"
            ws ["DI1"] = "fev1_pchg_lv6"
            ws ["DJ1"] = "fev1_pchg_lv7"
            ws ["DK1"] = "fev1_pchg_lv8"
            ws ["DL1"] = "fev1_pchg_lv9"
            ws ["DM1"] = "fev1_pred"
            ws ["DN1"] = "fev1_pre_ppred"
            ws ["DO1"] = "fev1_post_ppred"
            ws ["DP1"] = "pc_fev1"
            ws ["DQ1"] = "fev1_pc"
            ws ["DR1"] = "fev1dfvc_ref"
            ws ["DS1"] = "fev1dfvc_pre"
            ws ["DT1"] = "fev1dfvc_post"
            ws ["DU1"] = "fev1dfvc_pred"
            ws ["DV1"] = "fev1dfvc_tri1"
            ws ["DW1"] = "fev1dfvc_tri2"
            ws ["DX1"] = "fev1dfvc_tri3"
            ws ["DY1"] = "fev1dfvc_tri4"
            ws ["DZ1"] = "fev1dfvc_tri5"
            ws ["EA1"] = "fev1dfvc_tri6"
            ws ["EB1"] = "fev1dfvc_tri7"
            ws ["EC1"] = "fev1dfvc_tri8"
            ws ["ED1"] = "fev3dfvc_ref"
            ws ["EE1"] = "fev3dfvc_pre"
            ws ["EF1"] = "fev3dfvc_post"
            ws ["EG1"] = "fev3dfvc_pref_pre"
            ws ["EH1"] = "fev3dfvc_pref_post"
            ws ["EI1"] = "fev3dfvc_pchg"
            ws ["EJ1"] = "fef25_75_dose_lv1"
            ws ["EK1"] = "fef25_75_dose_lv2"
            ws ["EL1"] = "fef25_75_dose_lv3"
            ws ["EM1"] = "fef25_75_dose_lv4"
            ws ["EN1"] = "fef25_75_dose_lv5"
            ws ["EO1"] = "fef25_75_dose_lv6"
            ws ["EP1"] = "fef25_75_dose_lv7"
            ws ["EQ1"] = "fef25_75_dose_lv8"
            ws ["ER1"] = "fef25_75_dose_lv9"
            ws ["ES1"] = "fef25_75_ref"
            ws ["ET1"] = "fef25_75_pre"
            ws ["EU1"] = "fef25_75_post"
            ws ["EV1"] = "fef25_75_lv1"
            ws ["EW1"] = "fef25_75_lv2"
            ws ["EX1"] = "fef25_75_lv3"
            ws ["EY1"] = "fef25_75_lv4"
            ws ["EZ1"] = "fef25_75_lv5"
            ws ["FA1"] = "fef25_75_lv6"
            ws ["FB1"] = "fef25_75_lv7"
            ws ["FC1"] = "fef25_75_lv8"
            ws ["FD1"] = "fef25_75_lv9"
            ws ["FE1"] = "fef25_75_tri1"
            ws ["FF1"] = "fef25_75_tri2"
            ws ["FG1"] = "fef25_75_tri3"
            ws ["FH1"] = "fef25_75_tri4"
            ws ["FI1"] = "fef25_75_tri5"
            ws ["FJ1"] = "fef25_75_tri6"
            ws ["FK1"] = "fef25_75_tri7"
            ws ["FL1"] = "fef25_75_tri8"
            ws ["FM1"] = "fef25_75_pref_pre"
            ws ["FN1"] = "fef25_75_pref_post"
            ws ["FO1"] = "fef25_75_pref_lv1"
            ws ["FP1"] = "fef25_75_pref_lv2"
            ws ["FQ1"] = "fef25_75_pref_lv3"
            ws ["FR1"] = "fef25_75_pref_lv4"
            ws ["FS1"] = "fef25_75_pref_lv5"
            ws ["FT1"] = "fef25_75_pref_lv6"
            ws ["FU1"] = "fef25_75_pref_lv7"
            ws ["FV1"] = "fef25_75_pref_lv8"
            ws ["FW1"] = "fef25_75_pref_lv9"
            ws ["FX1"] = "fef25_75_pchg"
            ws ["FY1"] = "fef25_75_pchg_lv1"
            ws ["FZ1"] = "fef25_75_pchg_lv2"
            ws ["GA1"] = "fef25_75_pchg_lv3"
            ws ["GB1"] = "fef25_75_pchg_lv4"
            ws ["GC1"] = "fef25_75_pchg_lv5"
            ws ["GD1"] = "fef25_75_pchg_lv6"
            ws ["GE1"] = "fef25_75_pchg_lv7"
            ws ["GF1"] = "fef25_75_pchg_lv8"
            ws ["GG1"] = "fef25_75_pchg_lv9"
            ws ["GH1"] = "fef25_75_pred"
            ws ["GI1"] = "fef25_75_pre_ppred"
            ws ["GJ1"] = "fef25_75_post_ppred"
            ws ["GK1"] = "isofef25_75_pchg"
            ws ["GL1"] = "isofef25_75_pred"
            ws ["GM1"] = "isofef25_75_pre"
            ws ["GN1"] = "isofef25_75_pre_ppred"
            ws ["GO1"] = "isofef25_75_post"
            ws ["GP1"] = "isofef25_75_post_ppred"
            ws ["GQ1"] = "fef75_85_ref"
            ws ["GR1"] = "fef75_85_pre"
            ws ["GS1"] = "fef75_85_post"
            ws ["GT1"] = "fef75_85_pref_pre"
            ws ["GU1"] = "fef75_85_pref_post"
            ws ["GV1"] = "fef75_85_pchg"
            ws ["GW1"] = "fef75_85_pred"
            ws ["GX1"] = "fef75_85_pre_ppred"
            ws ["GY1"] = "fef75_85_post_ppred"
            ws ["GZ1"] = "fef25_ref"
            ws ["HA1"] = "fef25_pre"
            ws ["HB1"] = "fef25_post"
            ws ["HC1"] = "fef25_pref_pre"
            ws ["HD1"] = "fef25_pref_post"
            ws ["HE1"] = "fef25_pchg"
            ws ["HF1"] = "fef50_ref"
            ws ["HG1"] = "fef50_pre"
            ws ["HH1"] = "fef50_post"
            ws ["HI1"] = "fef50_pref_pre"
            ws ["HJ1"] = "fef50_pref_post"
            ws ["HK1"] = "fef50_pchg"
            ws ["HL1"] = "fef75_ref"
            ws ["HM1"] = "fef75_pre"
            ws ["HN1"] = "fef75_post"
            ws ["HO1"] = "fef75_pref_pre"
            ws ["HP1"] = "fef75_pref_post"
            ws ["HQ1"] = "fef75_pchg"
            ws ["HR1"] = "fef200_1200_ref"
            ws ["HS1"] = "fef200_1200_pre"
            ws ["HT1"] = "fef200_1200_post"
            ws ["HU1"] = "fef200_1200_pref_pre"
            ws ["HV1"] = "fef200_1200_pref_post"
            ws ["HW1"] = "fef200_1200_pchg"
            ws ["HX1"] = "pef_dose_lv1"
            ws ["HY1"] = "pef_dose_lv2"
            ws ["HZ1"] = "pef_dose_lv3"
            ws ["IA1"] = "pef_dose_lv4"
            ws ["IB1"] = "pef_dose_lv5"
            ws ["IC1"] = "pef_dose_lv6"
            ws ["ID1"] = "pef_dose_lv7"
            ws ["IE1"] = "pef_dose_lv8"
            ws ["IF1"] = "pef_dose_lv9"
            ws ["IG1"] = "pef_ref"
            ws ["IH1"] = "pef_pre"
            ws ["II1"] = "pef_post"
            ws ["IJ1"] = "pef_lv1"
            ws ["IK1"] = "pef_lv2"
            ws ["IL1"] = "pef_lv3"
            ws ["IM1"] = "pef_lv4"
            ws ["IN1"] = "pef_lv5"
            ws ["IO1"] = "pef_lv6"
            ws ["IP1"] = "pef_lv7"
            ws ["IQ1"] = "pef_lv8"
            ws ["IR1"] = "pef_lv9"
            ws ["IS1"] = "pef_tri1"
            ws ["IT1"] = "pef_tri2"
            ws ["IU1"] = "pef_tri3"
            ws ["IV1"] = "pef_tri4"
            ws ["IW1"] = "pef_tri5"
            ws ["IX1"] = "pef_tri6"
            ws ["IY1"] = "pef_tri7"
            ws ["IZ1"] = "pef_tri8"
            ws ["JA1"] = "pef_pref_pre"
            ws ["JB1"] = "pef_pref_post"
            ws ["JC1"] = "pef_pref_lv1"
            ws ["JD1"] = "pef_pref_lv2"
            ws ["JE1"] = "pef_pref_lv3"
            ws ["JF1"] = "pef_pref_lv4"
            ws ["JG1"] = "pef_pref_lv5"
            ws ["JH1"] = "pef_pref_lv6"
            ws ["JI1"] = "pef_pref_lv7"
            ws ["JJ1"] = "pef_pref_lv8"
            ws ["JK1"] = "pef_pref_lv9"
            ws ["JL1"] = "pef_pchg"
            ws ["JM1"] = "pef_pchg_lv1"
            ws ["JN1"] = "pef_pchg_lv2"
            ws ["JO1"] = "pef_pchg_lv3"
            ws ["JP1"] = "pef_pchg_lv4"
            ws ["JQ1"] = "pef_pchg_lv5"
            ws ["JR1"] = "pef_pchg_lv6"
            ws ["JS1"] = "pef_pchg_lv7"
            ws ["JT1"] = "pef_pchg_lv8"
            ws ["JU1"] = "pef_pchg_lv9"
            ws ["JV1"] = "pef_pred"
            ws ["JW1"] = "pef_pre_ppred"
            ws ["JX1"] = "pef_post_ppred"
            ws ["JY1"] = "peft_pchg"
            ws ["JZ1"] = "peft_pre"
            ws ["KA1"] = "peft_post"
            ws ["KB1"] = "peft_tri1"
            ws ["KC1"] = "peft_tri2"
            ws ["KD1"] = "peft_tri3"
            ws ["KE1"] = "peft_tri4"
            ws ["KF1"] = "peft_tri5"
            ws ["KG1"] = "peft_tri6"
            ws ["KH1"] = "peft_tri7"
            ws ["KI1"] = "peft_tri8"
            ws ["KJ1"] = "fet100_pchg"
            ws ["KK1"] = "fet100_pre"
            ws ["KL1"] = "fet100_post"
            ws ["KM1"] = "fet100_tri1"
            ws ["KN1"] = "fet100_tri2"
            ws ["KO1"] = "fet100_tri3"
            ws ["KP1"] = "fet100_tri4"
            ws ["KQ1"] = "fet100_tri5"
            ws ["KR1"] = "fet100_tri6"
            ws ["KS1"] = "fet100_tri7"
            ws ["KT1"] = "fet100_tri8"
            ws ["KU1"] = "fivc_ref"
            ws ["KV1"] = "fivc_pre"
            ws ["KW1"] = "fivc_post"
            ws ["KX1"] = "fivc_pref_pre"
            ws ["KY1"] = "fivc_pref_post"
            ws ["KZ1"] = "fivc_pchg"
            ws ["LA1"] = "fivc_pred"
            ws ["LB1"] = "fivc_pre_ppred"
            ws ["LC1"] = "fivc_post_ppred"
            ws ["LD1"] = "fivc_tri1"
            ws ["LE1"] = "fivc_tri2"
            ws ["LF1"] = "fivc_tri3"
            ws ["LG1"] = "fivc_tri4"
            ws ["LH1"] = "fivc_tri5"
            ws ["LI1"] = "fivc_tri6"
            ws ["LJ1"] = "fivc_tri7"
            ws ["LK1"] = "fivc_tri8"
            ws ["LL1"] = "fiv1_pchg"
            ws ["LM1"] = "fiv1_pre"
            ws ["LN1"] = "fiv1_post"
            ws ["LO1"] = "fefdfif50_pchg"
            ws ["LP1"] = "fefdfif50_pred"
            ws ["LQ1"] = "fefdfif50_pre"
            ws ["LR1"] = "fefdfif50_pre_ppred"
            ws ["LS1"] = "fefdfif50_post"
            ws ["LT1"] = "fefdfif50_post_ppred"
            ws ["LU1"] = "volextrap_pchg"
            ws ["LV1"] = "volextrap_pre"
            ws ["LW1"] = "volextrap_post"
            ws ["LX1"] = "volextrap_tri1"
            ws ["LY1"] = "volextrap_tri2"
            ws ["LZ1"] = "volextrap_tri3"
            ws ["MA1"] = "volextrap_tri4"
            ws ["MB1"] = "volextrap_tri5"
            ws ["MC1"] = "volextrap_tri6"
            ws ["MD1"] = "volextrap_tri7"
            ws ["ME1"] = "volextrap_tri8"
            ws ["MF1"] = "fvlecode_pre"
            ws ["MG1"] = "fvlecode_post"
            ws ["MH1"] = "fvlecode_tri1"
            ws ["MI1"] = "fvlecode_tri2"
            ws ["MJ1"] = "fvlecode_tri3"
            ws ["MK1"] = "fvlecode_tri4"
            ws ["ML1"] = "fvlecode_tri5"
            ws ["MM1"] = "fvlecode_tri6"
            ws ["MN1"] = "fvlecode_tri7"
            ws ["MO1"] = "fvlecode_tri8"
            ws ["MP1"] = "mvv_pred"
            ws ["MQ1"] = "tlc_ref"
            ws ["MR1"] = "tlc_pre"
            ws ["MS1"] = "tlc_pref_pre"
            ws ["MT1"] = "vc_ref"
            ws ["MU1"] = "vc_pre"
            ws ["MV1"] = "vc_pref_pre"
            ws ["MW1"] = "ic_ref"
            ws ["MX1"] = "ic_pre"
            ws ["MY1"] = "ic_pref_pre"
            ws ["MZ1"] = "frcpl_ref"
            ws ["NA1"] = "frcpl_pre"
            ws ["NB1"] = "frcpl_pref_pre"
            ws ["NC1"] = "erv_ref"
            ws ["ND1"] = "erv_pre"
            ws ["NE1"] = "erv_pref_pre"
            ws ["NF1"] = "rv_ref"
            ws ["NG1"] = "rv_pre"
            ws ["NH1"] = "rv_pref_pre"
            ws ["NI1"] = "rvdtlc_ref"
            ws ["NJ1"] = "rvdtlc_pre"
            ws ["NK1"] = "vtg_pre"
            ws ["NL1"] = "vt_pre"
            ws ["NM1"] = "dlco_ref"
            ws ["NN1"] = "dlco_pre"
            ws ["NO1"] = "dlco_pref_pre"
            ws ["NP1"] = "dladj_ref"
            ws ["NQ1"] = "dladj_pre"
            ws ["NR1"] = "dladj_pref_pre"
            ws ["NS1"] = "dlcodva_ref"
            ws ["NT1"] = "dlcodva_pre"
            ws ["NU1"] = "dlcodva_pref_pre"
            ws ["NV1"] = "dldvaadj_ref"
            ws ["NW1"] = "dldvaadj_pre"
            ws ["NX1"] = "dldvaadj_pref_pre"
            ws ["NY1"] = "va_pre"
            ws ["NZ1"] = "ivc_pre"
            ws ["OA1"] = "dlcoecode_pre"
            ws ["OB1"] = "rawtotal_pre"
            ws ["OC1"] = "rawinsp_pre"
            ws ["OD1"] = "rawexp_pre"
            ws ["OE1"] = "raw_ref"
            ws ["OF1"] = "raw_pre"
            ws ["OG1"] = "raw_pref_pre"
            ws ["OH1"] = "gaw_ref"
            ws ["OI1"] = "gaw_pre"
            ws ["OJ1"] = "gaw_pref_pre"
            ws ["OK1"] = "sraw_ref"
            ws ["OL1"] = "sraw_pre"
            ws ["OM1"] = "sraw_pref_pre"
            ws ["ON1"] = "sgaw_ref"
            ws ["OO1"] = "sgaw_pre"
            ws ["OP1"] = "sgaw_pref_pre"
            ws ["OQ1"] = "rawvtg_pre"
            ws ["OR1"] = "rawf_pre"
            wb.save(new_path + "/OOMII.xlsx")
            print("Save new OOMII excel file")

        else:
            wb = openpyxl.load_workbook(new_path + '/OOMII.xlsx')
            ws = wb.get_sheet_by_name("OOMII_DB")
            print("Load ori ONMI excel file")

        for img in ori_image_files:
            img_type = ""
            img_pid = ""
            img_date = ""
            img_age = ""
            img_height = ""
            img_weight = ""
            img_gender = ""
            img_fvc_dose_lv1 = ""
            img_fvc_dose_lv2 = ""
            img_fvc_dose_lv3 = ""
            img_fvc_dose_lv4 = ""
            img_fvc_dose_lv5 = ""
            img_fvc_dose_lv6 = ""
            img_fvc_dose_lv7 = ""
            img_fvc_dose_lv8 = ""
            img_fvc_dose_lv9 = ""
            img_fvc_ref = ""
            img_fvc_pre = ""
            img_fvc_post = ""
            img_fvc_lv1 = ""
            img_fvc_lv2 = ""
            img_fvc_lv3 = ""
            img_fvc_lv4 = ""
            img_fvc_lv5 = ""
            img_fvc_lv6 = ""
            img_fvc_lv7 = ""
            img_fvc_lv8 = ""
            img_fvc_lv9 = ""
            img_fvc_tri1 = ""
            img_fvc_tri2 = ""
            img_fvc_tri3 = ""
            img_fvc_tri4 = ""
            img_fvc_tri5 = ""
            img_fvc_tri6 = ""
            img_fvc_tri7 = ""
            img_fvc_tri8 = ""
            img_fvc_pref_pre = ""
            img_fvc_pref_post = ""
            img_fvc_pref_lv1 = ""
            img_fvc_pref_lv2 = ""
            img_fvc_pref_lv3 = ""
            img_fvc_pref_lv4 = ""
            img_fvc_pref_lv5 = ""
            img_fvc_pref_lv6 = ""
            img_fvc_pref_lv7 = ""
            img_fvc_pref_lv8 = ""
            img_fvc_pref_lv9 = ""
            img_fvc_pchg = ""
            img_fvc_pchg_lv1 = ""
            img_fvc_pchg_lv2 = ""
            img_fvc_pchg_lv3 = ""
            img_fvc_pchg_lv4 = ""
            img_fvc_pchg_lv5 = ""
            img_fvc_pchg_lv6 = ""
            img_fvc_pchg_lv7 = ""
            img_fvc_pchg_lv8 = ""
            img_fvc_pchg_lv9 = ""
            img_fvc_pred = ""
            img_fvc_pre_ppred = ""
            img_fvc_post_ppred = ""
            img_fev05_ref = ""
            img_fev05_pre = ""
            img_fev05_post = ""
            img_fev05_pref_pre = ""
            img_fev05_pref_post = ""
            img_fev05_pchg = ""
            img_fev1_dose_lv1 = ""
            img_fev1_dose_lv2 = ""
            img_fev1_dose_lv3 = ""
            img_fev1_dose_lv4 = ""
            img_fev1_dose_lv5 = ""
            img_fev1_dose_lv6 = ""
            img_fev1_dose_lv7 = ""
            img_fev1_dose_lv8 = ""
            img_fev1_dose_lv9 = ""
            img_fev1_ref = ""
            img_fev1_pre = ""
            img_fev1_post = ""
            img_fev1_lv1 = ""
            img_fev1_lv2 = ""
            img_fev1_lv3 = ""
            img_fev1_lv4 = ""
            img_fev1_lv5 = ""
            img_fev1_lv6 = ""
            img_fev1_lv7 = ""
            img_fev1_lv8 = ""
            img_fev1_lv9 = ""
            img_fev1_tri1 = ""
            img_fev1_tri2 = ""
            img_fev1_tri3 = ""
            img_fev1_tri4 = ""
            img_fev1_tri5 = ""
            img_fev1_tri6 = ""
            img_fev1_tri7 = ""
            img_fev1_tri8 = ""
            img_fev1_pref_pre = ""
            img_fev1_pref_post = ""
            img_fev1_pref_lv1 = ""
            img_fev1_pref_lv2 = ""
            img_fev1_pref_lv3 = ""
            img_fev1_pref_lv4 = ""
            img_fev1_pref_lv5 = ""
            img_fev1_pref_lv6 = ""
            img_fev1_pref_lv7 = ""
            img_fev1_pref_lv8 = ""
            img_fev1_pref_lv9 = ""
            img_fev1_pchg = ""
            img_fev1_pchg_lv1 = ""
            img_fev1_pchg_lv2 = ""
            img_fev1_pchg_lv3 = ""
            img_fev1_pchg_lv4 = ""
            img_fev1_pchg_lv5 = ""
            img_fev1_pchg_lv6 = ""
            img_fev1_pchg_lv7 = ""
            img_fev1_pchg_lv8 = ""
            img_fev1_pchg_lv9 = ""
            img_fev1_pred = ""
            img_fev1_pre_ppred = ""
            img_fev1_post_ppred = ""
            img_pc_fev1 = ""
            img_fev1_pc = ""
            img_fev1dfvc_ref = ""
            img_fev1dfvc_pre = ""
            img_fev1dfvc_post = ""
            img_fev1dfvc_pred = ""
            img_fev1dfvc_tri1 = ""
            img_fev1dfvc_tri2 = ""
            img_fev1dfvc_tri3 = ""
            img_fev1dfvc_tri4 = ""
            img_fev1dfvc_tri5 = ""
            img_fev1dfvc_tri6 = ""
            img_fev1dfvc_tri7 = ""
            img_fev1dfvc_tri8 = ""
            img_fev3dfvc_ref = ""
            img_fev3dfvc_pre = ""
            img_fev3dfvc_post = ""
            img_fev3dfvc_pref_pre = ""
            img_fev3dfvc_pref_post = ""
            img_fev3dfvc_pchg = ""
            img_fef25_75_dose_lv1 = ""
            img_fef25_75_dose_lv2 = ""
            img_fef25_75_dose_lv3 = ""
            img_fef25_75_dose_lv4 = ""
            img_fef25_75_dose_lv5 = ""
            img_fef25_75_dose_lv6 = ""
            img_fef25_75_dose_lv7 = ""
            img_fef25_75_dose_lv8 = ""
            img_fef25_75_dose_lv9 = ""
            img_fef25_75_ref = ""
            img_fef25_75_pre = ""
            img_fef25_75_post = ""
            img_fef25_75_lv1 = ""
            img_fef25_75_lv2 = ""
            img_fef25_75_lv3 = ""
            img_fef25_75_lv4 = ""
            img_fef25_75_lv5 = ""
            img_fef25_75_lv6 = ""
            img_fef25_75_lv7 = ""
            img_fef25_75_lv8 = ""
            img_fef25_75_lv9 = ""
            img_fef25_75_tri1 = ""
            img_fef25_75_tri2 = ""
            img_fef25_75_tri3 = ""
            img_fef25_75_tri4 = ""
            img_fef25_75_tri5 = ""
            img_fef25_75_tri6 = ""
            img_fef25_75_tri7 = ""
            img_fef25_75_tri8 = ""
            img_fef25_75_pref_pre = ""
            img_fef25_75_pref_post = ""
            img_fef25_75_pref_lv1 = ""
            img_fef25_75_pref_lv2 = ""
            img_fef25_75_pref_lv3 = ""
            img_fef25_75_pref_lv4 = ""
            img_fef25_75_pref_lv5 = ""
            img_fef25_75_pref_lv6 = ""
            img_fef25_75_pref_lv7 = ""
            img_fef25_75_pref_lv8 = ""
            img_fef25_75_pref_lv9 = ""
            img_fef25_75_pchg = ""
            img_fef25_75_pchg_lv1 = ""
            img_fef25_75_pchg_lv2 = ""
            img_fef25_75_pchg_lv3 = ""
            img_fef25_75_pchg_lv4 = ""
            img_fef25_75_pchg_lv5 = ""
            img_fef25_75_pchg_lv6 = ""
            img_fef25_75_pchg_lv7 = ""
            img_fef25_75_pchg_lv8 = ""
            img_fef25_75_pchg_lv9 = ""
            img_fef25_75_pred = ""
            img_fef25_75_pre_ppred = ""
            img_fef25_75_post_ppred = ""
            img_isofef25_75_pchg = ""
            img_isofef25_75_pred = ""
            img_isofef25_75_pre = ""
            img_isofef25_75_pre_ppred = ""
            img_isofef25_75_post = ""
            img_isofef25_75_post_ppred = ""
            img_fef75_85_ref = ""
            img_fef75_85_pre = ""
            img_fef75_85_post = ""
            img_fef75_85_pref_pre = ""
            img_fef75_85_pref_post = ""
            img_fef75_85_pchg = ""
            img_fef75_85_pred = ""
            img_fef75_85_pre_ppred = ""
            img_fef75_85_post_ppred = ""
            img_fef25_ref = ""
            img_fef25_pre = ""
            img_fef25_post = ""
            img_fef25_pref_pre = ""
            img_fef25_pref_post = ""
            img_fef25_pchg = ""
            img_fef50_ref = ""
            img_fef50_pre = ""
            img_fef50_post = ""
            img_fef50_pref_pre = ""
            img_fef50_pref_post = ""
            img_fef50_pchg = ""
            img_fef75_ref = ""
            img_fef75_pre = ""
            img_fef75_post = ""
            img_fef75_pref_pre = ""
            img_fef75_pref_post = ""
            img_fef75_pchg = ""
            img_fef200_1200_ref = ""
            img_fef200_1200_pre = ""
            img_fef200_1200_post = ""
            img_fef200_1200_pref_pre = ""
            img_fef200_1200_pref_post = ""
            img_fef200_1200_pchg = ""
            img_pef_dose_lv1 = ""
            img_pef_dose_lv2 = ""
            img_pef_dose_lv3 = ""
            img_pef_dose_lv4 = ""
            img_pef_dose_lv5 = ""
            img_pef_dose_lv6 = ""
            img_pef_dose_lv7 = ""
            img_pef_dose_lv8 = ""
            img_pef_dose_lv9 = ""
            img_pef_ref = ""
            img_pef_pre = ""
            img_pef_post = ""
            img_pef_lv1 = ""
            img_pef_lv2 = ""
            img_pef_lv3 = ""
            img_pef_lv4 = ""
            img_pef_lv5 = ""
            img_pef_lv6 = ""
            img_pef_lv7 = ""
            img_pef_lv8 = ""
            img_pef_lv9 = ""
            img_pef_tri1 = ""
            img_pef_tri2 = ""
            img_pef_tri3 = ""
            img_pef_tri4 = ""
            img_pef_tri5 = ""
            img_pef_tri6 = ""
            img_pef_tri7 = ""
            img_pef_tri8 = ""
            img_pef_pref_pre = ""
            img_pef_pref_post = ""
            img_pef_pref_lv1 = ""
            img_pef_pref_lv2 = ""
            img_pef_pref_lv3 = ""
            img_pef_pref_lv4 = ""
            img_pef_pref_lv5 = ""
            img_pef_pref_lv6 = ""
            img_pef_pref_lv7 = ""
            img_pef_pref_lv8 = ""
            img_pef_pref_lv9 = ""
            img_pef_pchg = ""
            img_pef_pchg_lv1 = ""
            img_pef_pchg_lv2 = ""
            img_pef_pchg_lv3 = ""
            img_pef_pchg_lv4 = ""
            img_pef_pchg_lv5 = ""
            img_pef_pchg_lv6 = ""
            img_pef_pchg_lv7 = ""
            img_pef_pchg_lv8 = ""
            img_pef_pchg_lv9 = ""
            img_pef_pred = ""
            img_pef_pre_ppred = ""
            img_pef_post_ppred = ""
            img_peft_pchg = ""
            img_peft_pre = ""
            img_peft_post = ""
            img_peft_tri1 = ""
            img_peft_tri2 = ""
            img_peft_tri3 = ""
            img_peft_tri4 = ""
            img_peft_tri5 = ""
            img_peft_tri6 = ""
            img_peft_tri7 = ""
            img_peft_tri8 = ""
            img_fet100_pchg = ""
            img_fet100_pre = ""
            img_fet100_post = ""
            img_fet100_tri1 = ""
            img_fet100_tri2 = ""
            img_fet100_tri3 = ""
            img_fet100_tri4 = ""
            img_fet100_tri5 = ""
            img_fet100_tri6 = ""
            img_fet100_tri7 = ""
            img_fet100_tri8 = ""
            img_fivc_ref = ""
            img_fivc_pre = ""
            img_fivc_post = ""
            img_fivc_pref_pre = ""
            img_fivc_pref_post = ""
            img_fivc_pchg = ""
            img_fivc_pred = ""
            img_fivc_pre_ppred = ""
            img_fivc_post_ppred = ""
            img_fivc_tri1 = ""
            img_fivc_tri2 = ""
            img_fivc_tri3 = ""
            img_fivc_tri4 = ""
            img_fivc_tri5 = ""
            img_fivc_tri6 = ""
            img_fivc_tri7 = ""
            img_fivc_tri8 = ""
            img_fiv1_pchg = ""
            img_fiv1_pre = ""
            img_fiv1_post = ""
            img_fefdfif50_pchg = ""
            img_fefdfif50_pred = ""
            img_fefdfif50_pre = ""
            img_fefdfif50_pre_ppred = ""
            img_fefdfif50_post = ""
            img_fefdfif50_post_ppred = ""
            img_volextrap_pchg = ""
            img_volextrap_pre = ""
            img_volextrap_post = ""
            img_volextrap_tri1 = ""
            img_volextrap_tri2 = ""
            img_volextrap_tri3 = ""
            img_volextrap_tri4 = ""
            img_volextrap_tri5 = ""
            img_volextrap_tri6 = ""
            img_volextrap_tri7 = ""
            img_volextrap_tri8 = ""
            img_fvlecode_pre = ""
            img_fvlecode_post = ""
            img_fvlecode_tri1 = ""
            img_fvlecode_tri2 = ""
            img_fvlecode_tri3 = ""
            img_fvlecode_tri4 = ""
            img_fvlecode_tri5 = ""
            img_fvlecode_tri6 = ""
            img_fvlecode_tri7 = ""
            img_fvlecode_tri8 = ""
            img_mvv_pred = ""
            img_tlc_ref = ""
            img_tlc_pre = ""
            img_tlc_pref_pre = ""
            img_vc_ref = ""
            img_vc_pre = ""
            img_vc_pref_pre = ""
            img_ic_ref = ""
            img_ic_pre = ""
            img_ic_pref_pre = ""
            img_frcpl_ref = ""
            img_frcpl_pre = ""
            img_frcpl_pref_pre = ""
            img_erv_ref = ""
            img_erv_pre = ""
            img_erv_pref_pre = ""
            img_rv_ref = ""
            img_rv_pre = ""
            img_rv_pref_pre = ""
            img_rvdtlc_ref = ""
            img_rvdtlc_pre = ""
            img_vtg_pre = ""
            img_vt_pre = ""
            img_dlco_ref = ""
            img_dlco_pre = ""
            img_dlco_pref_pre = ""
            img_dladj_ref = ""
            img_dladj_pre = ""
            img_dladj_pref_pre = ""
            img_dlcodva_ref = ""
            img_dlcodva_pre = ""
            img_dlcodva_pref_pre = ""
            img_dldvaadj_ref = ""
            img_dldvaadj_pre = ""
            img_dldvaadj_pref_pre = ""
            img_va_pre = ""
            img_ivc_pre = ""
            img_dlcoecode_pre = ""
            img_rawtotal_pre = ""
            img_rawinsp_pre = ""
            img_rawexp_pre = ""
            img_raw_ref = ""
            img_raw_pre = ""
            img_raw_pref_pre = ""
            img_gaw_ref = ""
            img_gaw_pre = ""
            img_gaw_pref_pre = ""
            img_sraw_ref = ""
            img_sraw_pre = ""
            img_sraw_pref_pre = ""
            img_sgaw_ref = ""
            img_sgaw_pre = ""
            img_sgaw_pref_pre = ""
            img_rawvtg_pre = ""
            img_rawf_pre = "" 
            img_name = os.path.basename(img)
            new_image_files = cv2.imread(img)
            ori_image_files_counter += 1

            #[s_y:e_y, s_x:e_x]
            if pytesseract.image_to_string(new_image_files[167:223, 239:445]) == 'Methacholine':
                img_type = 'type01'
                img_pid = pytesseract.image_to_string(new_image_files[9:33, 705:811])
                print(img_pid)
                img_date = pytesseract.image_to_string(new_image_files[136:159, 721:805])
                print(img_date)
                img_age = pytesseract.image_to_string(new_image_files[52:84, 709:745])
                print(img_age)
                img_height = pytesseract.image_to_string(new_image_files[54:85, 864:906])
                print(img_height)
                img_weight = pytesseract.image_to_string(new_image_files[80:110, 771:816])
                print(img_weight)
                img_gender = pytesseract.image_to_string(new_image_files[77:112, 893:965])
                print(img_gender)
                img_fvc_dose_lv1 = pytesseract.image_to_string(new_image_files[279:306, 255:335])
                print(img_fvc_dose_lv1)
                img_fvc_dose_lv2 = pytesseract.image_to_string(new_image_files[281:306, 342:396])
                print(img_fvc_dose_lv2)
                img_fvc_dose_lv3 = pytesseract.image_to_string(new_image_files[282:305, 411:466])
                print(img_fvc_dose_lv3)
                img_fvc_dose_lv4 = pytesseract.image_to_string(new_image_files[281:307, 430:529])
                print(img_fvc_dose_lv4)
                img_fvc_dose_lv5 = pytesseract.image_to_string(new_image_files[284:306, 542:597])
                print(img_fvc_dose_lv5)
                img_fvc_dose_lv6 = pytesseract.image_to_string(new_image_files[285:305, 595:661])
                print(img_fvc_dose_lv6)
                img_fvc_ref = pytesseract.image_to_string(new_image_files[303:328, 142:199])
                print(img_fvc_ref)
                img_fvc_pre = pytesseract.image_to_string(new_image_files[303:328, 211:268])
                print(img_fvc_pre)
                img_fvc_lv1 = pytesseract.image_to_string(new_image_files[303:328, 255:335])
                print(img_fvc_lv1)
                img_fvc_lv2 = pytesseract.image_to_string(new_image_files[303:328, 342:396])
                print(img_fvc_lv2)
                img_fvc_lv3 = pytesseract.image_to_string(new_image_files[303:328, 411:466])
                print(img_fvc_lv3)
                img_fvc_lv4 = pytesseract.image_to_string(new_image_files[303:328, 430:529])
                print(img_fvc_lv4)
                img_fvc_lv5 = pytesseract.image_to_string(new_image_files[303:328, 542:597])
                print(img_fvc_lv5)
                img_fvc_lv6 = pytesseract.image_to_string(new_image_files[303:328, 595:661])
                print(img_fvc_lv6)
                img_fvc_pref_pre = pytesseract.image_to_string(new_image_files[325:350, 211:268])
                print(img_fvc_pref_pre)
                img_fvc_pref_lv1 = pytesseract.image_to_string(new_image_files[325:350, 255:335])
                print(img_fvc_pref_lv1)
                img_fvc_pref_lv2 = pytesseract.image_to_string(new_image_files[325:350, 342:396])
                print(img_fvc_pref_lv2)
                img_fvc_pref_lv3 = pytesseract.image_to_string(new_image_files[325:350, 411:466])
                print(img_fvc_pref_lv3)
                img_fvc_pref_lv4 = pytesseract.image_to_string(new_image_files[325:350, 430:529])
                print(img_fvc_pref_lv4)
                img_fvc_pref_lv5 = pytesseract.image_to_string(new_image_files[325:350, 542:597])
                print(img_fvc_pref_lv5)
                img_fvc_pref_lv6 = pytesseract.image_to_string(new_image_files[325:350, 595:661])
                print(img_fvc_pref_lv6)
                img_fvc_pchg_lv1 = pytesseract.image_to_string(new_image_files[344:368, 254:329])
                print(img_fvc_pchg_lv1)
                img_fvc_pchg_lv2 = pytesseract.image_to_string(new_image_files[344:368, 342:396])
                print(img_fvc_pchg_lv2)
                img_fvc_pchg_lv3 = pytesseract.image_to_string(new_image_files[344:368, 411:466])
                print(img_fvc_pchg_lv3)
                img_fvc_pchg_lv4 = pytesseract.image_to_string(new_image_files[344:368, 430:529])
                print(img_fvc_pchg_lv4)
                img_fvc_pchg_lv5 = pytesseract.image_to_string(new_image_files[344:368, 542:597])
                print(img_fvc_pchg_lv5)
                img_fvc_pchg_lv6 = pytesseract.image_to_string(new_image_files[344:368, 596:661])
                print(img_fvc_pchg_lv6)
                img_fev1_dose_lv1 = pytesseract.image_to_string(new_image_files[372:399, 255:335])
                print(img_fev1_dose_lv1)
                img_fev1_dose_lv2 = pytesseract.image_to_string(new_image_files[372:399, 342:396])
                print(img_fev1_dose_lv2)
                img_fev1_dose_lv3 = pytesseract.image_to_string(new_image_files[372:399, 411:466])
                print(img_fev1_dose_lv3)
                img_fev1_dose_lv4 = pytesseract.image_to_string(new_image_files[372:399, 430:529])
                print(img_fev1_dose_lv4)
                img_fev1_dose_lv5 = pytesseract.image_to_string(new_image_files[372:399, 542:597])
                print(img_fev1_dose_lv5)
                img_fev1_dose_lv6 = pytesseract.image_to_string(new_image_files[372:399, 595:661])
                print(img_fev1_dose_lv6)
                img_fev1_ref = pytesseract.image_to_string(new_image_files[396:420, 142:199])
                print(img_fev1_ref)
                img_fev1_pre = pytesseract.image_to_string(new_image_files[396:420, 211:268])
                print(img_fev1_pre)
                img_fev1_lv1 = pytesseract.image_to_string(new_image_files[396:420, 255:335])
                print(img_fev1_lv1)
                img_fev1_lv2 = pytesseract.image_to_string(new_image_files[396:420, 342:396])
                print(img_fev1_lv2)
                img_fev1_lv3 = pytesseract.image_to_string(new_image_files[396:420, 422:466])
                print(img_fev1_lv3)
                img_fev1_lv4 = pytesseract.image_to_string(new_image_files[396:420, 430:529])
                print(img_fev1_lv4)
                img_fev1_lv5 = pytesseract.image_to_string(new_image_files[396:420, 542:597])
                print(img_fev1_lv5)
                img_fev1_lv6 = pytesseract.image_to_string(new_image_files[396:420, 595:661])
                print(img_fev1_lv6)
                img_fev1_pref_pre = pytesseract.image_to_string(new_image_files[416:438, 211:268])
                print(img_fev1_pref_pre)
                img_fev1_pref_lv1 = pytesseract.image_to_string(new_image_files[416:438, 255:335])
                print(img_fev1_pref_lv1)
                img_fev1_pref_lv2 = pytesseract.image_to_string(new_image_files[416:438, 342:396])
                print(img_fev1_pref_lv2)
                img_fev1_pref_lv3 = pytesseract.image_to_string(new_image_files[416:438, 422:466])
                print(img_fev1_pref_lv3)
                img_fev1_pref_lv4 = pytesseract.image_to_string(new_image_files[416:438, 430:529])
                print(img_fev1_pref_lv4)
                img_fev1_pref_lv5 = pytesseract.image_to_string(new_image_files[416:438, 542:597])
                print(img_fev1_pref_lv5)
                img_fev1_pref_lv6 = pytesseract.image_to_string(new_image_files[416:438, 595:661])
                print(img_fev1_pref_lv6)
                img_fev1_pchg_lv1 = pytesseract.image_to_string(new_image_files[435:458, 255:335])
                print(img_fev1_pchg_lv1)
                img_fev1_pchg_lv2 = pytesseract.image_to_string(new_image_files[435:458, 342:396])
                print(img_fev1_pchg_lv2)
                img_fev1_pchg_lv3 = pytesseract.image_to_string(new_image_files[435:458, 422:466])
                print(img_fev1_pchg_lv3)
                img_fev1_pchg_lv4 = pytesseract.image_to_string(new_image_files[435:458, 430:529])
                print(img_fev1_pchg_lv4)
                img_fev1_pchg_lv5 = pytesseract.image_to_string(new_image_files[435:458, 542:587])
                print(img_fev1_pchg_lv5)
                img_fev1_pchg_lv6 = pytesseract.image_to_string(new_image_files[435:458, 595:661])
                print(img_fev1_pchg_lv6)
                img_pc_fev1 = pytesseract.image_to_string(new_image_files[739:776, 746:790])
                print(img_pc_fev1)
                img_fev1_pc = pytesseract.image_to_string(new_image_files[739:776, 856:924])
                print(img_fev1_pc)
                img_fef25_75_dose_lv1 = pytesseract.image_to_string(new_image_files[373:399, 255:335])
                print(img_fef25_75_dose_lv1)
                img_fef25_75_dose_lv2 = pytesseract.image_to_string(new_image_files[373:399, 342:396])
                print(img_fef25_75_dose_lv2)
                img_fef25_75_dose_lv3 = pytesseract.image_to_string(new_image_files[373:399, 422:466])
                print(img_fef25_75_dose_lv3)
                img_fef25_75_dose_lv4 = pytesseract.image_to_string(new_image_files[373:399, 430:529])
                print(img_fef25_75_dose_lv4)
                img_fef25_75_dose_lv5 = pytesseract.image_to_string(new_image_files[373:399, 542:597])
                print(img_fef25_75_dose_lv5)
                img_fef25_75_dose_lv6 = pytesseract.image_to_string(new_image_files[373:399, 595:661])
                print(img_fef25_75_dose_lv6)
                img_fef25_75_ref = pytesseract.image_to_string(new_image_files[393:420, 142:199])
                print(img_fef25_75_ref)
                img_fef25_75_pre = pytesseract.image_to_string(new_image_files[393:420, 211:268])
                print(img_fef25_75_pre)
                img_fef25_75_lv1 = pytesseract.image_to_string(new_image_files[393:420, 255:335])
                print(img_fef25_75_lv1)
                img_fef25_75_lv2 = pytesseract.image_to_string(new_image_files[393:420, 342:396])
                print(img_fef25_75_lv2)
                img_fef25_75_lv3 = pytesseract.image_to_string(new_image_files[393:420, 422:466])
                print(img_fef25_75_lv3)
                img_fef25_75_lv4 = pytesseract.image_to_string(new_image_files[393:420, 430:529])
                print(img_fef25_75_lv4)
                img_fef25_75_lv5 = pytesseract.image_to_string(new_image_files[393:420, 542:597])
                print(img_fef25_75_lv5)
                img_fef25_75_lv6 = pytesseract.image_to_string(new_image_files[393:420, 595:661])
                print(img_fef25_75_lv6)
                img_fef25_75_pref_pre = pytesseract.image_to_string(new_image_files[510:533, 211:268])
                print(img_fef25_75_pref_pre)
                img_fef25_75_pref_lv1 = pytesseract.image_to_string(new_image_files[510:533, 255:335])
                print(img_fef25_75_pref_lv1)
                img_fef25_75_pref_lv2 = pytesseract.image_to_string(new_image_files[510:533, 342:396])
                print(img_fef25_75_pref_lv2)
                img_fef25_75_pref_lv3 = pytesseract.image_to_string(new_image_files[510:533, 422:466])
                print(img_fef25_75_pref_lv3)
                img_fef25_75_pref_lv4 = pytesseract.image_to_string(new_image_files[510:533, 430:529])
                print(img_fef25_75_pref_lv4)
                img_fef25_75_pref_lv5 = pytesseract.image_to_string(new_image_files[510:533, 542:597])
                print(img_fef25_75_pref_lv5)
                img_fef25_75_pref_lv6 = pytesseract.image_to_string(new_image_files[510:533, 595:661])
                print(img_fef25_75_pref_lv6)
                img_fef25_75_pchg_lv1 = pytesseract.image_to_string(new_image_files[530:556, 255:335])
                print(img_fef25_75_pchg_lv1)
                img_fef25_75_pchg_lv2 = pytesseract.image_to_string(new_image_files[530:556, 342:396])
                print(img_fef25_75_pchg_lv2)
                img_fef25_75_pchg_lv3 = pytesseract.image_to_string(new_image_files[530:556, 422:466])
                print(img_fef25_75_pchg_lv3)
                img_fef25_75_pchg_lv4 = pytesseract.image_to_string(new_image_files[530:556, 430:529])
                print(img_fef25_75_pchg_lv4)
                img_fef25_75_pchg_lv5 = pytesseract.image_to_string(new_image_files[530:556, 542:597])
                print(img_fef25_75_pchg_lv5)
                img_fef25_75_pchg_lv6 = pytesseract.image_to_string(new_image_files[530:556, 595:661])
                print(img_fef25_75_pchg_lv6)
                img_pef_dose_lv1 = pytesseract.image_to_string(new_image_files[554:579, 255:335])
                print(img_pef_dose_lv1)
                img_pef_dose_lv2 = pytesseract.image_to_string(new_image_files[554:579, 342:396])
                print(img_pef_dose_lv2)
                img_pef_dose_lv3 = pytesseract.image_to_string(new_image_files[554:579, 422:466])
                print(img_pef_dose_lv3)
                img_pef_dose_lv4 = pytesseract.image_to_string(new_image_files[554:579, 430:529])
                print(img_pef_dose_lv4)
                img_pef_dose_lv5 = pytesseract.image_to_string(new_image_files[554:579, 542:597])
                print(img_pef_dose_lv5)
                img_pef_dose_lv6 = pytesseract.image_to_string(new_image_files[554:579, 595:661])
                print(img_pef_dose_lv6)
                img_pef_ref = pytesseract.image_to_string(new_image_files[577:599, 142:199])
                print(img_pef_ref)
                img_pef_pre = pytesseract.image_to_string(new_image_files[577:599, 211:268])
                print(img_pef_pre)
                img_pef_lv1 = pytesseract.image_to_string(new_image_files[577:599, 255:335])
                print(img_pef_lv1)
                img_pef_lv2 = pytesseract.image_to_string(new_image_files[577:599, 342:396])
                print(img_pef_lv2)
                img_pef_lv3 = pytesseract.image_to_string(new_image_files[577:599, 422:466])
                print(img_pef_lv3)
                img_pef_lv4 = pytesseract.image_to_string(new_image_files[577:599, 430:529])
                print(img_pef_lv4)
                img_pef_lv5 = pytesseract.image_to_string(new_image_files[577:599, 542:597])
                print(img_pef_lv5)
                img_pef_lv6 = pytesseract.image_to_string(new_image_files[577:599, 595:661])
                print(img_pef_lv6)
                img_pef_pref_pre = pytesseract.image_to_string(new_image_files[597:618, 211:268])
                print(img_pef_pref_pre)
                img_pef_pref_lv1 = pytesseract.image_to_string(new_image_files[597:618, 255:335])
                print(img_pef_pref_lv1)
                img_pef_pref_lv2 = pytesseract.image_to_string(new_image_files[597:618, 342:396])
                print(img_pef_pref_lv2)
                img_pef_pref_lv3 = pytesseract.image_to_string(new_image_files[597:618, 422:466])
                print(img_pef_pref_lv3)
                img_pef_pref_lv4 = pytesseract.image_to_string(new_image_files[597:618, 430:529])
                print(img_pef_pref_lv4)
                img_pef_pref_lv5 = pytesseract.image_to_string(new_image_files[597:618, 542:597])
                print(img_pef_pref_lv5)
                img_pef_pref_lv6 = pytesseract.image_to_string(new_image_files[597:618, 595:661])
                print(img_pef_pref_lv6)
                img_pef_pchg_lv1 = pytesseract.image_to_string(new_image_files[616:641, 255:335])
                print(img_pef_pchg_lv1)
                img_pef_pchg_lv2 = pytesseract.image_to_string(new_image_files[616:641, 342:396])
                print(img_pef_pchg_lv2)
                img_pef_pchg_lv3 = pytesseract.image_to_string(new_image_files[616:641, 422:466])
                print(img_pef_pchg_lv3)
                img_pef_pchg_lv4 = pytesseract.image_to_string(new_image_files[616:641, 430:529])
                print(img_pef_pchg_lv4)
                img_pef_pchg_lv5 = pytesseract.image_to_string(new_image_files[616:641, 542:597])
                print(img_pef_pchg_lv5)
                img_pef_pchg_lv6 = pytesseract.image_to_string(new_image_files[616:641, 595:661])
                print(img_pef_pchg_lv6)
                type01_img_cnt += 1
                pass
 
            elif pytesseract.image_to_string(new_image_files[177:217, 156:247]) == 'aridol':
                img_type = 'type02'
                img_pid = pytesseract.image_to_string(new_image_files[3:35, 710:817])
                print(img_pid)
                img_date = pytesseract.image_to_string(new_image_files[134:160, 726:806])
                print(img_date)
                img_age = pytesseract.image_to_string(new_image_files[56:85, 712:755])
                print(img_age)
                img_height = pytesseract.image_to_string(new_image_files[55:87, 871:914])
                print(img_height)
                img_weight = pytesseract.image_to_string(new_image_files[82:111, 773:816])
                print(img_weight)
                img_gender = pytesseract.image_to_string(new_image_files[80:111, 888:959])
                print(img_gender)
                img_fvc_dose_lv1 = pytesseract.image_to_string(new_image_files[284:308, 293:339])
                print(img_fvc_dose_lv1)
                img_fvc_dose_lv2 = pytesseract.image_to_string(new_image_files[284:308, 354:407])
                print(img_fvc_dose_lv2)
                img_fvc_dose_lv3 = pytesseract.image_to_string(new_image_files[284:308, 424:471])
                print(img_fvc_dose_lv3)
                img_fvc_dose_lv4 = pytesseract.image_to_string(new_image_files[284:308, 485:537])
                print(img_fvc_dose_lv4)
                img_fvc_dose_lv5 = pytesseract.image_to_string(new_image_files[284:308, 555:607])
                print(img_fvc_dose_lv5)
                img_fvc_dose_lv6 = pytesseract.image_to_string(new_image_files[284:308, 622:675])
                print(img_fvc_dose_lv6)
                img_fvc_dose_lv7 = pytesseract.image_to_string(new_image_files[284:308, 689:743])
                print(img_fvc_dose_lv7)
                img_fvc_dose_lv8 = pytesseract.image_to_string(new_image_files[284:308, 755:809])
                print(img_fvc_dose_lv8)
                img_fvc_dose_lv9 = pytesseract.image_to_string(new_image_files[284:308, 817:865])
                print(img_fvc_dose_lv9)
                img_fvc_ref = pytesseract.image_to_string(new_image_files[303:331, 148:208])
                print(img_fvc_ref)
                img_fvc_pre = pytesseract.image_to_string(new_image_files[303:331, 215:271])
                print(img_fvc_pre)
                img_fvc_lv1 = pytesseract.image_to_string(new_image_files[303:331, 293:339])
                print(img_fvc_lv1)
                img_fvc_lv2 = pytesseract.image_to_string(new_image_files[303:331, 354:407])
                print(img_fvc_lv2)
                img_fvc_lv3 = pytesseract.image_to_string(new_image_files[303:331, 424:471])
                print(img_fvc_lv3)
                img_fvc_lv4 = pytesseract.image_to_string(new_image_files[303:331, 485:537])
                print(img_fvc_lv4)
                img_fvc_lv5 = pytesseract.image_to_string(new_image_files[303:331, 555:607])
                print(img_fvc_lv5)
                img_fvc_lv6 = pytesseract.image_to_string(new_image_files[303:331, 622:675])
                print(img_fvc_lv6)
                img_fvc_lv7 = pytesseract.image_to_string(new_image_files[303:331, 689:743])
                print(img_fvc_lv7)
                img_fvc_lv8 = pytesseract.image_to_string(new_image_files[303:331, 755:809])
                print(img_fvc_lv8)
                img_fvc_lv9 = pytesseract.image_to_string(new_image_files[303:331, 817:865])
                print(img_fvc_lv9)
                img_fvc_pref_pre = pytesseract.image_to_string(new_image_files[329:351, 215:271])
                print(img_fvc_pref_pre)
                img_fvc_pref_lv1 = pytesseract.image_to_string(new_image_files[329:351, 293:339])
                print(img_fvc_pref_lv1)
                img_fvc_pref_lv2 = pytesseract.image_to_string(new_image_files[329:351, 354:407])
                print(img_fvc_pref_lv2)
                img_fvc_pref_lv3 = pytesseract.image_to_string(new_image_files[329:351, 424:471])
                print(img_fvc_pref_lv3)
                img_fvc_pref_lv4 = pytesseract.image_to_string(new_image_files[329:351, 485:537])
                print(img_fvc_pref_lv4)
                img_fvc_pref_lv5 = pytesseract.image_to_string(new_image_files[329:351, 555:607])
                print(img_fvc_pref_lv5)
                img_fvc_pref_lv6 = pytesseract.image_to_string(new_image_files[329:351, 622:675])
                print(img_fvc_pref_lv6)
                img_fvc_pref_lv7 = pytesseract.image_to_string(new_image_files[329:351, 689:743])
                print(img_fvc_pref_lv7)
                img_fvc_pref_lv8 = pytesseract.image_to_string(new_image_files[329:351, 755:809])
                print(img_fvc_pref_lv8)
                img_fvc_pref_lv9 = pytesseract.image_to_string(new_image_files[329:351, 817:865])
                print(img_fvc_pref_lv9)
                img_fvc_pchg_lv1 = pytesseract.image_to_string(new_image_files[350:373, 293:339])
                print(img_fvc_pchg_lv1)
                img_fvc_pchg_lv2 = pytesseract.image_to_string(new_image_files[350:373, 354:407])
                print(img_fvc_pchg_lv2)
                img_fvc_pchg_lv3 = pytesseract.image_to_string(new_image_files[350:373, 424:471])
                print(img_fvc_pchg_lv3)
                img_fvc_pchg_lv4 = pytesseract.image_to_string(new_image_files[350:373, 485:537])
                print(img_fvc_pchg_lv4)
                img_fvc_pchg_lv5 = pytesseract.image_to_string(new_image_files[350:373, 555:607])
                print(img_fvc_pchg_lv5)
                img_fvc_pchg_lv6 = pytesseract.image_to_string(new_image_files[350:373, 622:675])
                print(img_fvc_pchg_lv6)
                img_fvc_pchg_lv7 = pytesseract.image_to_string(new_image_files[350:373, 689:743])
                print(img_fvc_pchg_lv7)
                img_fvc_pchg_lv8 = pytesseract.image_to_string(new_image_files[350:373, 755:809])
                print(img_fvc_pchg_lv8)
                img_fvc_pchg_lv9 = pytesseract.image_to_string(new_image_files[350:373, 817:865])
                print(img_fvc_pchg_lv9)
                img_fev1_dose_lv1 = pytesseract.image_to_string(new_image_files[371:401, 293:339])
                print(img_fev1_dose_lv1)
                img_fev1_dose_lv2 = pytesseract.image_to_string(new_image_files[371:401, 354:407])
                print(img_fev1_dose_lv2)
                img_fev1_dose_lv3 = pytesseract.image_to_string(new_image_files[371:401, 424:471])
                print(img_fev1_dose_lv3)
                img_fev1_dose_lv4 = pytesseract.image_to_string(new_image_files[371:401, 485:537])
                print(img_fev1_dose_lv4)
                img_fev1_dose_lv5 = pytesseract.image_to_string(new_image_files[371:401, 555:607])
                print(img_fev1_dose_lv5)
                img_fev1_dose_lv6 = pytesseract.image_to_string(new_image_files[371:401, 622:675])
                print(img_fev1_dose_lv6)
                img_fev1_dose_lv7 = pytesseract.image_to_string(new_image_files[371:401, 689:743])
                print(img_fev1_dose_lv7)
                img_fev1_dose_lv8 = pytesseract.image_to_string(new_image_files[371:401, 755:809])
                print(img_fev1_dose_lv8)
                img_fev1_dose_lv9 = pytesseract.image_to_string(new_image_files[371:401, 817:865])
                print(img_fev1_dose_lv9)
                img_fev1_ref = pytesseract.image_to_string(new_image_files[397:423, 148:208])
                print(img_fev1_ref)
                img_fev1_pre = pytesseract.image_to_string(new_image_files[397:423, 215:271])
                print(img_fev1_pre)
                img_fev1_lv1 = pytesseract.image_to_string(new_image_files[397:423, 293:339])
                print(img_fev1_lv1)
                img_fev1_lv2 = pytesseract.image_to_string(new_image_files[397:423, 354:407])
                print(img_fev1_lv2)
                img_fev1_lv3 = pytesseract.image_to_string(new_image_files[397:423, 424:471])
                print(img_fev1_lv3)
                img_fev1_lv4 = pytesseract.image_to_string(new_image_files[397:423, 485:537])
                print(img_fev1_lv4)
                img_fev1_lv5 = pytesseract.image_to_string(new_image_files[397:423, 555:607])
                print(img_fev1_lv5)
                img_fev1_lv6 = pytesseract.image_to_string(new_image_files[397:423, 622:675])
                print(img_fev1_lv6)
                img_fev1_lv7 = pytesseract.image_to_string(new_image_files[397:423, 689:743])
                print(img_fev1_lv7)
                img_fev1_lv8 = pytesseract.image_to_string(new_image_files[397:423, 755:809])
                print(img_fev1_lv8)
                img_fev1_lv9 = pytesseract.image_to_string(new_image_files[397:423, 817:865])
                print(img_fev1_lv9)
                img_fev1_pref_pre = pytesseract.image_to_string(new_image_files[418:442, 215:271])
                print(img_fev1_pref_pre)
                img_fev1_pref_lv1 = pytesseract.image_to_string(new_image_files[418:442, 293:339])
                print(img_fev1_pref_lv1)
                img_fev1_pref_lv2 = pytesseract.image_to_string(new_image_files[418:442, 354:407])
                print(img_fev1_pref_lv2)
                img_fev1_pref_lv3 = pytesseract.image_to_string(new_image_files[418:442, 424:471])
                print(img_fev1_pref_lv3)
                img_fev1_pref_lv4 = pytesseract.image_to_string(new_image_files[418:442, 485:537])
                print(img_fev1_pref_lv4)
                img_fev1_pref_lv5 = pytesseract.image_to_string(new_image_files[418:442, 555:607])
                print(img_fev1_pref_lv5)
                img_fev1_pref_lv6 = pytesseract.image_to_string(new_image_files[418:442, 622:675])
                print(img_fev1_pref_lv6)
                img_fev1_pref_lv7 = pytesseract.image_to_string(new_image_files[418:442, 689:743])
                print(img_fev1_pref_lv7)
                img_fev1_pref_lv8 = pytesseract.image_to_string(new_image_files[418:442, 755:809])
                print(img_fev1_pref_lv8)
                img_fev1_pref_lv9 = pytesseract.image_to_string(new_image_files[418:442, 817:865])
                print(img_fev1_pref_lv9)
                img_fev1_pchg_lv1 = pytesseract.image_to_string(new_image_files[438:465, 293:339])
                print(img_fev1_pchg_lv1)
                img_fev1_pchg_lv2 = pytesseract.image_to_string(new_image_files[438:465, 354:407])
                print(img_fev1_pchg_lv2)
                img_fev1_pchg_lv3 = pytesseract.image_to_string(new_image_files[438:465, 424:471])
                print(img_fev1_pchg_lv3)
                img_fev1_pchg_lv4 = pytesseract.image_to_string(new_image_files[438:465, 485:537])
                print(img_fev1_pchg_lv4)
                img_fev1_pchg_lv5 = pytesseract.image_to_string(new_image_files[438:465, 555:607])
                print(img_fev1_pchg_lv5)
                img_fev1_pchg_lv6 = pytesseract.image_to_string(new_image_files[438:465, 622:675])
                print(img_fev1_pchg_lv6)
                img_fev1_pchg_lv7 = pytesseract.image_to_string(new_image_files[438:465, 689:743])
                print(img_fev1_pchg_lv7)
                img_fev1_pchg_lv8 = pytesseract.image_to_string(new_image_files[438:465, 755:809])
                print(img_fev1_pchg_lv8)
                img_fev1_pchg_lv9 = pytesseract.image_to_string(new_image_files[438:465, 817:865])
                print(img_fev1_pchg_lv9)
                img_pc_fev1 = pytesseract.image_to_string(new_image_files[748:781, 755:803])
                print(img_pc_fev1)
                img_fev1_pc = pytesseract.image_to_string(new_image_files[750:783, 867:928])
                print(img_fev1_pc)
                img_fef25_75_dose_lv1 = pytesseract.image_to_string(new_image_files[468:493, 293:339])
                print(img_fef25_75_dose_lv1)
                img_fef25_75_dose_lv2 = pytesseract.image_to_string(new_image_files[468:493, 354:407])
                print(img_fef25_75_dose_lv2)
                img_fef25_75_dose_lv3 = pytesseract.image_to_string(new_image_files[468:493, 424:471])
                print(img_fef25_75_dose_lv3)
                img_fef25_75_dose_lv4 = pytesseract.image_to_string(new_image_files[468:493, 485:537])
                print(img_fef25_75_dose_lv4)
                img_fef25_75_dose_lv5 = pytesseract.image_to_string(new_image_files[468:493, 555:607])
                print(img_fef25_75_dose_lv5)
                img_fef25_75_dose_lv6 = pytesseract.image_to_string(new_image_files[468:493, 622:675])
                print(img_fef25_75_dose_lv6)
                img_fef25_75_dose_lv7 = pytesseract.image_to_string(new_image_files[468:493, 689:743])
                print(img_fef25_75_dose_lv7)
                img_fef25_75_dose_lv8 = pytesseract.image_to_string(new_image_files[468:493, 755:809])
                print(img_fef25_75_dose_lv8)
                img_fef25_75_dose_lv9 = pytesseract.image_to_string(new_image_files[468:493, 817:865])
                print(img_fef25_75_dose_lv9)
                img_fef25_75_ref = pytesseract.image_to_string(new_image_files[490:514, 148:208])
                print(img_fef25_75_ref)
                img_fef25_75_pre = pytesseract.image_to_string(new_image_files[490:514, 215:271])
                print(img_fef25_75_pre)
                img_fef25_75_lv1 = pytesseract.image_to_string(new_image_files[490:514, 293:339])
                print(img_fef25_75_lv1)
                img_fef25_75_lv2 = pytesseract.image_to_string(new_image_files[490:514, 354:407])
                print(img_fef25_75_lv2)
                img_fef25_75_lv3 = pytesseract.image_to_string(new_image_files[490:514, 424:471])
                print(img_fef25_75_lv3)
                img_fef25_75_lv4 = pytesseract.image_to_string(new_image_files[490:514, 485:537])
                print(img_fef25_75_lv4)
                img_fef25_75_lv5 = pytesseract.image_to_string(new_image_files[490:514, 555:607])
                print(img_fef25_75_lv5)
                img_fef25_75_lv6 = pytesseract.image_to_string(new_image_files[490:514, 622:675])
                print(img_fef25_75_lv6)
                img_fef25_75_lv7 = pytesseract.image_to_string(new_image_files[490:514, 689:743])
                print(img_fef25_75_lv7)
                img_fef25_75_lv8 = pytesseract.image_to_string(new_image_files[490:514, 755:809])
                print(img_fef25_75_lv8)
                img_fef25_75_lv9 = pytesseract.image_to_string(new_image_files[490:514, 817:865])
                print(img_fef25_75_lv9)
                img_fef25_75_pref_pre = pytesseract.image_to_string(new_image_files[515:539, 215:271])
                print(img_fef25_75_pref_pre)
                img_fef25_75_pref_lv1 = pytesseract.image_to_string(new_image_files[515:539, 293:339])
                print(img_fef25_75_pref_lv1)
                img_fef25_75_pref_lv2 = pytesseract.image_to_string(new_image_files[515:539, 354:407])
                print(img_fef25_75_pref_lv2)
                img_fef25_75_pref_lv3 = pytesseract.image_to_string(new_image_files[515:539, 424:471])
                print(img_fef25_75_pref_lv3)
                img_fef25_75_pref_lv4 = pytesseract.image_to_string(new_image_files[515:539, 485:537])
                print(img_fef25_75_pref_lv4)
                img_fef25_75_pref_lv5 = pytesseract.image_to_string(new_image_files[515:539, 555:607])
                print(img_fef25_75_pref_lv5)
                img_fef25_75_pref_lv6 = pytesseract.image_to_string(new_image_files[515:539, 622:675])
                print(img_fef25_75_pref_lv6)
                img_fef25_75_pref_lv7 = pytesseract.image_to_string(new_image_files[515:539, 689:743])
                print(img_fef25_75_pref_lv7)
                img_fef25_75_pref_lv8 = pytesseract.image_to_string(new_image_files[515:539, 755:809])
                print(img_fef25_75_pref_lv8)
                img_fef25_75_pref_lv9 = pytesseract.image_to_string(new_image_files[515:539, 817:865])
                print(img_fef25_75_pref_lv9)
                img_fef25_75_pchg_lv1 = pytesseract.image_to_string(new_image_files[537:564, 293:339])
                print(img_fef25_75_pchg_lv1)
                img_fef25_75_pchg_lv2 = pytesseract.image_to_string(new_image_files[537:564, 354:407])
                print(img_fef25_75_pchg_lv2)
                img_fef25_75_pchg_lv3 = pytesseract.image_to_string(new_image_files[537:564, 424:471])
                print(img_fef25_75_pchg_lv3)
                img_fef25_75_pchg_lv4 = pytesseract.image_to_string(new_image_files[537:564, 485:537])
                print(img_fef25_75_pchg_lv4)
                img_fef25_75_pchg_lv5 = pytesseract.image_to_string(new_image_files[537:564, 555:607])
                print(img_fef25_75_pchg_lv5)
                img_fef25_75_pchg_lv6 = pytesseract.image_to_string(new_image_files[537:564, 622:675])
                print(img_fef25_75_pchg_lv6)
                img_fef25_75_pchg_lv7 = pytesseract.image_to_string(new_image_files[537:564, 689:743])
                print(img_fef25_75_pchg_lv7)
                img_fef25_75_pchg_lv8 = pytesseract.image_to_string(new_image_files[537:564, 755:809])
                print(img_fef25_75_pchg_lv8)
                img_fef25_75_pchg_lv9 = pytesseract.image_to_string(new_image_files[537:564, 817:865])
                print(img_fef25_75_pchg_lv9)
                img_pef_dose_lv1 = pytesseract.image_to_string(new_image_files[560:586, 293:339])
                print(img_pef_dose_lv1)
                img_pef_dose_lv2 = pytesseract.image_to_string(new_image_files[560:586, 354:407])
                print(img_pef_dose_lv2)
                img_pef_dose_lv3 = pytesseract.image_to_string(new_image_files[560:586, 424:471])
                print(img_pef_dose_lv3)
                img_pef_dose_lv4 = pytesseract.image_to_string(new_image_files[560:586, 485:537])
                print(img_pef_dose_lv4)
                img_pef_dose_lv5 = pytesseract.image_to_string(new_image_files[560:586, 555:607])
                print(img_pef_dose_lv5)
                img_pef_dose_lv6 = pytesseract.image_to_string(new_image_files[560:586, 622:675])
                print(img_pef_dose_lv6)
                img_pef_dose_lv7 = pytesseract.image_to_string(new_image_files[560:586, 689:743])
                print(img_pef_dose_lv7)
                img_pef_dose_lv8 = pytesseract.image_to_string(new_image_files[560:586, 755:809])
                print(img_pef_dose_lv8)
                img_pef_dose_lv9 = pytesseract.image_to_string(new_image_files[560:586, 817:865])
                print(img_pef_dose_lv9)
                img_pef_ref = pytesseract.image_to_string(new_image_files[579:606, 148:208])
                print(img_pef_ref)
                img_pef_pre = pytesseract.image_to_string(new_image_files[579:606, 215:271])
                print(img_pef_pre)
                img_pef_lv1 = pytesseract.image_to_string(new_image_files[579:606, 293:339])
                print(img_pef_lv1)
                img_pef_lv2 = pytesseract.image_to_string(new_image_files[579:606, 354:407])
                print(img_pef_lv2)
                img_pef_lv3 = pytesseract.image_to_string(new_image_files[579:606, 424:471])
                print(img_pef_lv3)
                img_pef_lv4 = pytesseract.image_to_string(new_image_files[579:606, 485:537])
                print(img_pef_lv4)
                img_pef_lv5 = pytesseract.image_to_string(new_image_files[579:606, 555:607])
                print(img_pef_lv5)
                img_pef_lv6 = pytesseract.image_to_string(new_image_files[579:606, 622:675])
                print(img_pef_lv6)
                img_pef_lv7 = pytesseract.image_to_string(new_image_files[579:606, 689:743])
                print(img_pef_lv7)
                img_pef_lv8 = pytesseract.image_to_string(new_image_files[579:606, 755:809])
                print(img_pef_lv8)
                img_pef_lv9 = pytesseract.image_to_string(new_image_files[579:606, 817:865])
                print(img_pef_lv9)
                img_pef_pref_pre = pytesseract.image_to_string(new_image_files[602:626, 215:271])
                print(img_pef_pref_pre)
                img_pef_pref_lv1 = pytesseract.image_to_string(new_image_files[602:626, 293:339])
                print(img_pef_pref_lv1)
                img_pef_pref_lv2 = pytesseract.image_to_string(new_image_files[602:626, 354:407])
                print(img_pef_pref_lv2)
                img_pef_pref_lv3 = pytesseract.image_to_string(new_image_files[602:626, 424:471])
                print(img_pef_pref_lv3)
                img_pef_pref_lv4 = pytesseract.image_to_string(new_image_files[602:626, 485:537])
                print(img_pef_pref_lv4)
                img_pef_pref_lv5 = pytesseract.image_to_string(new_image_files[602:626, 555:607])
                print(img_pef_pref_lv5)
                img_pef_pref_lv6 = pytesseract.image_to_string(new_image_files[602:626, 622:675])
                print(img_pef_pref_lv6)
                img_pef_pref_lv7 = pytesseract.image_to_string(new_image_files[602:626, 689:743])
                print(img_pef_pref_lv7)
                img_pef_pref_lv8 = pytesseract.image_to_string(new_image_files[602:626, 755:809])
                print(img_pef_pref_lv8)
                img_pef_pref_lv9 = pytesseract.image_to_string(new_image_files[602:626, 817:865])
                print(img_pef_pref_lv9)
                img_pef_pchg_lv1 = pytesseract.image_to_string(new_image_files[622:650, 293:339])
                print(img_pef_pchg_lv1)
                img_pef_pchg_lv2 = pytesseract.image_to_string(new_image_files[622:650, 354:407])
                print(img_pef_pchg_lv2)
                img_pef_pchg_lv3 = pytesseract.image_to_string(new_image_files[622:650, 424:471])
                print(img_pef_pchg_lv3)
                img_pef_pchg_lv4 = pytesseract.image_to_string(new_image_files[622:650, 485:537])
                print(img_pef_pchg_lv4)
                img_pef_pchg_lv5 = pytesseract.image_to_string(new_image_files[622:650, 555:607])
                print(img_pef_pchg_lv5)
                img_pef_pchg_lv6 = pytesseract.image_to_string(new_image_files[622:650, 622:675])
                print(img_pef_pchg_lv6)
                img_pef_pchg_lv7 = pytesseract.image_to_string(new_image_files[622:650, 689:743])
                print(img_pef_pchg_lv7)
                img_pef_pchg_lv8 = pytesseract.image_to_string(new_image_files[622:650, 755:809])
                print(img_pef_pchg_lv8)
                img_pef_pchg_lv9 = pytesseract.image_to_string(new_image_files[622:650, 817:865])
                print(img_pef_pchg_lv9)
                type02_img_cnt += 1
                pass
          
            elif pytesseract.image_to_string(new_image_files[360:390, 38:111]) == 'Diffusing':
                img_type = 'type03'
                img_pid = pytesseract.image_to_string(new_image_files[54:83, 684:776])
                print(img_pid)
                img_date = pytesseract.image_to_string(new_image_files[4:35, 712:801])
                print(img_date)
                img_age = pytesseract.image_to_string(new_image_files[77:111, 695:737])
                print(img_age)
                img_height = pytesseract.image_to_string(new_image_files[77:110, 860:902])
                print(img_height)
                img_weight = pytesseract.image_to_string(new_image_files[105:136, 765:808])
                print(img_weight)
                img_gender = pytesseract.image_to_string(new_image_files[105:135, 881:966])
                print(img_gender)
                img_dlco_ref = pytesseract.image_to_string(new_image_files[386:405, 326:370])
                print(img_dlco_ref)
                img_dlco_pre = pytesseract.image_to_string(new_image_files[386:405, 387:436])
                print(img_dlco_pre)
                img_dlco_pref_pre = pytesseract.image_to_string(new_image_files[386:405, 464:497])
                print(img_dlco_pref_pre)
                img_dladj_ref = pytesseract.image_to_string(new_image_files[403:422, 326:370])
                print(img_dladj_ref)
                img_dladj_pre = pytesseract.image_to_string(new_image_files[403:422, 387:436])
                print(img_dladj_pre)
                img_dladj_pref_pre = pytesseract.image_to_string(new_image_files[403:422, 464:497])
                print(img_dladj_pref_pre)
                img_dlcodva_ref = pytesseract.image_to_string(new_image_files[420:440, 326:370])
                print(img_dlcodva_ref)
                img_dlcodva_pre = pytesseract.image_to_string(new_image_files[420:440, 387:436])
                print(img_dlcodva_pre)
                img_dlcodva_pref_pre = pytesseract.image_to_string(new_image_files[420:440, 464:497])
                print(img_dlcodva_pref_pre)
                img_dldvaadj_ref = pytesseract.image_to_string(new_image_files[440:459, 326:370])
                print(img_dldvaadj_ref)
                img_dldvaadj_pre = pytesseract.image_to_string(new_image_files[440:459, 387:436])
                print(img_dldvaadj_pre)
                img_dldvaadj_pref_pre = pytesseract.image_to_string(new_image_files[440:459, 464:497])
                print(img_dldvaadj_pref_pre)
                img_va_pre = pytesseract.image_to_string(new_image_files[456:475, 387:436])
                print(img_va_pre)
                img_ivc_pre = pytesseract.image_to_string(new_image_files[475:494, 387:436])
                print(img_ivc_pre)
                img_dlcoecode_pre = pytesseract.image_to_string(new_image_files[491:518, 387:436])
                print(img_dlcoecode_pre)
                type03_img_cnt += 1
                pass

            elif pytesseract.image_to_string(new_image_files[266:295, 1:89]) == 'Spirometry':
                img_type = 'type04'
                img_pid = pytesseract.image_to_string(new_image_files[55:85, 685:770])
                print(img_pid)
                img_date = pytesseract.image_to_string(new_image_files[3:36, 708:800])
                print(img_date)
                img_age = pytesseract.image_to_string(new_image_files[83:108, 698:728])
                print(img_age)
                img_height = pytesseract.image_to_string(new_image_files[79:110, 862:904])
                print(img_height)
                img_weight = pytesseract.image_to_string(new_image_files[108:134, 758:805])
                print(img_weight)
                img_gender = pytesseract.image_to_string(new_image_files[106:137, 882:959])
                print(img_gender)
                img_fvc_ref = pytesseract.image_to_string(new_image_files[314:391, 305:357])
                print(img_fvc_ref)
                img_fvc_pre = pytesseract.image_to_string(new_image_files[314:391, 365:429])
                print(img_fvc_pre)
                img_fvc_post = pytesseract.image_to_string(new_image_files[314:391, 504:573])
                print(img_fvc_post)
                img_fvc_pref_pre = pytesseract.image_to_string(new_image_files[314:391, 453:509])
                print(img_fvc_pref_pre)
                img_fvc_pref_post = pytesseract.image_to_string(new_image_files[314:391, 596:655])
                print(img_fvc_pref_post)
                img_fvc_pchg = pytesseract.image_to_string(new_image_files[314:391, 669:714])
                print(img_fvc_pchg)
                img_fev05_ref = pytesseract.image_to_string(new_image_files[313:334, 305:357])
                print(img_fev05_ref)
                img_fev05_pre = pytesseract.image_to_string(new_image_files[313:334, 365:429])
                print(img_fev05_pre)
                img_fev05_post = pytesseract.image_to_string(new_image_files[313:334, 504:573])
                print(img_fev05_post)
                img_fev05_pref_pre = pytesseract.image_to_string(new_image_files[313:334, 453:509])
                print(img_fev05_pref_pre)
                img_fev05_pref_post = pytesseract.image_to_string(new_image_files[313:334, 596:655])
                print(img_fev05_pref_post)
                img_fev05_pchg = pytesseract.image_to_string(new_image_files[313:334, 669:714])
                print(img_fev05_pchg)
                img_fev1_ref = pytesseract.image_to_string(new_image_files[333:354, 305:357])
                print(img_fev1_ref)
                img_fev1_pre = pytesseract.image_to_string(new_image_files[333:354, 365:429])
                print(img_fev1_pre)
                img_fev1_post = pytesseract.image_to_string(new_image_files[333:354, 504:573])
                print(img_fev1_post)
                img_fev1_pref_pre = pytesseract.image_to_string(new_image_files[333:354, 453:509])
                print(img_fev1_pref_pre)
                img_fev1_pref_post = pytesseract.image_to_string(new_image_files[333:354, 596:655])
                print(img_fev1_pref_post)
                img_fev1_pchg = pytesseract.image_to_string(new_image_files[333:354, 669:714])
                print(img_fev1_pchg)
                img_fev1dfvc_ref = pytesseract.image_to_string(new_image_files[352:372, 305:357])
                print(img_fev1dfvc_ref)
                img_fev1dfvc_pre = pytesseract.image_to_string(new_image_files[352:372, 365:429])
                print(img_fev1dfvc_pre)
                img_fev1dfvc_post = pytesseract.image_to_string(new_image_files[352:372, 504:573])
                print(img_fev1dfvc_post)
                img_fev3dfvc_ref = pytesseract.image_to_string(new_image_files[370:392, 305:357])
                print(img_fev3dfvc_ref)
                img_fev3dfvc_pre = pytesseract.image_to_string(new_image_files[370:392, 365:429])
                print(img_fev3dfvc_pre)
                img_fev3dfvc_post = pytesseract.image_to_string(new_image_files[370:392, 504:573])
                print(img_fev3dfvc_post)
                img_fef25_75_ref = pytesseract.image_to_string(new_image_files[393:412, 305:357])
                print(img_fef25_75_ref)
                img_fef25_75_pre = pytesseract.image_to_string(new_image_files[393:412, 365:429])
                print(img_fef25_75_pre)
                img_fef25_75_post = pytesseract.image_to_string(new_image_files[393:412, 504:573])
                print(img_fef25_75_post)
                img_fef25_75_pref_pre = pytesseract.image_to_string(new_image_files[393:412, 453:509])
                print(img_fef25_75_pref_pre)
                img_fef25_75_pref_post = pytesseract.image_to_string(new_image_files[393:412, 596:655])
                print(img_fef25_75_pref_post)
                img_fef25_75_pchg = pytesseract.image_to_string(new_image_files[393:412, 669:714])
                print(img_fef25_75_pchg)
                img_fef75_85_ref = pytesseract.image_to_string(new_image_files[410:431, 305:357])
                print(img_fef75_85_ref)
                img_fef75_85_pre = pytesseract.image_to_string(new_image_files[410:431, 365:429])
                print(img_fef75_85_pre)
                img_fef75_85_post = pytesseract.image_to_string(new_image_files[410:431, 504:573])
                print(img_fef75_85_post)
                img_fef75_85_pref_pre = pytesseract.image_to_string(new_image_files[410:431, 453:509])
                print(img_fef75_85_pref_pre)
                img_fef75_85_pref_post = pytesseract.image_to_string(new_image_files[410:431, 596:655])
                print(img_fef75_85_pref_post)
                img_fef75_85_pchg = pytesseract.image_to_string(new_image_files[410:431, 669:714])
                print(img_fef75_85_pchg)
                img_fef25_pre = pytesseract.image_to_string(new_image_files[429:450, 365:429])
                print(img_fef25_pre)
                img_fef25_post = pytesseract.image_to_string(new_image_files[429:450, 504:573])
                print(img_fef25_post)
                img_fef25_pchg = pytesseract.image_to_string(new_image_files[429:450, 669:714])
                print(img_fef25_pchg)
                img_fef50_ref = pytesseract.image_to_string(new_image_files[450:472, 305:357])
                print(img_fef50_ref)
                img_fef50_pre = pytesseract.image_to_string(new_image_files[450:472, 365:429])
                print(img_fef50_pre)
                img_fef50_post = pytesseract.image_to_string(new_image_files[450:472, 504:573])
                print(img_fef50_post)
                img_fef50_pref_pre = pytesseract.image_to_string(new_image_files[450:472, 453:509])
                print(img_fef50_pref_pre)
                img_fef50_pref_post = pytesseract.image_to_string(new_image_files[450:472, 596:655])
                print(img_fef50_pref_post)
                img_fef50_pchg = pytesseract.image_to_string(new_image_files[450:472, 669:714])
                print(img_fef50_pchg)
                img_fef75_ref = pytesseract.image_to_string(new_image_files[470:489, 305:357])
                print(img_fef75_ref)
                img_fef75_pre = pytesseract.image_to_string(new_image_files[470:489, 365:429])
                print(img_fef75_pre)
                img_fef75_post = pytesseract.image_to_string(new_image_files[470:489, 504:573])
                print(img_fef75_post)
                img_fef75_pref_pre = pytesseract.image_to_string(new_image_files[470:489, 453:509])
                print(img_fef75_pref_pre)
                img_fef75_pref_post = pytesseract.image_to_string(new_image_files[470:489, 596:655])
                print(img_fef75_pref_post)
                img_fef75_pchg = pytesseract.image_to_string(new_image_files[470:489, 669:714])
                print(img_fef75_pchg)
                img_fef200_1200_ref = pytesseract.image_to_string(new_image_files[487:510, 305:357])
                print(img_fef200_1200_ref)
                img_fef200_1200_pre = pytesseract.image_to_string(new_image_files[487:510, 365:429])
                print(img_fef200_1200_pre)
                img_fef200_1200_post = pytesseract.image_to_string(new_image_files[487:510, 504:573])
                print(img_fef200_1200_post)
                img_fef200_1200_pref_pre = pytesseract.image_to_string(new_image_files[487:510, 453:509])
                print(img_fef200_1200_pref_pre)
                img_fef200_1200_pref_post = pytesseract.image_to_string(new_image_files[487:510, 596:655])
                print(img_fef200_1200_pref_post)
                img_fef200_1200_pchg = pytesseract.image_to_string(new_image_files[487:510, 669:714])
                print(img_fef200_1200_pchg)
                img_pef_ref = pytesseract.image_to_string(new_image_files[508:527, 305:357])
                print(img_pef_ref)
                img_pef_pre = pytesseract.image_to_string(new_image_files[508:527, 365:429])
                print(img_pef_pre)
                img_pef_post = pytesseract.image_to_string(new_image_files[508:527, 504:573])
                print(img_pef_post)
                img_pef_pref_pre = pytesseract.image_to_string(new_image_files[508:527, 453:509])
                print(img_pef_pref_pre)
                img_pef_pref_post = pytesseract.image_to_string(new_image_files[508:527, 596:655])
                print(img_pef_pref_post)
                img_pef_pchg = pytesseract.image_to_string(new_image_files[508:527, 669:714])
                print(img_pef_pchg)
                img_fivc_ref = pytesseract.image_to_string(new_image_files[528:548, 305:357])
                print(img_fivc_ref)
                img_fivc_pre = pytesseract.image_to_string(new_image_files[528:548, 365:429])
                print(img_fivc_pre)
                img_fivc_post = pytesseract.image_to_string(new_image_files[528:548, 504:573])
                print(img_fivc_post)
                img_fivc_pref_pre = pytesseract.image_to_string(new_image_files[528:548, 453:509])
                print(img_fivc_pref_pre)
                img_fivc_pref_post = pytesseract.image_to_string(new_image_files[528:548, 596:655])
                print(img_fivc_pref_post)
                img_fivc_pchg = pytesseract.image_to_string(new_image_files[528:548, 669:714])
                print(img_fivc_pchg)
                img_fvlecode_pre = pytesseract.image_to_string(new_image_files[546:571, 365:429])
                print(img_fvlecode_pre)
                img_fvlecode_post = pytesseract.image_to_string(new_image_files[546:571, 504:573])
                print(img_fvlecode_post)
                img_tlc_ref = pytesseract.image_to_string(new_image_files[600:626, 305:357])
                print(img_tlc_ref)
                img_tlc_pre = pytesseract.image_to_string(new_image_files[600:626, 365:429])
                print(img_tlc_pre)
                img_tlc_pref_pre = pytesseract.image_to_string(new_image_files[600:626, 453:509])
                print(img_tlc_pref_pre)
                img_vc_ref = pytesseract.image_to_string(new_image_files[624:645, 305:357])
                print(img_vc_ref)
                img_vc_pre = pytesseract.image_to_string(new_image_files[624:645, 365:429])
                print(img_vc_pre)
                img_vc_pref_pre = pytesseract.image_to_string(new_image_files[624:645, 453:509])
                print(img_vc_pref_pre)
                img_ic_ref = pytesseract.image_to_string(new_image_files[644:665, 305:357])
                print(img_ic_ref)
                img_ic_pre = pytesseract.image_to_string(new_image_files[644:665, 365:429])
                print(img_ic_pre)
                img_ic_pref_pre = pytesseract.image_to_string(new_image_files[644:665, 453:509])
                print(img_ic_pref_pre)
                img_frcpl_ref = pytesseract.image_to_string(new_image_files[664:683, 305:357])
                print(img_frcpl_ref)
                img_frcpl_pre = pytesseract.image_to_string(new_image_files[664:683, 365:429])
                print(img_frcpl_pre)
                img_frcpl_pref_pre = pytesseract.image_to_string(new_image_files[664:683, 453:509])
                print(img_frcpl_pref_pre)
                img_erv_ref = pytesseract.image_to_string(new_image_files[683:703, 305:357])
                print(img_erv_ref)
                img_erv_pre = pytesseract.image_to_string(new_image_files[683:703, 365:429])
                print(img_erv_pre)
                img_erv_pref_pre = pytesseract.image_to_string(new_image_files[683:703, 453:509])
                print(img_erv_pref_pre)
                img_rv_ref = pytesseract.image_to_string(new_image_files[704:724, 305:357])
                print(img_rv_ref)
                img_rv_pre = pytesseract.image_to_string(new_image_files[704:724, 365:429])
                print(img_rv_pre)
                img_rv_pref_pre = pytesseract.image_to_string(new_image_files[704:724, 453:509])
                print(img_rv_pref_pre)
                img_rvdtlc_ref = pytesseract.image_to_string(new_image_files[722:742, 305:357])
                print(img_rvdtlc_ref)
                img_rvdtlc_pre = pytesseract.image_to_string(new_image_files[722:742, 365:429])
                print(img_rvdtlc_pre)
                img_vtg_pre = pytesseract.image_to_string(new_image_files[741:764, 378:442])
                print(img_vtg_pre)
                img_vt_pre = pytesseract.image_to_string(new_image_files[761:780, 387:433])
                print(img_vt_pre)
                img_dlco_ref = pytesseract.image_to_string(new_image_files[818:840, 305:357])
                print(img_dlco_ref)
                img_dlco_pre = pytesseract.image_to_string(new_image_files[818:840, 365:429])
                print(img_dlco_pre)
                img_dlco_pref_pre = pytesseract.image_to_string(new_image_files[818:840, 453:509])
                print(img_dlco_pref_pre)
                img_dladj_ref = pytesseract.image_to_string(new_image_files[841:861, 305:357])
                print(img_dladj_ref)
                img_dladj_pre = pytesseract.image_to_string(new_image_files[841:861, 365:429])
                print(img_dladj_pre)
                img_dladj_pref_pre = pytesseract.image_to_string(new_image_files[841:861, 453:509])
                print(img_dladj_pref_pre)
                img_dlcodva_ref = pytesseract.image_to_string(new_image_files[859:878, 305:357])
                print(img_dlcodva_ref)
                img_dlcodva_pre = pytesseract.image_to_string(new_image_files[859:878, 365:429])
                print(img_dlcodva_pre)
                img_dlcodva_pref_pre = pytesseract.image_to_string(new_image_files[859:878, 453:509])
                print(img_dlcodva_pref_pre)
                img_dldvaadj_ref = pytesseract.image_to_string(new_image_files[878:898, 305:357])
                print(img_dldvaadj_ref)
                img_dldvaadj_pre = pytesseract.image_to_string(new_image_files[878:898, 365:429])
                print(img_dldvaadj_pre)
                img_dldvaadj_pref_pre = pytesseract.image_to_string(new_image_files[878:898, 453:509])
                print(img_dldvaadj_pref_pre)
                img_va_pre = pytesseract.image_to_string(new_image_files[739:765, 382:429])
                print(img_va_pre)
                img_ivc_pre = pytesseract.image_to_string(new_image_files[916:939, 382:436])
                print(img_ivc_pre)
                img_rawtotal_pre = pytesseract.image_to_string(new_image_files[972:996, 382:440])
                print(img_rawtotal_pre)
                img_rawinsp_pre = pytesseract.image_to_string(new_image_files[996:1016, 387:431])
                print(img_rawinsp_pre)
                img_rawexp_pre = pytesseract.image_to_string(new_image_files[1016:1037, 288:429])
                print(img_rawexp_pre)
                img_raw_ref = pytesseract.image_to_string(new_image_files[1031:1055, 305:357])
                print(img_raw_ref)
                img_raw_pre = pytesseract.image_to_string(new_image_files[1031:1055, 365:429])
                print(img_raw_pre)
                img_raw_pref_pre = pytesseract.image_to_string(new_image_files[1031:1055, 453:509])
                print(img_raw_pref_pre)
                img_gaw_ref = pytesseract.image_to_string(new_image_files[1050:1075, 305:357])
                print(img_gaw_ref)
                img_gaw_pre = pytesseract.image_to_string(new_image_files[1050:1075, 365:429])
                print(img_gaw_pre)
                img_gaw_pref_pre = pytesseract.image_to_string(new_image_files[1050:1075, 453:509])
                print(img_gaw_pref_pre)
                img_sraw_ref = pytesseract.image_to_string(new_image_files[1071:1094, 305:357])
                print(img_sraw_ref)
                img_sraw_pre = pytesseract.image_to_string(new_image_files[1071:1094, 365:429])
                print(img_sraw_pre)
                img_sraw_pref_pre = pytesseract.image_to_string(new_image_files[1071:1094, 453:509])
                print(img_sraw_pref_pre)
                img_sgaw_ref = pytesseract.image_to_string(new_image_files[1089:1113, 305:357])
                print(img_sgaw_ref)
                img_sgaw_pre = pytesseract.image_to_string(new_image_files[1089:1113, 365:429])
                print(img_sgaw_pre)
                img_sgaw_pref_pre = pytesseract.image_to_string(new_image_files[1089:1113, 453:509])
                print(img_sgaw_pref_pre)
                img_rawvtg_pre = pytesseract.image_to_string(new_image_files[1112:1133, 379:435])
                print(img_rawvtg_pre)
                img_rawf_pre = pytesseract.image_to_string(new_image_files[1131:1154, 384:445])
                print(img_rawf_pre)
                type04_img_cnt += 1
                pass

            elif pytesseract.image_to_string(new_image_files[6:40, 245:367]) == 'CATHOLIC' and pytesseract.image_to_string(new_image_files[323:596, 657:914]) == '':
                img_type = 'type05'
                img_pid = pytesseract.image_to_string(new_image_files[120:150, 605:687])
                print(img_pid)
                img_date = pytesseract.image_to_string(new_image_files[146:175, 625:707])
                print(img_date)
                img_age = pytesseract.image_to_string(new_image_files[170:203, 91:131])
                print(img_age)
                img_height = pytesseract.image_to_string(new_image_files[197:225, 154:191])
                print(img_height)
                img_weight = pytesseract.image_to_string(new_image_files[196:224, 336:385])
                print(img_weight)
                img_gender = pytesseract.image_to_string(new_image_files[148:175, 123:193])
                print(img_gender)
                img_fvc_pre = pytesseract.image_to_string(new_image_files[331:355, 481:524])
                print(img_fvc_pre)
                img_fvc_pred = pytesseract.image_to_string(new_image_files[331:355, 344:402])
                print(img_fvc_pred)
                img_fvc_pre_ppred = pytesseract.image_to_string(new_image_files[331:355, 548:597])
                print(img_fvc_pre_ppred)
                img_fev1_pre = pytesseract.image_to_string(new_image_files[355:374, 481:524])
                print(img_fev1_pre)
                img_fev1_pred = pytesseract.image_to_string(new_image_files[355:374, 344:402])
                print(img_fev1_pred)
                img_fev1_pre_ppred = pytesseract.image_to_string(new_image_files[355:374, 548:597])
                print(img_fev1_pre_ppred)
                img_fev1dfvc_pre = pytesseract.image_to_string(new_image_files[370:391, 481:524])
                print(img_fev1dfvc_pre)
                img_fev1dfvc_pred = pytesseract.image_to_string(new_image_files[370:391, 344:402])
                print(img_fev1dfvc_pred)
                img_fef25_75_pre = pytesseract.image_to_string(new_image_files[399:408, 481:524])
                print(img_fef25_75_pre)
                img_fef25_75_pred = pytesseract.image_to_string(new_image_files[399:408, 344:402])
                print(img_fef25_75_pred)
                img_fef25_75_pre_ppred = pytesseract.image_to_string(new_image_files[399:408, 548:597])
                print(img_fef25_75_pre_ppred)
                img_isofef25_75_pred = pytesseract.image_to_string(new_image_files[409:427, 344:402])
                print(img_isofef25_75_pred)
                img_isofef25_75_pre = pytesseract.image_to_string(new_image_files[409:427, 481:524])
                print(img_isofef25_75_pre)
                img_isofef25_75_pre_ppred = pytesseract.image_to_string(new_image_files[409:427, 548:597])
                print(img_isofef25_75_pre_ppred)
                img_fef75_85_pre = pytesseract.image_to_string(new_image_files[421:444, 481:524])
                print(img_fef75_85_pre)
                img_fef75_85_pred = pytesseract.image_to_string(new_image_files[421:444, 344:402])
                print(img_fef75_85_pred)
                img_fef75_85_pre_ppred = pytesseract.image_to_string(new_image_files[421:444, 548:597])
                print(img_fef75_85_pre_ppred)
                img_pef_pre = pytesseract.image_to_string(new_image_files[443:461, 481:524])
                print(img_pef_pre)
                img_pef_pred = pytesseract.image_to_string(new_image_files[443:461, 344:402])
                print(img_pef_pred)
                img_pef_pre_ppred = pytesseract.image_to_string(new_image_files[443:461, 548:597])
                print(img_pef_pre_ppred)
                img_peft_pre = pytesseract.image_to_string(new_image_files[457:479, 485:528])
                print(img_peft_pre)
                img_fet100_pre = pytesseract.image_to_string(new_image_files[480:498, 487:533])
                print(img_fet100_pre)
                img_fivc_pre = pytesseract.image_to_string(new_image_files[501:520, 481:524])
                print(img_fivc_pre)
                img_fivc_pred = pytesseract.image_to_string(new_image_files[501:520, 344:402])
                print(img_fivc_pred)
                img_fivc_pre_ppred = pytesseract.image_to_string(new_image_files[501:520, 548:597])
                print(img_fivc_pre_ppred)
                img_fefdfif50_pred = pytesseract.image_to_string(new_image_files[534:556, 344:402])
                print(img_fefdfif50_pred)
                img_fefdfif50_pre = pytesseract.image_to_string(new_image_files[534:556, 481:524])
                print(img_fefdfif50_pre)
                img_volextrap_pre = pytesseract.image_to_string(new_image_files[555:574, 487:524])
                print(img_volextrap_pre)
                img_fvlecode_pre = pytesseract.image_to_string(new_image_files[572:588, 464:528])
                print(img_fvlecode_pre)
                img_mvv_pred = pytesseract.image_to_string(new_image_files[584:612, 357:404])
                print(img_mvv_pred)
                type05_img_cnt += 1
                pass

            elif pytesseract.image_to_string(new_image_files[6:40, 245:367]) == 'CATHOLIC':
                img_type = 'type06'
                img_pid = pytesseract.image_to_string(new_image_files[120:150, 605:687])
                print(img_pid)
                img_date = pytesseract.image_to_string(new_image_files[146:175, 625:707])
                print(img_date)
                img_age = pytesseract.image_to_string(new_image_files[170:203, 91:131])
                print(img_age)
                img_height = pytesseract.image_to_string(new_image_files[197:225, 154:191])
                print(img_height)
                img_weight = pytesseract.image_to_string(new_image_files[196:224, 336:385])
                print(img_weight)
                img_gender = pytesseract.image_to_string(new_image_files[148:175, 123:193])
                print(img_gender)
                img_fvc_pre = pytesseract.image_to_string(new_image_files[314:338, 469:524])
                print(img_fvc_pre)
                img_fvc_post = pytesseract.image_to_string(new_image_files[314:338, 654:718])
                print(img_fvc_post)
                img_fvc_pchg = pytesseract.image_to_string(new_image_files[314:338, 883:908])
                print(img_fvc_pchg)
                img_fvc_pred = pytesseract.image_to_string(new_image_files[314:338, 345:401])
                print(img_fvc_pred)
                img_fvc_pre_ppred = pytesseract.image_to_string(new_image_files[314:338, 561:597])
                print(img_fvc_pre_ppred)
                img_fvc_post_ppred = pytesseract.image_to_string(new_image_files[314:338, 752:791])
                print(img_fvc_post_ppred)
                img_fev1_pre = pytesseract.image_to_string(new_image_files[334:358, 469:524])
                print(img_fev1_pre)
                img_fev1_post = pytesseract.image_to_string(new_image_files[334:358, 654:718])
                print(img_fev1_post)
                img_fev1_pchg = pytesseract.image_to_string(new_image_files[334:358, 883:908])
                print(img_fev1_pchg)
                img_fev1_pred = pytesseract.image_to_string(new_image_files[334:358, 345:401])
                print(img_fev1_pred)
                img_fev1_pre_ppred = pytesseract.image_to_string(new_image_files[334:358, 561:597])
                print(img_fev1_pre_ppred)
                img_fev1_post_ppred = pytesseract.image_to_string(new_image_files[334:358, 752:791])
                print(img_fev1_post_ppred)
                img_fev1dfvc_pre = pytesseract.image_to_string(new_image_files[354:373, 469:524])
                print(img_fev1dfvc_pre)
                img_fev1dfvc_post = pytesseract.image_to_string(new_image_files[354:373, 654:718])
                print(img_fev1dfvc_post)
                img_fev1dfvc_pred = pytesseract.image_to_string(new_image_files[354:373, 345:401])
                print(img_fev1dfvc_pred)
                img_fef25_75_pre = pytesseract.image_to_string(new_image_files[372:391, 469:524])
                print(img_fef25_75_pre)
                img_fef25_75_post = pytesseract.image_to_string(new_image_files[372:391, 654:718])
                print(img_fef25_75_post)
                img_fef25_75_pchg = pytesseract.image_to_string(new_image_files[372:391, 883:908])
                print(img_fef25_75_pchg)
                img_fef25_75_pred = pytesseract.image_to_string(new_image_files[372:391, 345:401])
                print(img_fef25_75_pred)
                img_fef25_75_pre_ppred = pytesseract.image_to_string(new_image_files[372:391, 561:597])
                print(img_fef25_75_pre_ppred)
                img_fef25_75_post_ppred = pytesseract.image_to_string(new_image_files[372:391, 752:791])
                print(img_fef25_75_post_ppred)
                img_isofef25_75_pchg = pytesseract.image_to_string(new_image_files[391:410, 883:908])
                print(img_isofef25_75_pchg)
                img_isofef25_75_pred = pytesseract.image_to_string(new_image_files[391:410, 345:401])
                print(img_isofef25_75_pred)
                img_isofef25_75_pre = pytesseract.image_to_string(new_image_files[391:410, 469:524])
                print(img_isofef25_75_pre)
                img_isofef25_75_pre_ppred = pytesseract.image_to_string(new_image_files[391:410, 561:597])
                print(img_isofef25_75_pre_ppred)
                img_isofef25_75_post = pytesseract.image_to_string(new_image_files[391:410, 654:718])
                print(img_isofef25_75_post)
                img_isofef25_75_post_ppred = pytesseract.image_to_string(new_image_files[391:410, 752:791])
                print(img_isofef25_75_post_ppred)
                img_fef75_85_pre = pytesseract.image_to_string(new_image_files[407:427, 469:524])
                print(img_fef75_85_pre)
                img_fef75_85_post = pytesseract.image_to_string(new_image_files[407:427, 654:718])
                print(img_fef75_85_post)
                img_fef75_85_pchg = pytesseract.image_to_string(new_image_files[407:427, 883:908])
                print(img_fef75_85_pchg)
                img_fef75_85_pred = pytesseract.image_to_string(new_image_files[407:427, 345:401])
                print(img_fef75_85_pred)
                img_fef75_85_pre_ppred = pytesseract.image_to_string(new_image_files[407:427, 561:597])
                print(img_fef75_85_pre_ppred)
                img_fef75_85_post_ppred = pytesseract.image_to_string(new_image_files[407:427, 752:791])
                print(img_fef75_85_post_ppred)
                img_pef_pre = pytesseract.image_to_string(new_image_files[424:446, 469:524])
                print(img_pef_pre)
                img_pef_post = pytesseract.image_to_string(new_image_files[424:446, 654:718])
                print(img_pef_post)
                img_pef_pchg = pytesseract.image_to_string(new_image_files[424:446, 883:908])
                print(img_pef_pchg)
                img_pef_pred = pytesseract.image_to_string(new_image_files[424:446, 345:401])
                print(img_pef_pred)
                img_pef_pre_ppred = pytesseract.image_to_string(new_image_files[424:446, 561:597])
                print(img_pef_pre_ppred)
                img_pef_post_ppred = pytesseract.image_to_string(new_image_files[424:446, 752:791])
                print(img_pef_post_ppred)
                img_fet100_pchg = pytesseract.image_to_string(new_image_files[442:462, 883:908])
                print(img_fet100_pchg)
                img_fet100_pre = pytesseract.image_to_string(new_image_files[442:462, 469:524])
                print(img_fet100_pre)
                img_fet100_post = pytesseract.image_to_string(new_image_files[442:462, 654:718])
                print(img_fet100_post)
                img_fivc_pre = pytesseract.image_to_string(new_image_files[459:482, 469:524])
                print(img_fivc_pre)
                img_fivc_post = pytesseract.image_to_string(new_image_files[459:482, 654:718])
                print(img_fivc_post)
                img_fivc_pchg = pytesseract.image_to_string(new_image_files[459:482, 883:908])
                print(img_fivc_pchg)
                img_fivc_pred = pytesseract.image_to_string(new_image_files[459:482, 345:401])
                print(img_fivc_pred)
                img_fivc_pre_ppred = pytesseract.image_to_string(new_image_files[459:482, 561:597])
                print(img_fivc_pre_ppred)
                img_fivc_post_ppred = pytesseract.image_to_string(new_image_files[459:482, 752:791])
                print(img_fivc_post_ppred)
                img_fiv1_pchg = pytesseract.image_to_string(new_image_files[525:542, 883:908])
                print(img_fiv1_pchg)
                img_fiv1_pre = pytesseract.image_to_string(new_image_files[525:542, 469:524])
                print(img_fiv1_pre)
                img_fiv1_post = pytesseract.image_to_string(new_image_files[525:542, 654:718])
                print(img_fiv1_post)
                img_fefdfif50_pchg = pytesseract.image_to_string(new_image_files[541:562, 883:908])
                print(img_fefdfif50_pchg)
                img_fefdfif50_pred = pytesseract.image_to_string(new_image_files[541:562, 345:401])
                print(img_fefdfif50_pred)
                img_fefdfif50_pre = pytesseract.image_to_string(new_image_files[541:562, 469:524])
                print(img_fefdfif50_pre)
                img_fefdfif50_pre_ppred = pytesseract.image_to_string(new_image_files[541:562, 561:597])
                print(img_fefdfif50_pre_ppred)
                img_fefdfif50_post = pytesseract.image_to_string(new_image_files[541:562, 654:718])
                print(img_fefdfif50_post)
                img_fefdfif50_post_ppred = pytesseract.image_to_string(new_image_files[541:562, 752:791])
                print(img_fefdfif50_post_ppred)
                img_volextrap_pchg = pytesseract.image_to_string(new_image_files[559:580, 883:908])
                print(img_volextrap_pchg)
                img_volextrap_pre = pytesseract.image_to_string(new_image_files[559:580, 469:524])
                print(img_volextrap_pre)
                img_volextrap_post = pytesseract.image_to_string(new_image_files[559:580, 654:718])
                print(img_volextrap_post)
                img_fvlecode_pre = pytesseract.image_to_string(new_image_files[575:597, 469:524])
                print(img_fvlecode_pre)
                img_fvlecode_post = pytesseract.image_to_string(new_image_files[575:597, 654:718])
                print(img_fvlecode_post)
                img_mvv_pred = pytesseract.image_to_string(new_image_files[606:634, 357:412])
                print(img_mvv_pred)
                type06_img_cnt += 1
                pass

            elif pytesseract.image_to_string(new_image_files[66:99, 538:647]) == 'REPORT':
                img_type = 'type07'
                img_pid = pytesseract.image_to_string(new_image_files[94:121, 601:693])
                print(img_pid)
                img_date = pytesseract.image_to_string(new_image_files[118:142, 623:706])
                print(img_date)
                img_age = pytesseract.image_to_string(new_image_files[139:161, 91:123])
                print(img_age)
                img_height = pytesseract.image_to_string(new_image_files[154:186, 150:191])
                print(img_height)
                img_weight = pytesseract.image_to_string(new_image_files[159:183, 301:357])
                print(img_weight)
                img_gender = pytesseract.image_to_string(new_image_files[120:145, 119:199])
                print(img_gender)
                img_fvc_ref = pytesseract.image_to_string(new_image_files[249:271, 315:357])
                print(img_fvc_ref)
                img_fvc_pre = pytesseract.image_to_string(new_image_files[249:271, 363:421])
                print(img_fvc_pre)
                img_fvc_post = pytesseract.image_to_string(new_image_files[249:271, 491:551])
                print(img_fvc_post)
                img_fvc_tri1 = pytesseract.image_to_string(new_image_files[468:494, 315:357])
                print(img_fvc_tri1)
                img_fvc_tri2 = pytesseract.image_to_string(new_image_files[468:494, 363:421])
                print(img_fvc_tri2)
                img_fvc_tri3 = pytesseract.image_to_string(new_image_files[468:494, 445:486])
                print(img_fvc_tri3)
                img_fvc_tri4 = pytesseract.image_to_string(new_image_files[468:494, 491:551])
                print(img_fvc_tri4)
                img_fvc_tri5 = pytesseract.image_to_string(new_image_files[468:494, 572:615])
                print(img_fvc_tri5)
                img_fvc_tri6 = pytesseract.image_to_string(new_image_files[468:494, 630:682])
                print(img_fvc_tri6)
                img_fvc_tri7 = pytesseract.image_to_string(new_image_files[468:494, 696:743])
                print(img_fvc_tri7)
                img_fvc_tri8 = pytesseract.image_to_string(new_image_files[468:494, 764:805])
                print(img_fvc_tri8)
                img_fvc_pref_pre = pytesseract.image_to_string(new_image_files[249:271, 445:486])
                print(img_fvc_pref_pre)
                img_fvc_pref_post = pytesseract.image_to_string(new_image_files[249:271, 572:615])
                print(img_fvc_pref_post)
                img_fvc_pchg = pytesseract.image_to_string(new_image_files[249:271, 630:682])
                print(img_fvc_pchg)
                img_fev1_ref = pytesseract.image_to_string(new_image_files[270:289, 315:357])
                print(img_fev1_ref)
                img_fev1_pre = pytesseract.image_to_string(new_image_files[270:289, 363:421])
                print(img_fev1_pre)
                img_fev1_post = pytesseract.image_to_string(new_image_files[270:289, 491:551])
                print(img_fev1_post)
                img_fev1_tri1 = pytesseract.image_to_string(new_image_files[488:513, 315:357])
                print(img_fev1_tri1)
                img_fev1_tri2 = pytesseract.image_to_string(new_image_files[488:513, 363:421])
                print(img_fev1_tri2)
                img_fev1_tri3 = pytesseract.image_to_string(new_image_files[488:513, 445:486])
                print(img_fev1_tri3)
                img_fev1_tri4 = pytesseract.image_to_string(new_image_files[488:513, 491:551])
                print(img_fev1_tri4)
                img_fev1_tri5 = pytesseract.image_to_string(new_image_files[488:513, 572:615])
                print(img_fev1_tri5)
                img_fev1_tri6 = pytesseract.image_to_string(new_image_files[488:513, 630:682])
                print(img_fev1_tri6)
                img_fev1_tri7 = pytesseract.image_to_string(new_image_files[488:513, 696:743])
                print(img_fev1_tri7)
                img_fev1_tri8 = pytesseract.image_to_string(new_image_files[488:513, 764:805])
                print(img_fev1_tri8)
                img_fev1_pref_pre = pytesseract.image_to_string(new_image_files[270:289, 445:486])
                print(img_fev1_pref_pre)
                img_fev1_pref_post = pytesseract.image_to_string(new_image_files[270:289, 572:615])
                print(img_fev1_pref_post)
                img_fev1_pchg = pytesseract.image_to_string(new_image_files[270:289, 630:682])
                print(img_fev1_pchg)
                img_fev1dfvc_ref = pytesseract.image_to_string(new_image_files[289:306, 315:357])
                print(img_fev1dfvc_ref)
                img_fev1dfvc_pre = pytesseract.image_to_string(new_image_files[289:306, 363:421])
                print(img_fev1dfvc_pre)
                img_fev1dfvc_post = pytesseract.image_to_string(new_image_files[289:306, 491:551])
                print(img_fev1dfvc_post)
                img_fev1dfvc_tri1 = pytesseract.image_to_string(new_image_files[509:529, 315:357])
                print(img_fev1dfvc_tri1)
                img_fev1dfvc_tri2 = pytesseract.image_to_string(new_image_files[509:529, 363:421])
                print(img_fev1dfvc_tri2)
                img_fev1dfvc_tri3 = pytesseract.image_to_string(new_image_files[509:529, 445:486])
                print(img_fev1dfvc_tri3)
                img_fev1dfvc_tri4 = pytesseract.image_to_string(new_image_files[509:529, 491:551])
                print(img_fev1dfvc_tri4)
                img_fev1dfvc_tri5 = pytesseract.image_to_string(new_image_files[509:529, 572:615])
                print(img_fev1dfvc_tri5)
                img_fev1dfvc_tri6 = pytesseract.image_to_string(new_image_files[509:529, 630:682])
                print(img_fev1dfvc_tri6)
                img_fev1dfvc_tri7 = pytesseract.image_to_string(new_image_files[509:529, 696:743])
                print(img_fev1dfvc_tri7)
                img_fev1dfvc_tri8 = pytesseract.image_to_string(new_image_files[509:529, 764:805])
                print(img_fev1dfvc_tri8)
                img_fef25_75_ref = pytesseract.image_to_string(new_image_files[301:326, 315:357])
                print(img_fef25_75_ref)
                img_fef25_75_pre = pytesseract.image_to_string(new_image_files[301:326, 363:421])
                print(img_fef25_75_pre)
                img_fef25_75_post = pytesseract.image_to_string(new_image_files[301:326, 491:551])
                print(img_fef25_75_post)
                img_fef25_75_tri1 = pytesseract.image_to_string(new_image_files[526:548, 315:357])
                print(img_fef25_75_tri1)
                img_fef25_75_tri2 = pytesseract.image_to_string(new_image_files[526:548, 363:421])
                print(img_fef25_75_tri2)
                img_fef25_75_tri3 = pytesseract.image_to_string(new_image_files[526:548, 445:486])
                print(img_fef25_75_tri3)
                img_fef25_75_tri4 = pytesseract.image_to_string(new_image_files[526:548, 491:551])
                print(img_fef25_75_tri4)
                img_fef25_75_tri5 = pytesseract.image_to_string(new_image_files[526:548, 572:615])
                print(img_fef25_75_tri5)
                img_fef25_75_tri6 = pytesseract.image_to_string(new_image_files[526:548, 630:682])
                print(img_fef25_75_tri6)
                img_fef25_75_tri7 = pytesseract.image_to_string(new_image_files[526:548, 696:743])
                print(img_fef25_75_tri7)
                img_fef25_75_tri8 = pytesseract.image_to_string(new_image_files[526:548, 764:805])
                print(img_fef25_75_tri8)
                img_fef25_75_pref_pre = pytesseract.image_to_string(new_image_files[301:326, 445:486])
                print(img_fef25_75_pref_pre)
                img_fef25_75_pref_post = pytesseract.image_to_string(new_image_files[301:326, 572:615])
                print(img_fef25_75_pref_post)
                img_fef25_75_pchg = pytesseract.image_to_string(new_image_files[301:326, 630:682])
                print(img_fef25_75_pchg)
                img_pef_ref = pytesseract.image_to_string(new_image_files[322:344, 315:357])
                print(img_pef_ref)
                img_pef_pre = pytesseract.image_to_string(new_image_files[322:344, 363:421])
                print(img_pef_pre)
                img_pef_post = pytesseract.image_to_string(new_image_files[322:344, 491:551])
                print(img_pef_post)
                img_pef_tri1 = pytesseract.image_to_string(new_image_files[546:565, 315:357])
                print(img_pef_tri1)
                img_pef_tri2 = pytesseract.image_to_string(new_image_files[546:565, 363:421])
                print(img_pef_tri2)
                img_pef_tri3 = pytesseract.image_to_string(new_image_files[546:565, 445:486])
                print(img_pef_tri3)
                img_pef_tri4 = pytesseract.image_to_string(new_image_files[546:565, 491:551])
                print(img_pef_tri4)
                img_pef_tri5 = pytesseract.image_to_string(new_image_files[546:565, 572:615])
                print(img_pef_tri5)
                img_pef_tri6 = pytesseract.image_to_string(new_image_files[546:565, 630:682])
                print(img_pef_tri6)
                img_pef_tri7 = pytesseract.image_to_string(new_image_files[546:565, 696:743])
                print(img_pef_tri7)
                img_pef_tri8 = pytesseract.image_to_string(new_image_files[546:565, 764:805])
                print(img_pef_tri8)
                img_pef_pref_pre = pytesseract.image_to_string(new_image_files[322:344, 445:486])
                print(img_pef_pref_pre)
                img_pef_pref_post = pytesseract.image_to_string(new_image_files[322:344, 572:615])
                print(img_pef_pref_post)
                img_pef_pchg = pytesseract.image_to_string(new_image_files[322:344, 630:682])
                print(img_pef_pchg)
                img_peft_pchg = pytesseract.image_to_string(new_image_files[340:360, 630:682])
                print(img_peft_pchg)
                img_peft_pre = pytesseract.image_to_string(new_image_files[340:360, 363:421])
                print(img_peft_pre)
                img_peft_post = pytesseract.image_to_string(new_image_files[340:360, 491:551])
                print(img_peft_post)
                img_peft_tri1 = pytesseract.image_to_string(new_image_files[562:584, 315:357])
                print(img_peft_tri1)
                img_peft_tri2 = pytesseract.image_to_string(new_image_files[562:584, 363:421])
                print(img_peft_tri2)
                img_peft_tri3 = pytesseract.image_to_string(new_image_files[562:584, 445:486])
                print(img_peft_tri3)
                img_peft_tri4 = pytesseract.image_to_string(new_image_files[562:584, 491:551])
                print(img_peft_tri4)
                img_peft_tri5 = pytesseract.image_to_string(new_image_files[562:584, 572:615])
                print(img_peft_tri5)
                img_peft_tri6 = pytesseract.image_to_string(new_image_files[562:584, 630:682])
                print(img_peft_tri6)
                img_peft_tri7 = pytesseract.image_to_string(new_image_files[562:584, 696:743])
                print(img_peft_tri7)
                img_peft_tri8 = pytesseract.image_to_string(new_image_files[562:584, 764:805])
                print(img_peft_tri8)
                img_fet100_pchg = pytesseract.image_to_string(new_image_files[358:378, 630:682])
                print(img_fet100_pchg)
                img_fet100_pre = pytesseract.image_to_string(new_image_files[358:378, 363:421])
                print(img_fet100_pre)
                img_fet100_post = pytesseract.image_to_string(new_image_files[358:378, 491:551])
                print(img_fet100_post)
                img_fet100_tri1 = pytesseract.image_to_string(new_image_files[580:600, 315:357])
                print(img_fet100_tri1)
                img_fet100_tri2 = pytesseract.image_to_string(new_image_files[580:600, 363:421])
                print(img_fet100_tri2)
                img_fet100_tri3 = pytesseract.image_to_string(new_image_files[580:600, 445:486])
                print(img_fet100_tri3)
                img_fet100_tri4 = pytesseract.image_to_string(new_image_files[580:600, 491:551])
                print(img_fet100_tri4)
                img_fet100_tri5 = pytesseract.image_to_string(new_image_files[580:600, 572:615])
                print(img_fet100_tri5)
                img_fet100_tri6 = pytesseract.image_to_string(new_image_files[580:600, 630:682])
                print(img_fet100_tri6)
                img_fet100_tri7 = pytesseract.image_to_string(new_image_files[580:600, 696:743])
                print(img_fet100_tri7)
                img_fet100_tri8 = pytesseract.image_to_string(new_image_files[580:600, 764:805])
                print(img_fet100_tri8)
                img_fivc_ref = pytesseract.image_to_string(new_image_files[376:398, 315:357])
                print(img_fivc_ref)
                img_fivc_pre = pytesseract.image_to_string(new_image_files[376:398, 363:421])
                print(img_fivc_pre)
                img_fivc_post = pytesseract.image_to_string(new_image_files[376:398, 491:551])
                print(img_fivc_post)
                img_fivc_pref_pre = pytesseract.image_to_string(new_image_files[376:398, 445:486])
                print(img_fivc_pref_pre)
                img_fivc_pref_post = pytesseract.image_to_string(new_image_files[376:398, 572:615])
                print(img_fivc_pref_post)
                img_fivc_pchg = pytesseract.image_to_string(new_image_files[376:398, 630:682])
                print(img_fivc_pchg)
                img_fivc_tri1 = pytesseract.image_to_string(new_image_files[598:618, 315:357])
                print(img_fivc_tri1)
                img_fivc_tri2 = pytesseract.image_to_string(new_image_files[598:618, 363:421])
                print(img_fivc_tri2)
                img_fivc_tri3 = pytesseract.image_to_string(new_image_files[598:618, 445:486])
                print(img_fivc_tri3)
                img_fivc_tri4 = pytesseract.image_to_string(new_image_files[598:618, 491:551])
                print(img_fivc_tri4)
                img_fivc_tri5 = pytesseract.image_to_string(new_image_files[598:618, 572:615])
                print(img_fivc_tri5)
                img_fivc_tri6 = pytesseract.image_to_string(new_image_files[598:618, 630:682])
                print(img_fivc_tri6)
                img_fivc_tri7 = pytesseract.image_to_string(new_image_files[598:618, 696:743])
                print(img_fivc_tri7)
                img_fivc_tri8 = pytesseract.image_to_string(new_image_files[598:618, 764:805])
                print(img_fivc_tri8)
                img_volextrap_pchg = pytesseract.image_to_string(new_image_files[393:414, 630:682])
                print(img_volextrap_pchg)
                img_volextrap_pre = pytesseract.image_to_string(new_image_files[393:414, 363:421])
                print(img_volextrap_pre)
                img_volextrap_post = pytesseract.image_to_string(new_image_files[393:414, 491:551])
                print(img_volextrap_post)
                img_volextrap_tri1 = pytesseract.image_to_string(new_image_files[614:636, 315:357])
                print(img_volextrap_tri1)
                img_volextrap_tri2 = pytesseract.image_to_string(new_image_files[614:636, 363:421])
                print(img_volextrap_tri2)
                img_volextrap_tri3 = pytesseract.image_to_string(new_image_files[614:636, 445:486])
                print(img_volextrap_tri3)
                img_volextrap_tri4 = pytesseract.image_to_string(new_image_files[614:636, 491:551])
                print(img_volextrap_tri4)
                img_volextrap_tri5 = pytesseract.image_to_string(new_image_files[614:636, 572:615])
                print(img_volextrap_tri5)
                img_volextrap_tri6 = pytesseract.image_to_string(new_image_files[614:636, 630:682])
                print(img_volextrap_tri6)
                img_volextrap_tri7 = pytesseract.image_to_string(new_image_files[614:636, 696:743])
                print(img_volextrap_tri7)
                img_volextrap_tri8 = pytesseract.image_to_string(new_image_files[614:636, 764:805])
                print(img_volextrap_tri8)
                img_fvlecode_pre = pytesseract.image_to_string(new_image_files[412:434, 363:421])
                print(img_fvlecode_pre)
                img_fvlecode_post = pytesseract.image_to_string(new_image_files[412:434, 491:551])
                print(img_fvlecode_post)
                img_fvlecode_tri1 = pytesseract.image_to_string(new_image_files[635:650, 315:357])
                print(img_fvlecode_tri1)
                img_fvlecode_tri2 = pytesseract.image_to_string(new_image_files[635:650, 363:421])
                print(img_fvlecode_tri2)
                img_fvlecode_tri3 = pytesseract.image_to_string(new_image_files[635:650, 445:486])
                print(img_fvlecode_tri3)
                img_fvlecode_tri4 = pytesseract.image_to_string(new_image_files[635:650, 491:551])
                print(img_fvlecode_tri4)
                img_fvlecode_tri5 = pytesseract.image_to_string(new_image_files[635:650, 572:615])
                print(img_fvlecode_tri5)
                img_fvlecode_tri6 = pytesseract.image_to_string(new_image_files[635:650, 630:682])
                print(img_fvlecode_tri6)
                img_fvlecode_tri7 = pytesseract.image_to_string(new_image_files[635:650, 696:743])
                print(img_fvlecode_tri7)
                img_fvlecode_tri8 = pytesseract.image_to_string(new_image_files[635:650, 764:805])
                print(img_fvlecode_tri8)
                type07_img_cnt += 1
                pass

            elif pytesseract.image_to_string(new_image_files[266:303, 3:44]) == 'Lung':
                img_type = 'type08'
                img_pid = pytesseract.image_to_string(new_image_files[55:85, 685:770])
                print(img_pid)
                img_date = pytesseract.image_to_string(new_image_files[3:36, 708:800])
                print(img_date)
                img_age = pytesseract.image_to_string(new_image_files[83:108, 698:728])
                print(img_age)
                img_height = pytesseract.image_to_string(new_image_files[79:110, 862:904])
                print(img_height)
                img_weight = pytesseract.image_to_string(new_image_files[108:134, 758:805])
                print(img_weight)
                img_gender = pytesseract.image_to_string(new_image_files[106:137, 882:959])
                print(img_gender)
                img_tlc_ref = pytesseract.image_to_string(new_image_files[290:315, 308:360])
                print(img_tlc_ref)
                img_tlc_pre = pytesseract.image_to_string(new_image_files[290:315, 376:432])
                print(img_tlc_pre)
                img_tlc_pref_pre = pytesseract.image_to_string(new_image_files[290:315, 451:503])
                print(img_tlc_pref_pre)
                img_vc_ref = pytesseract.image_to_string(new_image_files[314:335, 308:360])
                print(img_vc_ref)
                img_vc_pre = pytesseract.image_to_string(new_image_files[314:335, 376:432])
                print(img_vc_pre)
                img_vc_pref_pre = pytesseract.image_to_string(new_image_files[314:335, 451:503])
                print(img_vc_pref_pre)
                img_ic_ref = pytesseract.image_to_string(new_image_files[334:355, 308:360])
                print(img_ic_ref)
                img_ic_pre = pytesseract.image_to_string(new_image_files[334:355, 376:432])
                print(img_ic_pre)
                img_ic_pref_pre = pytesseract.image_to_string(new_image_files[334:355, 451:503])
                print(img_ic_pref_pre)
                img_frcpl_ref = pytesseract.image_to_string(new_image_files[353:376, 308:360])
                print(img_frcpl_ref)
                img_frcpl_pre = pytesseract.image_to_string(new_image_files[353:376, 376:432])
                print(img_frcpl_pre)
                img_frcpl_pref_pre = pytesseract.image_to_string(new_image_files[353:376, 451:503])
                print(img_frcpl_pref_pre)
                img_erv_ref = pytesseract.image_to_string(new_image_files[372:395, 308:360])
                print(img_erv_ref)
                img_erv_pre = pytesseract.image_to_string(new_image_files[372:395, 376:432])
                print(img_erv_pre)
                img_erv_pref_pre = pytesseract.image_to_string(new_image_files[372:395, 451:503])
                print(img_erv_pref_pre)
                img_rv_ref = pytesseract.image_to_string(new_image_files[393:414, 308:360])
                print(img_rv_ref)
                img_rv_pre = pytesseract.image_to_string(new_image_files[393:414, 376:432])
                print(img_rv_pre)
                img_rv_pref_pre = pytesseract.image_to_string(new_image_files[393:414, 451:503])
                print(img_rv_pref_pre)
                img_rvdtlc_ref = pytesseract.image_to_string(new_image_files[411:435, 308:360])
                print(img_rvdtlc_ref)
                img_rvdtlc_pre = pytesseract.image_to_string(new_image_files[411:435, 376:432])
                print(img_rvdtlc_pre)
                img_vtg_pre = pytesseract.image_to_string(new_image_files[432:455, 385:428])
                print(img_vtg_pre)
                img_vt_pre = pytesseract.image_to_string(new_image_files[448:478, 382:436])
                print(img_vt_pre)
                img_rawinsp_pre = pytesseract.image_to_string(new_image_files[509:530, 384:439])
                print(img_rawinsp_pre)
                img_rawexp_pre = pytesseract.image_to_string(new_image_files[530:552, 380:433])
                print(img_rawexp_pre)
                img_raw_ref = pytesseract.image_to_string(new_image_files[548:573, 308:360])
                print(img_raw_ref)
                img_raw_pre = pytesseract.image_to_string(new_image_files[548:573, 376:432])
                print(img_raw_pre)
                img_raw_pref_pre = pytesseract.image_to_string(new_image_files[548:573, 451:503])
                print(img_raw_pref_pre)
                img_gaw_ref = pytesseract.image_to_string(new_image_files[568:591, 308:360])
                print(img_gaw_ref)
                img_gaw_pre = pytesseract.image_to_string(new_image_files[568:591, 376:432])
                print(img_gaw_pre)
                img_gaw_pref_pre = pytesseract.image_to_string(new_image_files[568:591, 451:503])
                print(img_gaw_pref_pre)
                img_sraw_ref = pytesseract.image_to_string(new_image_files[587:611, 308:360])
                print(img_sraw_ref)
                img_sraw_pre = pytesseract.image_to_string(new_image_files[587:611, 376:432])
                print(img_sraw_pre)
                img_sraw_pref_pre = pytesseract.image_to_string(new_image_files[587:611, 451:503])
                print(img_sraw_pref_pre)
                img_sgaw_ref = pytesseract.image_to_string(new_image_files[607:630, 308:360])
                print(img_sgaw_ref)
                img_sgaw_pre = pytesseract.image_to_string(new_image_files[607:630, 376:432])
                print(img_sgaw_pre)
                img_sgaw_pref_pre = pytesseract.image_to_string(new_image_files[607:630, 451:503])
                print(img_sgaw_pref_pre)
                img_rawvtg_pre = pytesseract.image_to_string(new_image_files[628:648, 386:431])
                print(img_rawvtg_pre)
                img_rawf_pre = pytesseract.image_to_string(new_image_files[646:668, 391:430])
                print(img_rawf_pre)
                type08_img_cnt += 1
                pass

            else:
                img_type = 'typeUK'
                typeUK_img_cnt += 1
                pass
            
            # create new directory
            tmp_dir = new_path
            new_path += "/" + img_type #new_path = tmp_dir + "/" + img_type
            
            if os.path.exists (new_path) == 0:
                os.mkdir(new_path)

            # add data into OOMII_DB
            ws_cnt = str(ws.max_row + 1)
            print ("ws_cnt: " + ws_cnt)
            ws ["A" + ws_cnt] = img_type
            ws ["B" + ws_cnt] = img_pid
            ws ["C" + ws_cnt] = img_date
            ws ["D" + ws_cnt] = img_age
            ws ["E" + ws_cnt] = img_height
            ws ["F" + ws_cnt] = img_weight
            ws ["G" + ws_cnt] = img_gender
            ws ["H" + ws_cnt] = img_fvc_dose_lv1
            ws ["I" + ws_cnt] = img_fvc_dose_lv2
            ws ["J" + ws_cnt] = img_fvc_dose_lv3
            ws ["K" + ws_cnt] = img_fvc_dose_lv4
            ws ["L" + ws_cnt] = img_fvc_dose_lv5
            ws ["M" + ws_cnt] = img_fvc_dose_lv6
            ws ["N" + ws_cnt] = img_fvc_dose_lv7
            ws ["O" + ws_cnt] = img_fvc_dose_lv8
            ws ["P" + ws_cnt] = img_fvc_dose_lv9
            ws ["Q" + ws_cnt] = img_fvc_ref
            ws ["R" + ws_cnt] = img_fvc_pre
            ws ["S" + ws_cnt] = img_fvc_post
            ws ["T" + ws_cnt] = img_fvc_lv1
            ws ["U" + ws_cnt] = img_fvc_lv2
            ws ["V" + ws_cnt] = img_fvc_lv3
            ws ["W" + ws_cnt] = img_fvc_lv4
            ws ["X" + ws_cnt] = img_fvc_lv5
            ws ["Y" + ws_cnt] = img_fvc_lv6
            ws ["Z" + ws_cnt] = img_fvc_lv7
            ws ["AA" + ws_cnt] = img_fvc_lv8
            ws ["AB" + ws_cnt] = img_fvc_lv9
            ws ["AC" + ws_cnt] = img_fvc_tri1
            ws ["AD" + ws_cnt] = img_fvc_tri2
            ws ["AE" + ws_cnt] = img_fvc_tri3
            ws ["AF" + ws_cnt] = img_fvc_tri4
            ws ["AG" + ws_cnt] = img_fvc_tri5
            ws ["AH" + ws_cnt] = img_fvc_tri6
            ws ["AI" + ws_cnt] = img_fvc_tri7
            ws ["AJ" + ws_cnt] = img_fvc_tri8
            ws ["AK" + ws_cnt] = img_fvc_pref_pre
            ws ["AL" + ws_cnt] = img_fvc_pref_post
            ws ["AM" + ws_cnt] = img_fvc_pref_lv1
            ws ["AN" + ws_cnt] = img_fvc_pref_lv2
            ws ["AO" + ws_cnt] = img_fvc_pref_lv3
            ws ["AP" + ws_cnt] = img_fvc_pref_lv4
            ws ["AQ" + ws_cnt] = img_fvc_pref_lv5
            ws ["AR" + ws_cnt] = img_fvc_pref_lv6
            ws ["AS" + ws_cnt] = img_fvc_pref_lv7
            ws ["AT" + ws_cnt] = img_fvc_pref_lv8
            ws ["AU" + ws_cnt] = img_fvc_pref_lv9
            ws ["AV" + ws_cnt] = img_fvc_pchg
            ws ["AW" + ws_cnt] = img_fvc_pchg_lv1
            ws ["AX" + ws_cnt] = img_fvc_pchg_lv2
            ws ["AY" + ws_cnt] = img_fvc_pchg_lv3
            ws ["AZ" + ws_cnt] = img_fvc_pchg_lv4
            ws ["BA" + ws_cnt] = img_fvc_pchg_lv5
            ws ["BB" + ws_cnt] = img_fvc_pchg_lv6
            ws ["BC" + ws_cnt] = img_fvc_pchg_lv7
            ws ["BD" + ws_cnt] = img_fvc_pchg_lv8
            ws ["BE" + ws_cnt] = img_fvc_pchg_lv9
            ws ["BF" + ws_cnt] = img_fvc_pred
            ws ["BG" + ws_cnt] = img_fvc_pre_ppred
            ws ["BH" + ws_cnt] = img_fvc_post_ppred
            ws ["BI" + ws_cnt] = img_fev05_ref
            ws ["BJ" + ws_cnt] = img_fev05_pre
            ws ["BK" + ws_cnt] = img_fev05_post
            ws ["BL" + ws_cnt] = img_fev05_pref_pre
            ws ["BM" + ws_cnt] = img_fev05_pref_post
            ws ["BN" + ws_cnt] = img_fev05_pchg
            ws ["BO" + ws_cnt] = img_fev1_dose_lv1
            ws ["BP" + ws_cnt] = img_fev1_dose_lv2
            ws ["BQ" + ws_cnt] = img_fev1_dose_lv3
            ws ["BR" + ws_cnt] = img_fev1_dose_lv4
            ws ["BS" + ws_cnt] = img_fev1_dose_lv5
            ws ["BT" + ws_cnt] = img_fev1_dose_lv6
            ws ["BU" + ws_cnt] = img_fev1_dose_lv7
            ws ["BV" + ws_cnt] = img_fev1_dose_lv8
            ws ["BW" + ws_cnt] = img_fev1_dose_lv9
            ws ["BX" + ws_cnt] = img_fev1_ref
            ws ["BY" + ws_cnt] = img_fev1_pre
            ws ["BZ" + ws_cnt] = img_fev1_post
            ws ["CA" + ws_cnt] = img_fev1_lv1
            ws ["CB" + ws_cnt] = img_fev1_lv2
            ws ["CC" + ws_cnt] = img_fev1_lv3
            ws ["CD" + ws_cnt] = img_fev1_lv4
            ws ["CE" + ws_cnt] = img_fev1_lv5
            ws ["CF" + ws_cnt] = img_fev1_lv6
            ws ["CG" + ws_cnt] = img_fev1_lv7
            ws ["CH" + ws_cnt] = img_fev1_lv8
            ws ["CI" + ws_cnt] = img_fev1_lv9
            ws ["CJ" + ws_cnt] = img_fev1_tri1
            ws ["CK" + ws_cnt] = img_fev1_tri2
            ws ["CL" + ws_cnt] = img_fev1_tri3
            ws ["CM" + ws_cnt] = img_fev1_tri4
            ws ["CN" + ws_cnt] = img_fev1_tri5
            ws ["CO" + ws_cnt] = img_fev1_tri6
            ws ["CP" + ws_cnt] = img_fev1_tri7
            ws ["CQ" + ws_cnt] = img_fev1_tri8
            ws ["CR" + ws_cnt] = img_fev1_pref_pre
            ws ["CS" + ws_cnt] = img_fev1_pref_post
            ws ["CT" + ws_cnt] = img_fev1_pref_lv1
            ws ["CU" + ws_cnt] = img_fev1_pref_lv2
            ws ["CV" + ws_cnt] = img_fev1_pref_lv3
            ws ["CW" + ws_cnt] = img_fev1_pref_lv4
            ws ["CX" + ws_cnt] = img_fev1_pref_lv5
            ws ["CY" + ws_cnt] = img_fev1_pref_lv6
            ws ["CZ" + ws_cnt] = img_fev1_pref_lv7
            ws ["DA" + ws_cnt] = img_fev1_pref_lv8
            ws ["DB" + ws_cnt] = img_fev1_pref_lv9
            ws ["DC" + ws_cnt] = img_fev1_pchg
            ws ["DD" + ws_cnt] = img_fev1_pchg_lv1
            ws ["DE" + ws_cnt] = img_fev1_pchg_lv2
            ws ["DF" + ws_cnt] = img_fev1_pchg_lv3
            ws ["DG" + ws_cnt] = img_fev1_pchg_lv4
            ws ["DH" + ws_cnt] = img_fev1_pchg_lv5
            ws ["DI" + ws_cnt] = img_fev1_pchg_lv6
            ws ["DJ" + ws_cnt] = img_fev1_pchg_lv7
            ws ["DK" + ws_cnt] = img_fev1_pchg_lv8
            ws ["DL" + ws_cnt] = img_fev1_pchg_lv9
            ws ["DM" + ws_cnt] = img_fev1_pred
            ws ["DN" + ws_cnt] = img_fev1_pre_ppred
            ws ["DO" + ws_cnt] = img_fev1_post_ppred
            ws ["DP" + ws_cnt] = img_pc_fev1
            ws ["DQ" + ws_cnt] = img_fev1_pc
            ws ["DR" + ws_cnt] = img_fev1dfvc_ref
            ws ["DS" + ws_cnt] = img_fev1dfvc_pre
            ws ["DT" + ws_cnt] = img_fev1dfvc_post
            ws ["DU" + ws_cnt] = img_fev1dfvc_pred
            ws ["DV" + ws_cnt] = img_fev1dfvc_tri1
            ws ["DW" + ws_cnt] = img_fev1dfvc_tri2
            ws ["DX" + ws_cnt] = img_fev1dfvc_tri3
            ws ["DY" + ws_cnt] = img_fev1dfvc_tri4
            ws ["DZ" + ws_cnt] = img_fev1dfvc_tri5
            ws ["EA" + ws_cnt] = img_fev1dfvc_tri6
            ws ["EB" + ws_cnt] = img_fev1dfvc_tri7
            ws ["EC" + ws_cnt] = img_fev1dfvc_tri8
            ws ["ED" + ws_cnt] = img_fev3dfvc_ref
            ws ["EE" + ws_cnt] = img_fev3dfvc_pre
            ws ["EF" + ws_cnt] = img_fev3dfvc_post
            ws ["EG" + ws_cnt] = img_fev3dfvc_pref_pre
            ws ["EH" + ws_cnt] = img_fev3dfvc_pref_post
            ws ["EI" + ws_cnt] = img_fev3dfvc_pchg
            ws ["EJ" + ws_cnt] = img_fef25_75_dose_lv1
            ws ["EK" + ws_cnt] = img_fef25_75_dose_lv2
            ws ["EL" + ws_cnt] = img_fef25_75_dose_lv3
            ws ["EM" + ws_cnt] = img_fef25_75_dose_lv4
            ws ["EN" + ws_cnt] = img_fef25_75_dose_lv5
            ws ["EO" + ws_cnt] = img_fef25_75_dose_lv6
            ws ["EP" + ws_cnt] = img_fef25_75_dose_lv7
            ws ["EQ" + ws_cnt] = img_fef25_75_dose_lv8
            ws ["ER" + ws_cnt] = img_fef25_75_dose_lv9
            ws ["ES" + ws_cnt] = img_fef25_75_ref
            ws ["ET" + ws_cnt] = img_fef25_75_pre
            ws ["EU" + ws_cnt] = img_fef25_75_post
            ws ["EV" + ws_cnt] = img_fef25_75_lv1
            ws ["EW" + ws_cnt] = img_fef25_75_lv2
            ws ["EX" + ws_cnt] = img_fef25_75_lv3
            ws ["EY" + ws_cnt] = img_fef25_75_lv4
            ws ["EZ" + ws_cnt] = img_fef25_75_lv5
            ws ["FA" + ws_cnt] = img_fef25_75_lv6
            ws ["FB" + ws_cnt] = img_fef25_75_lv7
            ws ["FC" + ws_cnt] = img_fef25_75_lv8
            ws ["FD" + ws_cnt] = img_fef25_75_lv9
            ws ["FE" + ws_cnt] = img_fef25_75_tri1
            ws ["FF" + ws_cnt] = img_fef25_75_tri2
            ws ["FG" + ws_cnt] = img_fef25_75_tri3
            ws ["FH" + ws_cnt] = img_fef25_75_tri4
            ws ["FI" + ws_cnt] = img_fef25_75_tri5
            ws ["FJ" + ws_cnt] = img_fef25_75_tri6
            ws ["FK" + ws_cnt] = img_fef25_75_tri7
            ws ["FL" + ws_cnt] = img_fef25_75_tri8
            ws ["FM" + ws_cnt] = img_fef25_75_pref_pre
            ws ["FN" + ws_cnt] = img_fef25_75_pref_post
            ws ["FO" + ws_cnt] = img_fef25_75_pref_lv1
            ws ["FP" + ws_cnt] = img_fef25_75_pref_lv2
            ws ["FQ" + ws_cnt] = img_fef25_75_pref_lv3
            ws ["FR" + ws_cnt] = img_fef25_75_pref_lv4
            ws ["FS" + ws_cnt] = img_fef25_75_pref_lv5
            ws ["FT" + ws_cnt] = img_fef25_75_pref_lv6
            ws ["FU" + ws_cnt] = img_fef25_75_pref_lv7
            ws ["FV" + ws_cnt] = img_fef25_75_pref_lv8
            ws ["FW" + ws_cnt] = img_fef25_75_pref_lv9
            ws ["FX" + ws_cnt] = img_fef25_75_pchg
            ws ["FY" + ws_cnt] = img_fef25_75_pchg_lv1
            ws ["FZ" + ws_cnt] = img_fef25_75_pchg_lv2
            ws ["GA" + ws_cnt] = img_fef25_75_pchg_lv3
            ws ["GB" + ws_cnt] = img_fef25_75_pchg_lv4
            ws ["GC" + ws_cnt] = img_fef25_75_pchg_lv5
            ws ["GD" + ws_cnt] = img_fef25_75_pchg_lv6
            ws ["GE" + ws_cnt] = img_fef25_75_pchg_lv7
            ws ["GF" + ws_cnt] = img_fef25_75_pchg_lv8
            ws ["GG" + ws_cnt] = img_fef25_75_pchg_lv9
            ws ["GH" + ws_cnt] = img_fef25_75_pred
            ws ["GI" + ws_cnt] = img_fef25_75_pre_ppred
            ws ["GJ" + ws_cnt] = img_fef25_75_post_ppred
            ws ["GK" + ws_cnt] = img_isofef25_75_pchg
            ws ["GL" + ws_cnt] = img_isofef25_75_pred
            ws ["GM" + ws_cnt] = img_isofef25_75_pre
            ws ["GN" + ws_cnt] = img_isofef25_75_pre_ppred
            ws ["GO" + ws_cnt] = img_isofef25_75_post
            ws ["GP" + ws_cnt] = img_isofef25_75_post_ppred
            ws ["GQ" + ws_cnt] = img_fef75_85_ref
            ws ["GR" + ws_cnt] = img_fef75_85_pre
            ws ["GS" + ws_cnt] = img_fef75_85_post
            ws ["GT" + ws_cnt] = img_fef75_85_pref_pre
            ws ["GU" + ws_cnt] = img_fef75_85_pref_post
            ws ["GV" + ws_cnt] = img_fef75_85_pchg
            ws ["GW" + ws_cnt] = img_fef75_85_pred
            ws ["GX" + ws_cnt] = img_fef75_85_pre_ppred
            ws ["GY" + ws_cnt] = img_fef75_85_post_ppred
            ws ["GZ" + ws_cnt] = img_fef25_ref
            ws ["HA" + ws_cnt] = img_fef25_pre
            ws ["HB" + ws_cnt] = img_fef25_post
            ws ["HC" + ws_cnt] = img_fef25_pref_pre
            ws ["HD" + ws_cnt] = img_fef25_pref_post
            ws ["HE" + ws_cnt] = img_fef25_pchg
            ws ["HF" + ws_cnt] = img_fef50_ref
            ws ["HG" + ws_cnt] = img_fef50_pre
            ws ["HH" + ws_cnt] = img_fef50_post
            ws ["HI" + ws_cnt] = img_fef50_pref_pre
            ws ["HJ" + ws_cnt] = img_fef50_pref_post
            ws ["HK" + ws_cnt] = img_fef50_pchg
            ws ["HL" + ws_cnt] = img_fef75_ref
            ws ["HM" + ws_cnt] = img_fef75_pre
            ws ["HN" + ws_cnt] = img_fef75_post
            ws ["HO" + ws_cnt] = img_fef75_pref_pre
            ws ["HP" + ws_cnt] = img_fef75_pref_post
            ws ["HQ" + ws_cnt] = img_fef75_pchg
            ws ["HR" + ws_cnt] = img_fef200_1200_ref
            ws ["HS" + ws_cnt] = img_fef200_1200_pre
            ws ["HT" + ws_cnt] = img_fef200_1200_post
            ws ["HU" + ws_cnt] = img_fef200_1200_pref_pre
            ws ["HV" + ws_cnt] = img_fef200_1200_pref_post
            ws ["HW" + ws_cnt] = img_fef200_1200_pchg
            ws ["HX" + ws_cnt] = img_pef_dose_lv1
            ws ["HY" + ws_cnt] = img_pef_dose_lv2
            ws ["HZ" + ws_cnt] = img_pef_dose_lv3
            ws ["IA" + ws_cnt] = img_pef_dose_lv4
            ws ["IB" + ws_cnt] = img_pef_dose_lv5
            ws ["IC" + ws_cnt] = img_pef_dose_lv6
            ws ["ID" + ws_cnt] = img_pef_dose_lv7
            ws ["IE" + ws_cnt] = img_pef_dose_lv8
            ws ["IF" + ws_cnt] = img_pef_dose_lv9
            ws ["IG" + ws_cnt] = img_pef_ref
            ws ["IH" + ws_cnt] = img_pef_pre
            ws ["II" + ws_cnt] = img_pef_post
            ws ["IJ" + ws_cnt] = img_pef_lv1
            ws ["IK" + ws_cnt] = img_pef_lv2
            ws ["IL" + ws_cnt] = img_pef_lv3
            ws ["IM" + ws_cnt] = img_pef_lv4
            ws ["IN" + ws_cnt] = img_pef_lv5
            ws ["IO" + ws_cnt] = img_pef_lv6
            ws ["IP" + ws_cnt] = img_pef_lv7
            ws ["IQ" + ws_cnt] = img_pef_lv8
            ws ["IR" + ws_cnt] = img_pef_lv9
            ws ["IS" + ws_cnt] = img_pef_tri1
            ws ["IT" + ws_cnt] = img_pef_tri2
            ws ["IU" + ws_cnt] = img_pef_tri3
            ws ["IV" + ws_cnt] = img_pef_tri4
            ws ["IW" + ws_cnt] = img_pef_tri5
            ws ["IX" + ws_cnt] = img_pef_tri6
            ws ["IY" + ws_cnt] = img_pef_tri7
            ws ["IZ" + ws_cnt] = img_pef_tri8
            ws ["JA" + ws_cnt] = img_pef_pref_pre
            ws ["JB" + ws_cnt] = img_pef_pref_post
            ws ["JC" + ws_cnt] = img_pef_pref_lv1
            ws ["JD" + ws_cnt] = img_pef_pref_lv2
            ws ["JE" + ws_cnt] = img_pef_pref_lv3
            ws ["JF" + ws_cnt] = img_pef_pref_lv4
            ws ["JG" + ws_cnt] = img_pef_pref_lv5
            ws ["JH" + ws_cnt] = img_pef_pref_lv6
            ws ["JI" + ws_cnt] = img_pef_pref_lv7
            ws ["JJ" + ws_cnt] = img_pef_pref_lv8
            ws ["JK" + ws_cnt] = img_pef_pref_lv9
            ws ["JL" + ws_cnt] = img_pef_pchg
            ws ["JM" + ws_cnt] = img_pef_pchg_lv1
            ws ["JN" + ws_cnt] = img_pef_pchg_lv2
            ws ["JO" + ws_cnt] = img_pef_pchg_lv3
            ws ["JP" + ws_cnt] = img_pef_pchg_lv4
            ws ["JQ" + ws_cnt] = img_pef_pchg_lv5
            ws ["JR" + ws_cnt] = img_pef_pchg_lv6
            ws ["JS" + ws_cnt] = img_pef_pchg_lv7
            ws ["JT" + ws_cnt] = img_pef_pchg_lv8
            ws ["JU" + ws_cnt] = img_pef_pchg_lv9
            ws ["JV" + ws_cnt] = img_pef_pred
            ws ["JW" + ws_cnt] = img_pef_pre_ppred
            ws ["JX" + ws_cnt] = img_pef_post_ppred
            ws ["JY" + ws_cnt] = img_peft_pchg
            ws ["JZ" + ws_cnt] = img_peft_pre
            ws ["KA" + ws_cnt] = img_peft_post
            ws ["KB" + ws_cnt] = img_peft_tri1
            ws ["KC" + ws_cnt] = img_peft_tri2
            ws ["KD" + ws_cnt] = img_peft_tri3
            ws ["KE" + ws_cnt] = img_peft_tri4
            ws ["KF" + ws_cnt] = img_peft_tri5
            ws ["KG" + ws_cnt] = img_peft_tri6
            ws ["KH" + ws_cnt] = img_peft_tri7
            ws ["KI" + ws_cnt] = img_peft_tri8
            ws ["KJ" + ws_cnt] = img_fet100_pchg
            ws ["KK" + ws_cnt] = img_fet100_pre
            ws ["KL" + ws_cnt] = img_fet100_post
            ws ["KM" + ws_cnt] = img_fet100_tri1
            ws ["KN" + ws_cnt] = img_fet100_tri2
            ws ["KO" + ws_cnt] = img_fet100_tri3
            ws ["KP" + ws_cnt] = img_fet100_tri4
            ws ["KQ" + ws_cnt] = img_fet100_tri5
            ws ["KR" + ws_cnt] = img_fet100_tri6
            ws ["KS" + ws_cnt] = img_fet100_tri7
            ws ["KT" + ws_cnt] = img_fet100_tri8
            ws ["KU" + ws_cnt] = img_fivc_ref
            ws ["KV" + ws_cnt] = img_fivc_pre
            ws ["KW" + ws_cnt] = img_fivc_post
            ws ["KX" + ws_cnt] = img_fivc_pref_pre
            ws ["KY" + ws_cnt] = img_fivc_pref_post
            ws ["KZ" + ws_cnt] = img_fivc_pchg
            ws ["LA" + ws_cnt] = img_fivc_pred
            ws ["LB" + ws_cnt] = img_fivc_pre_ppred
            ws ["LC" + ws_cnt] = img_fivc_post_ppred
            ws ["LD" + ws_cnt] = img_fivc_tri1
            ws ["LE" + ws_cnt] = img_fivc_tri2
            ws ["LF" + ws_cnt] = img_fivc_tri3
            ws ["LG" + ws_cnt] = img_fivc_tri4
            ws ["LH" + ws_cnt] = img_fivc_tri5
            ws ["LI" + ws_cnt] = img_fivc_tri6
            ws ["LJ" + ws_cnt] = img_fivc_tri7
            ws ["LK" + ws_cnt] = img_fivc_tri8
            ws ["LL" + ws_cnt] = img_fiv1_pchg
            ws ["LM" + ws_cnt] = img_fiv1_pre
            ws ["LN" + ws_cnt] = img_fiv1_post
            ws ["LO" + ws_cnt] = img_fefdfif50_pchg
            ws ["LP" + ws_cnt] = img_fefdfif50_pred
            ws ["LQ" + ws_cnt] = img_fefdfif50_pre
            ws ["LR" + ws_cnt] = img_fefdfif50_pre_ppred
            ws ["LS" + ws_cnt] = img_fefdfif50_post
            ws ["LT" + ws_cnt] = img_fefdfif50_post_ppred
            ws ["LU" + ws_cnt] = img_volextrap_pchg
            ws ["LV" + ws_cnt] = img_volextrap_pre
            ws ["LW" + ws_cnt] = img_volextrap_post
            ws ["LX" + ws_cnt] = img_volextrap_tri1
            ws ["LY" + ws_cnt] = img_volextrap_tri2
            ws ["LZ" + ws_cnt] = img_volextrap_tri3
            ws ["MA" + ws_cnt] = img_volextrap_tri4
            ws ["MB" + ws_cnt] = img_volextrap_tri5
            ws ["MC" + ws_cnt] = img_volextrap_tri6
            ws ["MD" + ws_cnt] = img_volextrap_tri7
            ws ["ME" + ws_cnt] = img_volextrap_tri8
            ws ["MF" + ws_cnt] = img_fvlecode_pre
            ws ["MG" + ws_cnt] = img_fvlecode_post
            ws ["MH" + ws_cnt] = img_fvlecode_tri1
            ws ["MI" + ws_cnt] = img_fvlecode_tri2
            ws ["MJ" + ws_cnt] = img_fvlecode_tri3
            ws ["MK" + ws_cnt] = img_fvlecode_tri4
            ws ["ML" + ws_cnt] = img_fvlecode_tri5
            ws ["MM" + ws_cnt] = img_fvlecode_tri6
            ws ["MN" + ws_cnt] = img_fvlecode_tri7
            ws ["MO" + ws_cnt] = img_fvlecode_tri8
            ws ["MP" + ws_cnt] = img_mvv_pred
            ws ["MQ" + ws_cnt] = img_tlc_ref
            ws ["MR" + ws_cnt] = img_tlc_pre
            ws ["MS" + ws_cnt] = img_tlc_pref_pre
            ws ["MT" + ws_cnt] = img_vc_ref
            ws ["MU" + ws_cnt] = img_vc_pre
            ws ["MV" + ws_cnt] = img_vc_pref_pre
            ws ["MW" + ws_cnt] = img_ic_ref
            ws ["MX" + ws_cnt] = img_ic_pre
            ws ["MY" + ws_cnt] = img_ic_pref_pre
            ws ["MZ" + ws_cnt] = img_frcpl_ref
            ws ["NA" + ws_cnt] = img_frcpl_pre
            ws ["NB" + ws_cnt] = img_frcpl_pref_pre
            ws ["NC" + ws_cnt] = img_erv_ref
            ws ["ND" + ws_cnt] = img_erv_pre
            ws ["NE" + ws_cnt] = img_erv_pref_pre
            ws ["NF" + ws_cnt] = img_rv_ref
            ws ["NG" + ws_cnt] = img_rv_pre
            ws ["NH" + ws_cnt] = img_rv_pref_pre
            ws ["NI" + ws_cnt] = img_rvdtlc_ref
            ws ["NJ" + ws_cnt] = img_rvdtlc_pre
            ws ["NK" + ws_cnt] = img_vtg_pre
            ws ["NL" + ws_cnt] = img_vt_pre
            ws ["NM" + ws_cnt] = img_dlco_ref
            ws ["NN" + ws_cnt] = img_dlco_pre
            ws ["NO" + ws_cnt] = img_dlco_pref_pre
            ws ["NP" + ws_cnt] = img_dladj_ref
            ws ["NQ" + ws_cnt] = img_dladj_pre
            ws ["NR" + ws_cnt] = img_dladj_pref_pre
            ws ["NS" + ws_cnt] = img_dlcodva_ref
            ws ["NT" + ws_cnt] = img_dlcodva_pre
            ws ["NU" + ws_cnt] = img_dlcodva_pref_pre
            ws ["NV" + ws_cnt] = img_dldvaadj_ref
            ws ["NW" + ws_cnt] = img_dldvaadj_pre
            ws ["NX" + ws_cnt] = img_dldvaadj_pref_pre
            ws ["NY" + ws_cnt] = img_va_pre
            ws ["NZ" + ws_cnt] = img_ivc_pre
            ws ["OA" + ws_cnt] = img_dlcoecode_pre
            ws ["OB" + ws_cnt] = img_rawtotal_pre
            ws ["OC" + ws_cnt] = img_rawinsp_pre
            ws ["OD" + ws_cnt] = img_rawexp_pre
            ws ["OE" + ws_cnt] = img_raw_ref
            ws ["OF" + ws_cnt] = img_raw_pre
            ws ["OG" + ws_cnt] = img_raw_pref_pre
            ws ["OH" + ws_cnt] = img_gaw_ref
            ws ["OI" + ws_cnt] = img_gaw_pre
            ws ["OJ" + ws_cnt] = img_gaw_pref_pre
            ws ["OK" + ws_cnt] = img_sraw_ref
            ws ["OL" + ws_cnt] = img_sraw_pre
            ws ["OM" + ws_cnt] = img_sraw_pref_pre
            ws ["ON" + ws_cnt] = img_sgaw_ref
            ws ["OO" + ws_cnt] = img_sgaw_pre
            ws ["OP" + ws_cnt] = img_sgaw_pref_pre
            ws ["OQ" + ws_cnt] = img_rawvtg_pre
            ws ["OR" + ws_cnt] = img_rawf_pre
            wb.save(tmp_dir + "/OOMII.xlsx")
            
            # create new image
            cv2.imwrite(new_path + "/" + img_type + "_" + img_name, new_image_files)
            new_path = tmp_dir

        self.label3.setText("Total: " + str(ori_image_files_counter) + "\n\n"
                          + "type01: " + str(type01_img_cnt) + "\n"
                          + "type02: " + str(type02_img_cnt) + "\n"
                          + "type03: " + str(type03_img_cnt) + "\n"
                          + "type04: " + str(type04_img_cnt) + "\n"
                          + "type05: " + str(type05_img_cnt) + "\n"
                          + "type06: " + str(type06_img_cnt) + "\n"
			              + "type07: " + str(type07_img_cnt) + "\n"
			              + "type08: " + str(type08_img_cnt) + "\n"
                          + "typeUK: " + str(typeUK_img_cnt) + "\n")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    ex.show()
    sys.exit(app.exec_())


# In[ ]:
